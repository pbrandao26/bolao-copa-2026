#!/usr/bin/env python3
"""
============================================================================
VALIDADOR T495 v2 — Bolão Copa 2026 Turim MFO
============================================================================
Mostra, para cada participante afetado pela T495 errada, EXATAMENTE quais
jogos da R32 mudariam — com nomes de seleções — e em quem ele pôs o X.

O que mudou da v1
-----------------
Além da chave e do "1A vs 3H → 3C", agora traduz cada confronto para os
NOMES reais das seleções (na visão dos palpites do participante) e indica
em qual seleção ele apostou que avançaria (o X da linha 6).

Estrutura usada nas planilhas dos participantes
-----------------------------------------------
• Calc!B88        → chave T495 (8 letras).
• Calc linhas 53–64 → classificação por grupo (col A=grupo; col D=3º colocado).
• Aba 'Apostas - Mata-Mata':
    linha 3 → rótulos M73..M88
    linha 5 → nomes das seleções (R32 nas colunas B..AG, em pares)
    linha 6 → "X" abaixo da seleção que o participante crê que avança
  Jogo j (0..15): coluna esquerda = 2 + j*2, coluna direita = 3 + j*2.

Mapa posição-T495 → jogo R32 (derivado do bracket FIFA do app):
    1A→M79  1B→M85  1D→M81  1E→M74
    1G→M82  1I→M77  1K→M87  1L→M80
Em cada um desses jogos, um lado é o líder do grupo (não muda) e o outro é
o 3º colocado (o lado que a T495 errada definiu errado).

Uso
---
    python verificador.py . matchups_r32_copa_2026.xlsx
============================================================================
"""

import sys
import glob
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────
PASTA_PADRAO     = "."
PLANILHA_CORRETA = "matchups_r32_copa_2026.xlsx"
ABA_CORRETA      = "R32 Matchups Certo"

ABA_CALC      = "Calc"
CELULA_CHAVE  = "B88"
CALC_GRP_INI  = 53          # linha inicial da classificação por grupo
CALC_GRP_FIM  = 64          # linha final
CALC_3RD_COL  = 4           # coluna D = 3º colocado

ABA_MM     = "Apostas - Mata-Mata"
MM_ROW_SEL = 5              # linha com nomes das seleções
MM_ROW_X   = 6              # linha com o X

COLUNAS_CONFRONTO = ["1A", "1B", "1D", "1E", "1G", "1I", "1K", "1L"]

# Posição na T495 (0..7) → rótulo do jogo Mxx onde aquele 3º colocado entra.
POS_TO_GAME = {
    0: "M79",   # 1A vs 3?
    1: "M85",   # 1B vs 3?
    2: "M81",   # 1D vs 3?
    3: "M74",   # 1E vs 3?
    4: "M82",   # 1G vs 3?
    5: "M77",   # 1I vs 3?
    6: "M87",   # 1K vs 3?
    7: "M80",   # 1L vs 3?
}
# Líder de grupo de cada posição (o lado que NÃO muda).
POS_TO_LEADER = {0: "A", 1: "B", 2: "D", 3: "E", 4: "G", 5: "I", 6: "K", 7: "L"}


# ── CARREGAR TABELA CORRETA ────────────────────────────────────────────
def carregar_t495_correta(caminho, aba):
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb[aba]
    tabela = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        chave = str(row[0]).strip()
        if len(chave) == 8:
            tabela[chave] = [str(v).strip() if v is not None else "?" for v in row[1:9]]
    return tabela


def carregar_t495_da_planilha(wb):
    nome = next((s for s in wb.sheetnames if s.strip().upper() == "T495"), None)
    if not nome:
        return {}
    ws = wb[nome]
    tabela = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        chave = str(row[0]).strip()
        if len(chave) == 8:
            tabela[chave] = [str(v).strip() if v is not None else "?" for v in row[1:9]]
    return tabela


def ler_terceiros_por_grupo(wb):
    """Calc linhas 53–64: {grupo: nome do 3º colocado (palpite)}."""
    ws = wb[ABA_CALC]
    thirds = {}
    for r in range(CALC_GRP_INI, CALC_GRP_FIM + 1):
        g = ws.cell(row=r, column=1).value
        t3 = ws.cell(row=r, column=CALC_3RD_COL).value
        if g:
            thirds[str(g).strip()] = (str(t3).strip() if t3 else "?")
    return thirds


def ler_confrontos_mm(wb):
    """Lê a aba Apostas - Mata-Mata → {Mxx: {'esq':(time,tem_x),'dir':(time,tem_x)}}."""
    ws = wb[ABA_MM]
    jogos = {}
    for j in range(16):
        c1 = 2 + j * 2
        c2 = c1 + 1
        rot = ws.cell(row=3, column=c1).value  # rótulo Mxx fica na col esquerda
        rot = str(rot).strip() if rot else f"J{j}"
        t1 = ws.cell(row=MM_ROW_SEL, column=c1).value
        t2 = ws.cell(row=MM_ROW_SEL, column=c2).value
        x1 = ws.cell(row=MM_ROW_X, column=c1).value
        x2 = ws.cell(row=MM_ROW_X, column=c2).value
        has_x1 = bool(x1) and str(x1).strip().upper() == "X"
        has_x2 = bool(x2) and str(x2).strip().upper() == "X"
        jogos[rot] = {
            "esq": (str(t1).strip() if t1 else "?", has_x1),
            "dir": (str(t2).strip() if t2 else "?", has_x2),
        }
    return jogos


# ── ANÁLISE DE UM PARTICIPANTE ─────────────────────────────────────────
def analisar_participante(caminho, t495_correta):
    nome_arq = os.path.basename(caminho)
    nome = (nome_arq
            .replace("Bolao_Copa2026_TurimMFO_", "")
            .replace(".xlsx", "")
            .replace("_", " ")
            .strip())

    try:
        wb = load_workbook(caminho, data_only=True)
    except Exception as e:
        return {"nome": nome, "status": "ERRO_LEITURA", "detalhe": str(e)}

    # ignora planilhas que não são de participante (sem aba Calc)
    if ABA_CALC not in wb.sheetnames:
        return {"nome": nome, "status": "NAO_PARTICIPANTE",
                "detalhe": "sem aba Calc (consolidada/tabela auxiliar)"}

    chave = wb[ABA_CALC][CELULA_CHAVE].value
    if chave is None:
        return {"nome": nome, "status": "CHAVE_VAZIA",
                "detalhe": "Calc!B88 vazia (abra e salve no Excel para recalcular)"}
    chave = str(chave).strip()
    if len(chave) != 8:
        return {"nome": nome, "status": "CHAVE_INVALIDA",
                "detalhe": f"chave='{chave}'", "chave": chave}

    t495_errada = carregar_t495_da_planilha(wb)
    linha_errada = t495_errada.get(chave)
    linha_certa  = t495_correta.get(chave)
    if linha_certa is None:
        return {"nome": nome, "status": "CHAVE_SEM_CORRETA", "chave": chave}
    if linha_errada == linha_certa:
        return {"nome": nome, "status": "OK", "chave": chave}

    # Há diferença → traduz para nomes
    thirds   = ler_terceiros_por_grupo(wb)   # {grupo: 3º colocado}
    confronto = ler_confrontos_mm(wb)        # {Mxx: esq/dir (time, x)}

    mudancas = []
    for pos, (col, ve, vc) in enumerate(zip(COLUNAS_CONFRONTO,
                                            linha_errada or ["?"] * 8,
                                            linha_certa)):
        if ve == vc:
            continue
        game = POS_TO_GAME[pos]
        leader_grp = POS_TO_LEADER[pos]
        # nomes: ve/vc são como '3H' → grupo H
        grp_errado = ve[-1] if ve and ve != "?" else "?"
        grp_certo  = vc[-1] if vc and vc != "?" else "?"
        time_errado = thirds.get(grp_errado, "?")
        time_certo  = thirds.get(grp_certo, "?")

        # o jogo na planilha: descobrir qual lado é o líder e qual é o 3º
        jg = confronto.get(game, {})
        esq_team, esq_x = jg.get("esq", ("?", False))
        dir_team, dir_x = jg.get("dir", ("?", False))
        lider_team = thirds_leader(wb, leader_grp)  # 1º colocado do grupo do líder

        # qual lado tem o time_errado (3º colocado errado que apareceu na planilha)?
        if esq_team == time_errado:
            lado_3o, time_3o_x = "esq", esq_x
            lado_lider, lider_x = "dir", dir_x
            lider_team = dir_team
        elif dir_team == time_errado:
            lado_3o, time_3o_x = "dir", dir_x
            lado_lider, lider_x = "esq", esq_x
            lider_team = esq_team
        else:
            # fallback: não casou pelo nome (ex. palpite incompleto)
            lado_3o, time_3o_x = "?", False
            lider_x = False

        # em quem o participante pôs o X nesse jogo?
        if lider_x:
            x_em = f"{lider_team} (líder grupo {leader_grp})"
        elif time_3o_x:
            x_em = f"{time_errado} (3º — adversário ERRADO)"
        else:
            x_em = "— (sem X / fase não preenchida)"

        mudancas.append({
            "game": game,
            "leader_grp": leader_grp,
            "lider_team": lider_team,
            "time_errado": time_errado, "grp_errado": grp_errado,
            "time_certo": time_certo,  "grp_certo": grp_certo,
            "x_em": x_em,
            "x_no_3o_errado": time_3o_x,
        })

    return {"nome": nome, "status": "AFETADO", "chave": chave, "mudancas": mudancas}


def thirds_leader(wb, grupo):
    """1º colocado do grupo (Calc 53–64, col B)."""
    ws = wb[ABA_CALC]
    for r in range(CALC_GRP_INI, CALC_GRP_FIM + 1):
        if str(ws.cell(row=r, column=1).value).strip() == grupo:
            v = ws.cell(row=r, column=2).value
            return str(v).strip() if v else "?"
    return "?"


# ── MAIN ────────────────────────────────────────────────────────────────
def main():
    pasta = sys.argv[1] if len(sys.argv) > 1 else PASTA_PADRAO
    correta_path = sys.argv[2] if len(sys.argv) > 2 else PLANILHA_CORRETA

    print("=" * 76)
    print("VALIDADOR T495 v2 — jogos da R32 que mudariam, por seleção e com o X")
    print("=" * 76)
    print(f"Pasta: {pasta}  |  Tabela correta: {correta_path} → '{ABA_CORRETA}'\n")

    t495_correta = carregar_t495_correta(correta_path, ABA_CORRETA)
    print(f"Tabela correta: {len(t495_correta)} chaves.\n")

    arquivos = sorted(glob.glob(os.path.join(pasta, "*.xlsx")))
    arquivos = [a for a in arquivos if not os.path.basename(a).startswith("~$")]
    print(f"{len(arquivos)} planilha(s) encontradas.\n")

    afetados, ok, problemas = [], [], []
    for caminho in arquivos:
        r = analisar_participante(caminho, t495_correta)
        if r["status"] == "OK":
            ok.append(r)
        elif r["status"] == "AFETADO":
            afetados.append(r)
        elif r["status"] == "NAO_PARTICIPANTE":
            pass  # silencioso
        else:
            problemas.append(r)

    print("─" * 76)
    print(f"RESUMO: {len(afetados)} afetado(s) · {len(ok)} ok · "
          f"{len(problemas)} a verificar")
    print("─" * 76)

    if afetados:
        print("\n🔴 PARTICIPANTES AFETADOS — jogos da R32 que mudariam:\n")
        for r in afetados:
            print(f"╔═ {r['nome']}  (chave {r['chave']})")
            for mc in r["mudancas"]:
                print(f"║  {mc['game']} — líder do grupo {mc['leader_grp']} "
                      f"({mc['lider_team']}):")
                print(f"║      adversário ERRADO : {mc['time_errado']} "
                      f"(3º do grupo {mc['grp_errado']})")
                print(f"║      adversário CERTO  : {mc['time_certo']} "
                      f"(3º do grupo {mc['grp_certo']})")
                print(f"║      apostou que avança: {mc['x_em']}")
                if mc["x_no_3o_errado"]:
                    print(f"║      ⚠️  ele marcou o ADVERSÁRIO ERRADO para avançar — "
                          f"esse X fica sem sentido com o time certo!")
                print("║")
            print("╚" + "═" * 60 + "\n")
    else:
        print("\n✅ Ninguém afetado.\n")

    if problemas:
        print("\n⚠️  A VERIFICAR:\n")
        for r in problemas:
            print(f"  ● {r['nome']:30s} [{r['status']}] {r.get('detalhe','')}")


if __name__ == "__main__":
    main()