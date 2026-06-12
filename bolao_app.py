#!/usr/bin/env python3
"""
Bolão Copa do Mundo 2026 — Turim MFO
Dashboard v5 · Paleta Turim/Tori · Sidebar nativa
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
import os, glob, re, base64, hashlib, pickle, unicodedata
from concurrent.futures import ThreadPoolExecutor
from collections import OrderedDict, Counter
from datetime import date
from pathlib import Path
from dotenv import load_dotenv
import numpy as np

load_dotenv()

try:
    _PWD = st.secrets["APP_PASSWORD"]
except Exception:
    _PWD = os.getenv("APP_PASSWORD")

# ══════════════════════════════════════════════════════════════════════
# PAGE CONFIG — sidebar nativa funciona, 3 pontinhos aparecem
# ══════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Bolão Copa 2026 · Turim MFO",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════
# PALETA TURIM (extraída do PPTX corporativo)
# Primary: #0D2B40 (navy escuro), #123A56 (navy médio)
# Accent:  #D6B864 (gold), #DC884A (laranja), #B2584E (vermelho)
# Tori:    #0D8587 (teal)
# Text:    #4D4D4F / #FFFFFF
# Fontes:  Effra, Gotham (fallback: Inter)
# ══════════════════════════════════════════════════════════════════════

# CSS — only styles custom HTML divs, inherits text for dark/light compat
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,300;0,400;0,600;0,700;0,800;0,900;1,400&display=swap');

/* Fonte global sem quebrar sidebar */
html, body, [class*="css"] {
    font-family: 'Inter', 'Effra', 'Gotham Book', sans-serif !important;
}

/* Remove padding top excessivo */
.block-container { padding-top: 1.2rem !important; max-width: 1400px; }

/* ── Hero banner */
.hero {
    background: linear-gradient(135deg, #0D2B40 0%, #123A56 50%, #0D8587 100%);
    border-radius: 16px; padding: 28px 40px; margin-bottom: 20px;
    position: relative; overflow: hidden;
    box-shadow: 0 6px 28px rgba(13,43,64,.35);
}
.hero::after {
    content: ""; position: absolute; right: -20px; top: -20px;
    width: 200px; height: 200px; border-radius: 50%;
    background: rgba(214,184,100,.12);
}
.hero-title {
    font-size: 1.85rem; font-weight: 900; color: #FFFFFF !important;
    margin: 0; letter-spacing: -.5px;
}
.hero-sub { font-size: .88rem; color: rgba(255,255,255,.72) !important; margin-top: 5px; }
.hero-badge {
    display: inline-block; background: rgba(214,184,100,.25);
    border: 1px solid rgba(214,184,100,.5); border-radius: 20px;
    padding: 3px 12px; font-size: .75rem; color: #D6B864 !important; margin-top: 10px;
}

/* ── Metric mini-card */
.mc {
    border: 1px solid rgba(128,128,128,.15);
    border-radius: 12px; padding: 14px 16px; text-align: center;
}
.mc-v { font-size: 1.7rem; font-weight: 900; color: #D6B864 !important; line-height: 1; }
.mc-l { font-size: .63rem; color: inherit; opacity: .6; text-transform: uppercase;
         letter-spacing: 1px; margin-top: 4px; }

/* ── Ranking card */
.rc {
    border-radius: 11px; padding: 12px 16px; margin-bottom: 6px;
    display: flex; align-items: center; gap: 12px;
    border: 1px solid rgba(128,128,128,.15);
}
.rc1 { border-left: 4px solid #D6B864; }
.rc2 { border-left: 4px solid #7F7F7F; }
.rc3 { border-left: 4px solid #DC884A; }
.rcN { border-left: 4px solid #123A56; }
.rc-name { font-size: .94rem; font-weight: 700; flex: 1; }
.rc-sub  { font-size: .67rem; opacity: .6; margin-top: 2px; }
.rc-pts  { font-size: 1.55rem; font-weight: 900; color: #D6B864 !important; }
.rc-pl   { font-size: .62rem; color: #5C5F62; text-align: right; }
.bar-bg  { height: 3px; border-radius: 3px; background: rgba(18,58,86,.12); margin-top: 5px; }
.bar-fg  { height: 3px; border-radius: 3px;
           background: linear-gradient(90deg, #123A56, #D6B864); }

/* ── Group table */
.gb {
    border: 1px solid rgba(128,128,128,.12); border-radius: 11px;
    padding: 11px 12px; margin-bottom: 9px;
}
.gb-hdr { font-weight: 800; font-size: .88rem; margin-bottom: 6px; }
.gt { width: 100%; border-collapse: collapse; font-size: .76rem; }
.gt th { font-weight: 700; text-align: center; padding: 4px 4px;
          background: #0D2B40; color: #FFFFFF !important;
          border-bottom: 2px solid rgba(214,184,100,.3); }
.gt td { padding: 4px 4px; text-align: center; color: inherit; }
.gt td.nm { text-align: left; font-weight: 600; white-space: nowrap; }
.row-q1 { background: rgba(13,133,135,.14); }
.row-q2 { background: rgba(13,133,135,.07); }
.row-q3 { background: rgba(220,136,74,.12); }
.row-q4 { background: rgba(178,88,78,.10); }
.dot { display: inline-block; width: 7px; height: 7px; border-radius: 50%;
       margin-right: 4px; vertical-align: middle; }

/* ── Score badges */
.b5 { background: #14532D; color: #86EFAC; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.b3 { background: #0D4040; color: #5EEAD4; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.b2 { background: #7C2D12; color: #FED7AA; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.b0 { background: #7F1D1D; color: #FECACA; border-radius: 4px; padding: 1px 7px; font-size: .7rem; font-weight: 700; }
.bN { border-radius: 4px; padding: 1px 7px; font-size: .7rem; opacity: .5; }

/* ── Match row (group detail) */
.mr {
    border: 1px solid rgba(128,128,128,.10); border-radius: 8px;
    padding: 6px 10px; margin-bottom: 3px;
    display: flex; align-items: center; gap: 8px; font-size: .78rem;
}
.mr-t { flex: 1; font-weight: 600; }
.mr-s { font-size: .73rem; color: #5C5F62; white-space: nowrap; }

/* ── MM table */
.mm-wrap { overflow-x: auto; }
.mm-tbl { border-collapse: collapse; font-size: .78rem; width: 100%; min-width: 500px; }
.mm-tbl th {
    background: #0D2B40; color: #FFFFFF !important;
    padding: 7px 9px; text-align: center; border: 1px solid rgba(255,255,255,.1);
    white-space: nowrap; font-weight: 700;
}
.mm-tbl th.nm { text-align: left; min-width: 110px; }
.mm-tbl td { padding: 6px 9px; text-align: center; white-space: nowrap;
             color: inherit; border: 1px solid rgba(128,128,128,.10); }
.mm-tbl td.nm { text-align: left; font-weight: 600; }
.mm-tbl tr:nth-child(even) td { background: rgba(18,58,86,.04); }
.mm-sub { font-size: .66rem; opacity: .6; line-height: 1.3; }

/* ── MM detail row */
.mmr {
    border: 1px solid rgba(128,128,128,.10); border-radius: 8px;
    padding: 7px 12px; margin-bottom: 4px; font-size: .80rem;
}
.mmr-id { font-weight: 700; color: #0D8587 !important; display: inline-block; min-width: 50px; }

/* ── Bonus card */
.bc {
    border: 1px solid rgba(128,128,128,.12); border-radius: 11px;
    padding: 16px; text-align: center;
}
.bc-lbl { font-size: .63rem; color: #5C5F62 !important; text-transform: uppercase;
           letter-spacing: 1.1px; margin-bottom: 5px; }
.bc-ico  { font-size: 1.5rem; }
.bc-bet  { font-size: .88rem; font-weight: 600; margin: 4px 0 2px; }
.bc-real { font-size: .76rem; opacity: .6; }
.bc-pts  { font-size: 1.35rem; font-weight: 900; color: #D6B864 !important; margin-top: 6px; }

/* ── Section header */
.sh {
    font-size: 1.02rem; font-weight: 800;
    border-bottom: 2px solid #0D2B40; padding-bottom: 5px; margin: 14px 0 10px;
}

/* ── Sidebar logo container */
.sb-logo { text-align: center; padding: 8px 0 4px; }
.sb-divider { border: none; border-top: 1px solid rgba(18,58,86,.15); margin: 8px 0; }

/* ── Mascote strip */
.mascote-strip {
    background: linear-gradient(135deg, #0D2B40, #123A56);
    border-radius: 12px; padding: 10px; text-align: center; margin-bottom: 14px;
}

/* ── Seleção pills (Mata-Mata) */
.pill { display: inline-block; margin-right: 16px; margin-bottom: 4px; white-space: nowrap; }
.pill-real { color: inherit; }
.pill-hit  { color: #22C55E; font-weight: 700; }
.pill-miss { opacity: .55; }
.pill-wait { opacity: .7; }
            
button[data-testid="stNumberInputStepUp"],
button[data-testid="stNumberInputStepDown"] { display: none; }            
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# HELPERS — logo como base64
# ══════════════════════════════════════════════════════════════════════
def img_to_b64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""

# Paths dos assets (ao lado do script ou em subpasta assets/)
SCRIPT_DIR = Path(__file__).parent

LOGOS_DIR = SCRIPT_DIR / "logos"
APOSTAS_DIR = SCRIPT_DIR / "apostas"
GABARITO_DIR = SCRIPT_DIR / "gabarito"
CONSOLIDADA_PATH = APOSTAS_DIR / "Bolao_Copa2026_TurimMFO - Consolidada.xlsx"

def find_asset(name):
    for p in [LOGOS_DIR / name, SCRIPT_DIR / name]:
        if p.exists():
            return str(p)
    return ""

LOGO_TURIM_BRANCA = find_asset("logo_turim_branca.png")
LOGO_TURIM_AZUL   = find_asset("logo_turim_azul.png")
LOGO_TORI         = find_asset("logo_tori.png")
MASCOTES          = find_asset("mascotes.png")
FRONT_PAGE        = find_asset("front_page.png")
FAVICON           = find_asset("favicon.png")

# ── Feature flags
MOSTRAR_SIMULACAO = True   # True → mostra aba 🔮 Simulação; False → sistema idêntico ao atual

# Lista de (jogador, pais) para simulação de artilheiro — preencha antes da Copa
JOGADORES_ARTILHEIRO: list[tuple[str, str]] = [
    ('Mohammed Al-Owais', 'Arabia Saudita'),
    ('Nawaf Al-Aqidi', 'Arabia Saudita'),
    ('Ahmed Al-Kassar', 'Arabia Saudita'),
    ('Ali Al-Bulayhi', 'Arabia Saudita'),
    ('Hassan Kadesh', 'Arabia Saudita'),
    ('Yasir Al-Shahrani', 'Arabia Saudita'),
    ('Nawaf Boushal', 'Arabia Saudita'),
    ('Sultan Al-Ghannam', 'Arabia Saudita'),
    ('Nasser Al-Dawsari', 'Arabia Saudita'),
    ('Musab Al-Juwayr', 'Arabia Saudita'),
    ('Faisal Al-Ghamdi', 'Arabia Saudita'),
    ('Ayman Yahya', 'Arabia Saudita'),
    ('Abdulrahman Ghareeb', 'Arabia Saudita'),
    ('Firas Al-Buraikan', 'Arabia Saudita'),
    ('Abdullah Al-Hamdan', 'Arabia Saudita'),
    ('Saleh Al-Shehri', 'Arabia Saudita'),
    ('Saud Abdulhamid', 'Arabia Saudita'),
    ('Mohammed Kanno', 'Arabia Saudita'),
    ('Salem Al-Dawsari', 'Arabia Saudita'),
    ('Ahmed Al-Ghamdi', 'Arabia Saudita'),
    ('Mat Ryan', 'Australia'),
    ('Joe Gauci', 'Australia'),
    ('Patrick Beach', 'Australia'),
    ('Harry Souttar', 'Australia'),
    ('Kye Rowles', 'Australia'),
    ('Alessandro Circati', 'Australia'),
    ('Cameron Burgess', 'Australia'),
    ('Aziz Behich', 'Australia'),
    ('Jordan Bos', 'Australia'),
    ('Nathaniel Atkinson', 'Australia'),
    ('Jackson Irvine', 'Australia'),
    ("Aiden O'Neill", 'Australia'),
    ('Riley McGree', 'Australia'),
    ('Connor Metcalfe', 'Australia'),
    ('Martin Boyle', 'Australia'),
    ('Craig Goodwin', 'Australia'),
    ('Kusini Yengi', 'Australia'),
    ('Brandon Borrello', 'Australia'),
    ('Daniel Bennie', 'Australia'),
    ('Raphael Borges Rodrigues', 'Australia'),
    ('Awer Mabil', 'Australia'),
    ('Jalal Hassan', 'Iraque'),
    ('Ahmed Basil', 'Iraque'),
    ('Kumel Al-Rkabe', 'Iraque'),
    ('Ahmed Yahya', 'Iraque'),
    ('Rebin Sulaka', 'Iraque'),
    ('Frans Putros', 'Iraque'),
    ('Hussein Ali', 'Iraque'),
    ('Zaid Tahseen', 'Iraque'),
    ('Ali Jasim', 'Iraque'),
    ('Ibrahim Bayesh', 'Iraque'),
    ('Zidane Iqbal', 'Iraque'),
    ('Amir Al-Ammari', 'Iraque'),
    ('Osama Rashid', 'Iraque'),
    ('Youssef Amyn', 'Iraque'),
    ('Mohanad Ali', 'Iraque'),
    ('Aymen Hussein', 'Iraque'),
    ('Ali Al-Hamadi', 'Iraque'),
    ('Marko Farji', 'Iraque'),
    ('Peter Gwargis', 'Iraque'),
    ('Hassan Abdulkarim', 'Iraque'),
    ('Alireza Beiranvand', 'Ira'),
    ('Payam Niazmand', 'Ira'),
    ('Hossein Hosseini', 'Ira'),
    ('Hossein Kanaanizadegan', 'Ira'),
    ('Shoja Khalilzadeh', 'Ira'),
    ('Ramin Rezaeian', 'Ira'),
    ("Ali Ne'mati", 'Ira'),
    ('Milad Mohammadi', 'Ira'),
    ('Roozbeh Cheshmi', 'Ira'),
    ('Saman Ghoddos', 'Ira'),
    ('Saeid Ezatolahi', 'Ira'),
    ('Alireza Jahanbakhsh', 'Ira'),
    ('Mehdi Ghayedi', 'Ira'),
    ('Mehdi Taremi', 'Ira'),
    ('Mohammad Mohebi', 'Ira'),
    ('Mohammad Ghorbani', 'Ira'),
    ('Sardar Azmoun', 'Ira'),
    ('Allahyar Sayyadmanesh', 'Ira'),
    ('Omid Noorafkan', 'Ira'),
    ('Mehdi Torabi', 'Ira'),
    ('Ehsan Hajsafi', 'Ira'),
    ('Zion Suzuki', 'Japao'),
    ('Keisuke Osako', 'Japao'),
    ('Tomoki Hayakawa', 'Japao'),
    ('Ko Itakura', 'Japao'),
    ('Hiroki Ito', 'Japao'),
    ('Shogo Taniguchi', 'Japao'),
    ('Ayumu Seko', 'Japao'),
    ('Yukinari Sugawara', 'Japao'),
    ('Wataru Endo', 'Japao'),
    ('Hidemasa Morita', 'Japao'),
    ('Reo Hatate', 'Japao'),
    ('Takefusa Kubo', 'Japao'),
    ('Kaoru Mitoma', 'Japao'),
    ('Ritsu Doan', 'Japao'),
    ('Daichi Kamada', 'Japao'),
    ('Junya Ito', 'Japao'),
    ('Ayase Ueda', 'Japao'),
    ('Shuto Machino', 'Japao'),
    ('Junnosuke Suzuki', 'Japao'),
    ('Joel Chima Fujita', 'Japao'),
    ('Takehiro Tomiyasu', 'Japao'),
    ('Yazeed Abu Laila', 'Jordania'),
    ('Abdullah Nasib', 'Jordania'),
    ('Yazan Al-Arab', 'Jordania'),
    ('Noor Al-Rawabdeh', 'Jordania'),
    ('Nizar Al-Rashdan', 'Jordania'),
    ('Ali Olwan', 'Jordania'),
    ('Mousa Al-Tamari', 'Jordania'),
    ('Yazan Al-Naimat', 'Jordania'),
    ('Mahmoud Al-Mardi', 'Jordania'),
    ('Saleh Ratib', 'Jordania'),
    ('Mohammad Abu Hasheesh', 'Jordania'),
    ('Anas Bani Yaseen', 'Jordania'),
    ('Amer Jamous', 'Jordania'),
    ('Ibrahim Sabra', 'Jordania'),
    ('Meshaal Barsham', 'Qatar'),
    ('Mahmoud Abunada', 'Qatar'),
    ('Salah Zakaria', 'Qatar'),
    ('Tarek Salman', 'Qatar'),
    ('Pedro Miguel', 'Qatar'),
    ('Boualem Khoukhi', 'Qatar'),
    ('Lucas Mendes', 'Qatar'),
    ('Homam Al-Amin', 'Qatar'),
    ('Karim Boudiaf', 'Qatar'),
    ('Assim Madibo', 'Qatar'),
    ('Mohammed Waad', 'Qatar'),
    ('Ahmed Alaaeldin', 'Qatar'),
    ('Akram Afif', 'Qatar'),
    ('Almoez Ali', 'Qatar'),
    ('Mohammed Manai', 'Qatar'),
    ('Mustapha Tariq', 'Qatar'),
    ('Tahsin Mohammed Jamshid', 'Qatar'),
    ('Sebastian Soria', 'Qatar'),
    ('Jo Hyeon-woo', 'Coreia do Sul'),
    ('Kim Seung-gyu', 'Coreia do Sul'),
    ('Kim Min-jae', 'Coreia do Sul'),
    ('Kim Jin-su', 'Coreia do Sul'),
    ('Seol Young-woo', 'Coreia do Sul'),
    ('Kim Young-gwon', 'Coreia do Sul'),
    ('Hwang In-beom', 'Coreia do Sul'),
    ('Lee Jae-sung', 'Coreia do Sul'),
    ('Lee Kang-in', 'Coreia do Sul'),
    ('Hwang Hee-chan', 'Coreia do Sul'),
    ('Son Heung-min', 'Coreia do Sul'),
    ('Cho Gue-sung', 'Coreia do Sul'),
    ('Oh Hyeon-gyu', 'Coreia do Sul'),
    ('Bae Jun-ho', 'Coreia do Sul'),
    ('Yang Min-hyeok', 'Coreia do Sul'),
    ('Jeong Sang-bin', 'Coreia do Sul'),
    ('Hong Hyun-seok', 'Coreia do Sul'),
    ('Utkir Yusupov', 'Uzbequistao'),
    ('Abduvohid Nematov', 'Uzbequistao'),
    ('Abdukodir Khusanov', 'Uzbequistao'),
    ('Rustam Ashurmatov', 'Uzbequistao'),
    ('Farrukh Sayfiev', 'Uzbequistao'),
    ('Odiljon Hamrobekov', 'Uzbequistao'),
    ('Abbosbek Fayzullaev', 'Uzbequistao'),
    ('Jasurbek Jaloliddinov', 'Uzbequistao'),
    ('Jaloliddin Masharipov', 'Uzbequistao'),
    ('Oston Urunov', 'Uzbequistao'),
    ('Eldor Shomurodov', 'Uzbequistao'),
    ('Umarali Rakhmonaliev', 'Uzbequistao'),
    ("Muhammadali O'rinboyev", 'Uzbequistao'),
    ('Sardor Rashidov', 'Uzbequistao'),
    ('Bobur Abdukholikov', 'Uzbequistao'),
    ('Anthony Mandrea', 'Argelia'),
    ('Alexandre Oukidja', 'Argelia'),
    ('Rayan Aït-Nouri', 'Argelia'),
    ('Ramy Bensebaini', 'Argelia'),
    ('Aïssa Mandi', 'Argelia'),
    ('Zineddine Belaïd', 'Argelia'),
    ('Ismaël Bennacer', 'Argelia'),
    ('Houssem Aouar', 'Argelia'),
    ('Farès Chaïbi', 'Argelia'),
    ('Riyad Mahrez', 'Argelia'),
    ('Mohamed Amoura', 'Argelia'),
    ('Amine Gouiri', 'Argelia'),
    ('Baghdad Bounedjah', 'Argelia'),
    ('Saïd Benrahma', 'Argelia'),
    ('Anis Hadj Moussa', 'Argelia'),
    ('Ibrahim Maza', 'Argelia'),
    ('Bruno Varela', 'Cabo Verde'),
    ('Vozinha', 'Cabo Verde'),
    ('Roberto Lopes', 'Cabo Verde'),
    ('Logan Costa', 'Cabo Verde'),
    ('Stopira', 'Cabo Verde'),
    ('Jojo', 'Cabo Verde'),
    ('Kevin Pina', 'Cabo Verde'),
    ('Laros Duarte', 'Cabo Verde'),
    ('Jamiro Monteiro', 'Cabo Verde'),
    ('Bebé', 'Cabo Verde'),
    ('Ryan Mendes', 'Cabo Verde'),
    ('Jovane Cabral', 'Cabo Verde'),
    ('Benchimol', 'Cabo Verde'),
    ('Gilson Tavares', 'Cabo Verde'),
    ('Kevin Lenini', 'Cabo Verde'),
    ('Yahia Fofana', 'Costa do Marfim'),
    ('Alban Lafont', 'Costa do Marfim'),
    ('Odilon Kossounou', 'Costa do Marfim'),
    ('Evan Ndicka', 'Costa do Marfim'),
    ('Wilfried Singo', 'Costa do Marfim'),
    ('Ghislain Konan', 'Costa do Marfim'),
    ('Franck Kessie', 'Costa do Marfim'),
    ('Seko Fofana', 'Costa do Marfim'),
    ('Ibrahim Sangare', 'Costa do Marfim'),
    ('Amad Diallo', 'Costa do Marfim'),
    ('Simon Adingra', 'Costa do Marfim'),
    ('Sebastien Haller', 'Costa do Marfim'),
    ('Nicolas Pepe', 'Costa do Marfim'),
    ('Oumar Diakite', 'Costa do Marfim'),
    ('Yan Diomande', 'Costa do Marfim'),
    ('Ange-Yoan Bonny', 'Costa do Marfim'),
    ('Karim Konaté', 'Costa do Marfim'),
    ('Lionel Mpasi', 'RD Congo'),
    ('Theo Fayulu', 'RD Congo'),
    ('Matthieu Epolo', 'RD Congo'),
    ('Chancel Mbemba', 'RD Congo'),
    ('Aaron Wan-Bissaka', 'RD Congo'),
    ('Arthur Masuaku', 'RD Congo'),
    ('Joris Kayembe', 'RD Congo'),
    ('Axel Tuanzebe', 'RD Congo'),
    ('Dylan Batubinsika', 'RD Congo'),
    ('Edo Kayembe', 'RD Congo'),
    ('Samuel Moutoussamy', 'RD Congo'),
    ('Noah Sadiki', 'RD Congo'),
    ("Ngal'ayel Mukau", 'RD Congo'),
    ('Yoane Wissa', 'RD Congo'),
    ('Cedric Bakambu', 'RD Congo'),
    ('Simon Banza', 'RD Congo'),
    ('Theo Bongonda', 'RD Congo'),
    ('Mohamed El Shenawy', 'Egito'),
    ('Mohamed Sobhy', 'Egito'),
    ('Mohamed Abdelmonem', 'Egito'),
    ('Ahmed Hegazi', 'Egito'),
    ('Omar Kamal', 'Egito'),
    ('Mohamed Hany', 'Egito'),
    ('Hamdi Fathi', 'Egito'),
    ('Marwan Atiya', 'Egito'),
    ('Mohamed Elneny', 'Egito'),
    ('Trezeguet', 'Egito'),
    ('Omar Marmoush', 'Egito'),
    ('Mohamed Salah', 'Egito'),
    ('Mostafa Mohamed', 'Egito'),
    ('Ibrahim Adel', 'Egito'),
    ('Mustafa Fathi', 'Egito'),
    ('Lawrence Ati-Zigi', 'Gana'),
    ('Joseph Wollacott', 'Gana'),
    ('Alexander Djiku', 'Gana'),
    ('Mohammed Salisu', 'Gana'),
    ('Tariq Lamptey', 'Gana'),
    ('Gideon Mensah', 'Gana'),
    ('Thomas Partey', 'Gana'),
    ('Elisha Owusu', 'Gana'),
    ('Mohammed Kudus', 'Gana'),
    ('Abdul Fatawu Issahaku', 'Gana'),
    ('Antoine Semenyo', 'Gana'),
    ('Inaki Williams', 'Gana'),
    ('Jordan Ayew', 'Gana'),
    ('Ernest Nuamah', 'Gana'),
    ('Ibrahim Osman', 'Gana'),
    ('Salis Abdul Samed', 'Gana'),
    ('Bono', 'Marrocos'),
    ('Munir El Kajoui', 'Marrocos'),
    ('Achraf Hakimi', 'Marrocos'),
    ('Noussair Mazraoui', 'Marrocos'),
    ('Nayef Aguerd', 'Marrocos'),
    ('Romain Saiss', 'Marrocos'),
    ('Sofyan Amrabat', 'Marrocos'),
    ('Azzedine Ounahi', 'Marrocos'),
    ('Bilal El Khannouss', 'Marrocos'),
    ('Brahim Diaz', 'Marrocos'),
    ('Hakim Ziyech', 'Marrocos'),
    ('Amine Adli', 'Marrocos'),
    ('Youssef En-Nesyri', 'Marrocos'),
    ('Ayoub El Kaabi', 'Marrocos'),
    ('Eliesse Ben Seghir', 'Marrocos'),
    ('Chadi Riad', 'Marrocos'),
    ('Sofiane Boufal', 'Marrocos'),
    ('Abde Ezzalzouli', 'Marrocos'),
    ('Edouard Mendy', 'Senegal'),
    ('Mory Diaw', 'Senegal'),
    ('Kalidou Koulibaly', 'Senegal'),
    ('Moussa Niakhate', 'Senegal'),
    ('Ismail Jakobs', 'Senegal'),
    ('Krépin Diatta', 'Senegal'),
    ('Idrissa Gana Gueye', 'Senegal'),
    ('Pape Matar Sarr', 'Senegal'),
    ('Lamine Camara', 'Senegal'),
    ('Nicolas Jackson', 'Senegal'),
    ('Ismaila Sarr', 'Senegal'),
    ('Iliman Ndiaye', 'Senegal'),
    ('Boulaye Dia', 'Senegal'),
    ('Habib Diallo', 'Senegal'),
    ('Sadio Mane', 'Senegal'),
    ('Mikayil Faye', 'Senegal'),
    ('Abdallah Sima', 'Senegal'),
    ('Ronwen Williams', 'Africa do Sul'),
    ('Siyabonga Ngezana', 'Africa do Sul'),
    ('Grant Kekana', 'Africa do Sul'),
    ('Khuliso Mudau', 'Africa do Sul'),
    ('Teboho Mokoena', 'Africa do Sul'),
    ('Mothobi Mvala', 'Africa do Sul'),
    ('Themba Zwane', 'Africa do Sul'),
    ('Percy Tau', 'Africa do Sul'),
    ('Oswin Appollis', 'Africa do Sul'),
    ('Evidence Makgopa', 'Africa do Sul'),
    ('Iqraam Rayners', 'Africa do Sul'),
    ('Relebohile Mofokeng', 'Africa do Sul'),
    ('Shandre Campbell', 'Africa do Sul'),
    ('Aubrey Modiba', 'Africa do Sul'),
    ('Nkosinathi Sibisi', 'Africa do Sul'),
    ('Aymen Dahmen', 'Tunisia'),
    ('Bechir Ben Said', 'Tunisia'),
    ('Montassar Talbi', 'Tunisia'),
    ('Dylan Bronn', 'Tunisia'),
    ('Ali Abdi', 'Tunisia'),
    ('Wajdi Kechrida', 'Tunisia'),
    ('Ellyes Skhiri', 'Tunisia'),
    ('Aissa Laidouni', 'Tunisia'),
    ('Hannibal Mejbri', 'Tunisia'),
    ('Mohamed Ali Ben Romdhane', 'Tunisia'),
    ('Anis Ben Slimane', 'Tunisia'),
    ('Elias Achouri', 'Tunisia'),
    ('Seifeddine Jaziri', 'Tunisia'),
    ('Haythem Jouini', 'Tunisia'),
    ('Youssef Msakni', 'Tunisia'),
    ('Maxime Crépeau', 'Canada'),
    ('Dayne St. Clair', 'Canada'),
    ('Alistair Johnston', 'Canada'),
    ('Alphonso Davies', 'Canada'),
    ('Moïse Bombito', 'Canada'),
    ('Derek Cornelius', 'Canada'),
    ('Ismaël Koné', 'Canada'),
    ('Stephen Eustaquio', 'Canada'),
    ('Jonathan David', 'Canada'),
    ('Cyle Larin', 'Canada'),
    ('Tajon Buchanan', 'Canada'),
    ('Jacob Shaffelburg', 'Canada'),
    ('Mathieu Choinière', 'Canada'),
    ('Richie Laryea', 'Canada'),
    ('Luc de Fougerolles', 'Canada'),
    ('Ali Ahmed', 'Canada'),
    ('Eloy Room', 'Curacao'),
    ('Cuco Martina', 'Curacao'),
    ('Jurien Gaari', 'Curacao'),
    ('Sherel Floranus', 'Curacao'),
    ('Juninho Bacuna', 'Curacao'),
    ('Leandro Bacuna', 'Curacao'),
    ('Vurnon Anita', 'Curacao'),
    ('Kenji Gorré', 'Curacao'),
    ('Jeremy Antonisse', 'Curacao'),
    ('Gervane Kastaneer', 'Curacao'),
    ('Rangelo Janga', 'Curacao'),
    ('Joshua Zimmerman', 'Curacao'),
    ('Livano Comenencia', 'Curacao'),
    ('Johny Placide', 'Haiti'),
    ('Alexandre Pierre', 'Haiti'),
    ('Ricardo Adé', 'Haiti'),
    ('Carlens Arcus', 'Haiti'),
    ('Alex Christian Jr.', 'Haiti'),
    ('Danley Jean Jacques', 'Haiti'),
    ('Bryan Alceus', 'Haiti'),
    ('Leverton Pierre', 'Haiti'),
    ('Louicius Don Deedson', 'Haiti'),
    ('Duckens Nazon', 'Haiti'),
    ('Frantzdy Pierrot', 'Haiti'),
    ('Mondy Prunier', 'Haiti'),
    ('Ruben Providence', 'Haiti'),
    ('Fafa Picault', 'Haiti'),
    ('Luis Ángel Malagón', 'Mexico'),
    ('Guillermo Ochoa', 'Mexico'),
    ('Johan Vásquez', 'Mexico'),
    ('César Montes', 'Mexico'),
    ('Jesús Gallardo', 'Mexico'),
    ('Jorge Sánchez', 'Mexico'),
    ('Edson Álvarez', 'Mexico'),
    ('Luis Chávez', 'Mexico'),
    ('Orbelín Pineda', 'Mexico'),
    ('Marcel Ruiz', 'Mexico'),
    ('Roberto Alvarado', 'Mexico'),
    ('César Huerta', 'Mexico'),
    ('Santiago Giménez', 'Mexico'),
    ('Raúl Jiménez', 'Mexico'),
    ('Hirving Lozano', 'Mexico'),
    ('Julián Quiñones', 'Mexico'),
    ('Fidel Ambriz', 'Mexico'),
    ('Gilberto Mora', 'Mexico'),
    ('Orlando Mosquera', 'Panama'),
    ('César Blackman', 'Panama'),
    ('Éric Davis', 'Panama'),
    ('Fidel Escobar', 'Panama'),
    ('Andrés Andrade', 'Panama'),
    ('José Córdoba', 'Panama'),
    ('Michael Murillo', 'Panama'),
    ('Adalberto Carrasquilla', 'Panama'),
    ('Aníbal Godoy', 'Panama'),
    ('Cristian Martínez', 'Panama'),
    ('Ismael Díaz', 'Panama'),
    ('José Fajardo', 'Panama'),
    ('Eduardo Guerrero', 'Panama'),
    ('José Luis Rodríguez', 'Panama'),
    ('Kahiser Lenis', 'Panama'),
    ('Matt Turner', 'Estados Unidos'),
    ('Zack Steffen', 'Estados Unidos'),
    ('Joe Scally', 'Estados Unidos'),
    ('Sergiño Dest', 'Estados Unidos'),
    ('Antonee Robinson', 'Estados Unidos'),
    ('Chris Richards', 'Estados Unidos'),
    ('Tim Ream', 'Estados Unidos'),
    ('Tyler Adams', 'Estados Unidos'),
    ('Weston McKennie', 'Estados Unidos'),
    ('Yunus Musah', 'Estados Unidos'),
    ('Gio Reyna', 'Estados Unidos'),
    ('Christian Pulisic', 'Estados Unidos'),
    ('Tim Weah', 'Estados Unidos'),
    ('Folarin Balogun', 'Estados Unidos'),
    ('Ricardo Pepi', 'Estados Unidos'),
    ('Josh Sargent', 'Estados Unidos'),
    ('Malik Tillman', 'Estados Unidos'),
    ('Diego Luna', 'Estados Unidos'),
    ('Paxten Aaronson', 'Estados Unidos'),
    ('Emiliano Martínez', 'Argentina'),
    ('Gerónimo Rulli', 'Argentina'),
    ('Cristian Romero', 'Argentina'),
    ('Lisandro Martínez', 'Argentina'),
    ('Nicolás Otamendi', 'Argentina'),
    ('Nahuel Molina', 'Argentina'),
    ('Marcos Acuña', 'Argentina'),
    ('Alexis Mac Allister', 'Argentina'),
    ('Enzo Fernández', 'Argentina'),
    ('Rodrigo De Paul', 'Argentina'),
    ('Leandro Paredes', 'Argentina'),
    ('Exequiel Palacios', 'Argentina'),
    ('Lionel Messi', 'Argentina'),
    ('Julián Álvarez', 'Argentina'),
    ('Lautaro Martínez', 'Argentina'),
    ('Nico González', 'Argentina'),
    ('Thiago Almada', 'Argentina'),
    ('Alejandro Garnacho', 'Argentina'),
    ('Alisson', 'Brasil'),
    ('Ederson', 'Brasil'),
    ('Marquinhos', 'Brasil'),
    ('Gabriel Magalhães', 'Brasil'),
    ('Éder Militão', 'Brasil'),
    ('Beraldo', 'Brasil'),
    ('Guilherme Arana', 'Brasil'),
    ('Danilo', 'Brasil'),
    ('Danilo Santos','Brasil'),
    ('Bruno Guimarães', 'Brasil'),
    ('João Gomes', 'Brasil'),
    ('Lucas Paquetá', 'Brasil'),
    ('Rodrygo', 'Brasil'),
    ('Vinícius Júnior', 'Brasil'),
    ('Raphinha', 'Brasil'),
    ('Endrick', 'Brasil'),
    ('Gabriel Martinelli', 'Brasil'),
    ('Savinho', 'Brasil'),
    ('João Pedro', 'Brasil'),
    ('Estevão', 'Brasil'),
    ('Neymar', 'Brasil'),
    ('Camilo Vargas', 'Colombia'),
    ('Álvaro Montero', 'Colombia'),
    ('Davinson Sánchez', 'Colombia'),
    ('Jhon Lucumí', 'Colombia'),
    ('Daniel Muñoz', 'Colombia'),
    ('Deiver Machado', 'Colombia'),
    ('Johan Mojica', 'Colombia'),
    ('Jefferson Lerma', 'Colombia'),
    ('Richard Ríos', 'Colombia'),
    ('Kevin Castaño', 'Colombia'),
    ('James Rodríguez', 'Colombia'),
    ('Jhon Arias', 'Colombia'),
    ('Luis Díaz', 'Colombia'),
    ('Jhon Durán', 'Colombia'),
    ('Rafael Santos Borré', 'Colombia'),
    ('Andrés Gómez', 'Colombia'),
    ('Luis Sinisterra', 'Colombia'),
    ('Yáser Asprilla', 'Colombia'),
    ('Hernán Galíndez', 'Equador'),
    ('Alexander Domínguez', 'Equador'),
    ('Willian Pacho', 'Equador'),
    ('Piero Hincapié', 'Equador'),
    ('Félix Torres', 'Equador'),
    ('Angelo Preciado', 'Equador'),
    ('Moisés Caicedo', 'Equador'),
    ('Alan Franco', 'Equador'),
    ('Kendry Páez', 'Equador'),
    ('Jeremy Sarmiento', 'Equador'),
    ('Gonzalo Plata', 'Equador'),
    ('Kevin Rodríguez', 'Equador'),
    ('Enner Valencia', 'Equador'),
    ('John Yeboah', 'Equador'),
    ('Nilson Angulo', 'Equador'),
    ('Justin Lerma', 'Equador'),
    ('Carlos Coronel', 'Paraguai'),
    ('Gatito Fernández', 'Paraguai'),
    ('Fabián Balbuena', 'Paraguai'),
    ('Omar Alderete', 'Paraguai'),
    ('Juan José Cáceres', 'Paraguai'),
    ('Diego Gómez', 'Paraguai'),
    ('Mathías Villasanti', 'Paraguai'),
    ('Andrés Cubas', 'Paraguai'),
    ('Miguel Almirón', 'Paraguai'),
    ('Julio Enciso', 'Paraguai'),
    ('Antonio Sanabria', 'Paraguai'),
    ('Ramón Sosa', 'Paraguai'),
    ('Enso González', 'Paraguai'),
    ('Diego León', 'Paraguai'),
    ('Gabriel Ávalos', 'Paraguai'),
    ('Sergio Rochet', 'Uruguai'),
    ('Santiago Mele', 'Uruguai'),
    ('Ronald Araújo', 'Uruguai'),
    ('José María Giménez', 'Uruguai'),
    ('Sebastián Cáceres', 'Uruguai'),
    ('Mathías Olivera', 'Uruguai'),
    ('Nahitan Nández', 'Uruguai'),
    ('Federico Valverde', 'Uruguai'),
    ('Manuel Ugarte', 'Uruguai'),
    ('Rodrigo Bentancur', 'Uruguai'),
    ('Nicolás de la Cruz', 'Uruguai'),
    ('Darwin Núñez', 'Uruguai'),
    ('Facundo Pellistri', 'Uruguai'),
    ('Maximiliano Araújo', 'Uruguai'),
    ('Luciano Rodríguez', 'Uruguai'),
    ('Puma Rodríguez', 'Uruguai'),
    ('Giorgian de Arrascaeta', 'Uruguai'),
    ('Franco González', 'Uruguai'),
    ('Brian Rodríguez', 'Uruguai'),
    ('Max Crocombe', 'Nova Zelandia'),
    ('Michael Boxall', 'Nova Zelandia'),
    ('Tommy Smith', 'Nova Zelandia'),
    ('Liberato Cacace', 'Nova Zelandia'),
    ('Tim Payne', 'Nova Zelandia'),
    ('Tyler Bindon', 'Nova Zelandia'),
    ('Marko Stamenic', 'Nova Zelandia'),
    ('Joe Bell', 'Nova Zelandia'),
    ('Sarpreet Singh', 'Nova Zelandia'),
    ('Matt Garbett', 'Nova Zelandia'),
    ('Elijah Just', 'Nova Zelandia'),
    ('Ben Waine', 'Nova Zelandia'),
    ('Chris Wood', 'Nova Zelandia'),
    ('Callum McCowatt', 'Nova Zelandia'),
    ('Alexander Schlager', 'Austria'),
    ('Patrick Pentz', 'Austria'),
    ('David Alaba', 'Austria'),
    ('Kevin Danso', 'Austria'),
    ('Philipp Lienhart', 'Austria'),
    ('Stefan Posch', 'Austria'),
    ('Alexander Prass', 'Austria'),
    ('Konrad Laimer', 'Austria'),
    ('Nicolas Seiwald', 'Austria'),
    ('Christoph Baumgartner', 'Austria'),
    ('Marcel Sabitzer', 'Austria'),
    ('Romano Schmid', 'Austria'),
    ('Marko Arnautović', 'Austria'),
    ('Michael Gregoritsch', 'Austria'),
    ('Samson Baidoo', 'Austria'),
    ('Raul Florucz', 'Austria'),
    ('Koen Casteels', 'Belgica'),
    ('Matz Sels', 'Belgica'),
    ('Wout Faes', 'Belgica'),
    ('Arthur Theate', 'Belgica'),
    ('Zeno Debast', 'Belgica'),
    ('Timothy Castagne', 'Belgica'),
    ('Amadou Onana', 'Belgica'),
    ('Orel Mangala', 'Belgica'),
    ('Youri Tielemans', 'Belgica'),
    ('Kevin De Bruyne', 'Belgica'),
    ('Jérémy Doku', 'Belgica'),
    ('Charles De Ketelaere', 'Belgica'),
    ('Lois Openda', 'Belgica'),
    ('Romelu Lukaku', 'Belgica'),
    ('Arthur Vermeeren', 'Belgica'),
    ('Malick Fofana', 'Belgica'),
    ('Julien Duranville', 'Belgica'),
    ('Leandro Trossard', 'Belgica'),
    ('Nikola Vasilj', 'Bosnia e Herzegovina'),
    ('Ibrahim Šehić', 'Bosnia e Herzegovina'),
    ('Amar Dedić', 'Bosnia e Herzegovina'),
    ('Sead Kolašinac', 'Bosnia e Herzegovina'),
    ('Dennis Hadžikadunić', 'Bosnia e Herzegovina'),
    ('Jusuf Gazibegović', 'Bosnia e Herzegovina'),
    ('Benjamin Tahirović', 'Bosnia e Herzegovina'),
    ('Haris Hajradinović', 'Bosnia e Herzegovina'),
    ('Ivan Bašić', 'Bosnia e Herzegovina'),
    ('Dženis Burnić', 'Bosnia e Herzegovina'),
    ('Esmir Bajraktarević', 'Bosnia e Herzegovina'),
    ('Edin Džeko', 'Bosnia e Herzegovina'),
    ('Ermedin Demirović', 'Bosnia e Herzegovina'),
    ('Smail Prevljak', 'Bosnia e Herzegovina'),
    ('Dominik Livaković', 'Croacia'),
    ('Josip Šutalo', 'Croacia'),
    ('Joško Gvardiol', 'Croacia'),
    ('Josip Stanišić', 'Croacia'),
    ('Borna Sosa', 'Croacia'),
    ('Luka Modrić', 'Croacia'),
    ('Mateo Kovačić', 'Croacia'),
    ('Marcelo Brozović', 'Croacia'),
    ('Luka Sučić', 'Croacia'),
    ('Martin Baturina', 'Croacia'),
    ('Mario Pašalić', 'Croacia'),
    ('Andrej Kramarić', 'Croacia'),
    ('Ante Budimir', 'Croacia'),
    ('Ivan Perišić', 'Croacia'),
    ('Lovro Majer', 'Croacia'),
    ('Petar Sučić', 'Croacia'),
    ('Dion Drena Beljo', 'Croacia'),
    ('Jindřich Staněk', 'Tchequia'),
    ('Tomáš Vaclík', 'Tchequia'),
    ('Tomáš Holeš', 'Tchequia'),
    ('Robin Hranáč', 'Tchequia'),
    ('Vladimír Coufal', 'Tchequia'),
    ('David Douděra', 'Tchequia'),
    ('Tomáš Souček', 'Tchequia'),
    ('Antonín Barák', 'Tchequia'),
    ('Lukáš Provod', 'Tchequia'),
    ('Adam Hložek', 'Tchequia'),
    ('Patrik Schick', 'Tchequia'),
    ('Václav Černý', 'Tchequia'),
    ('Jan Kuchta', 'Tchequia'),
    ('Matěj Jurásek', 'Tchequia'),
    ('Adam Karabec', 'Tchequia'),
    ('Jordan Pickford', 'Inglaterra'),
    ('Aaron Ramsdale', 'Inglaterra'),
    ('John Stones', 'Inglaterra'),
    ('Marc Guéhi', 'Inglaterra'),
    ('Ezri Konsa', 'Inglaterra'),
    ('Luke Shaw', 'Inglaterra'),
    ('Trent Alexander-Arnold', 'Inglaterra'),
    ('Rico Lewis', 'Inglaterra'),
    ('Declan Rice', 'Inglaterra'),
    ('Jude Bellingham', 'Inglaterra'),
    ('Cole Palmer', 'Inglaterra'),
    ('Bukayo Saka', 'Inglaterra'),
    ('Phil Foden', 'Inglaterra'),
    ('Anthony Gordon', 'Inglaterra'),
    ('Harry Kane', 'Inglaterra'),
    ('Ollie Watkins', 'Inglaterra'),
    ('Ivan Toney', 'Inglaterra'),
    ('Jarell Quansah', 'Inglaterra'),
    ('Ethan Nwaneri', 'Inglaterra'),
    ('Reece James', 'Inglaterra'),
    ('Mike Maignan', 'Franca'),
    ('Brice Samba', 'Franca'),
    ('William Saliba', 'Franca'),
    ('Ibrahima Konaté', 'Franca'),
    ('Dayot Upamecano', 'Franca'),
    ('Jules Koundé', 'Franca'),
    ('Theo Hernandez', 'Franca'),
    ('Aurélien Tchouaméni', 'Franca'),
    ('Eduardo Camavinga', 'Franca'),
    ('Warren Zaïre-Emery', 'Franca'),
    ('Michael Olise', 'Franca'),
    ('Ousmane Dembélé', 'Franca'),
    ('Bradley Barcola', 'Franca'),
    ('Kylian Mbappé', 'Franca'),
    ('Marcus Thuram', 'Franca'),
    ('Randal Kolo Muani', 'Franca'),
    ('Désiré Doué', 'Franca'),
    ('Mathys Tel', 'Franca'),
    ('Youssouf Fofana', 'Franca'),
    ('Marc-André ter Stegen', 'Alemanha'),
    ('Manuel Neuer', 'Alemanha'),
    ('Antonio Rüdiger', 'Alemanha'),
    ('Jonathan Tah', 'Alemanha'),
    ('Nico Schlotterbeck', 'Alemanha'),
    ('David Raum', 'Alemanha'),
    ('Joshua Kimmich', 'Alemanha'),
    ('Ilkay Gündoğan', 'Alemanha'),
    ('Robert Andrich', 'Alemanha'),
    ('Aleksandar Pavlović', 'Alemanha'),
    ('Florian Wirtz', 'Alemanha'),
    ('Jamal Musiala', 'Alemanha'),
    ('Kai Havertz', 'Alemanha'),
    ('Leroy Sané', 'Alemanha'),
    ('Karim Adeyemi', 'Alemanha'),
    ('Niclas Füllkrug', 'Alemanha'),
    ('Maximilian Beier', 'Alemanha'),
    ('Brajan Gruda', 'Alemanha'),
    ('Bart Verbruggen', 'Holanda'),
    ('Mark Flekken', 'Holanda'),
    ('Virgil van Dijk', 'Holanda'),
    ('Matthijs de Ligt', 'Holanda'),
    ('Nathan Aké', 'Holanda'),
    ('Jeremie Frimpong', 'Holanda'),
    ('Denzel Dumfries', 'Holanda'),
    ('Frenkie de Jong', 'Holanda'),
    ('Teun Koopmeiners', 'Holanda'),
    ('Tijjani Reijnders', 'Holanda'),
    ('Xavi Simons', 'Holanda'),
    ('Cody Gakpo', 'Holanda'),
    ('Joshua Zirkzee', 'Holanda'),
    ('Brian Brobbey', 'Holanda'),
    ('Noa Lang', 'Holanda'),
    ('Jorrel Hato', 'Holanda'),
    ('Ian Maatsen', 'Holanda'),
    ('Memphis Depay', 'Holanda'),
    ('Ørjan Nyland', 'Noruega'),
    ('Kristoffer Ajer', 'Noruega'),
    ('Leo Østigård', 'Noruega'),
    ('Julian Ryerson', 'Noruega'),
    ('Andreas Hanche-Olsen', 'Noruega'),
    ('Martin Ødegaard', 'Noruega'),
    ('Sander Berge', 'Noruega'),
    ('Patrick Berg', 'Noruega'),
    ('Oscar Bobb', 'Noruega'),
    ('Antonio Nusa', 'Noruega'),
    ('Erling Haaland', 'Noruega'),
    ('Alexander Sørloth', 'Noruega'),
    ('Andreas Schjelderup', 'Noruega'),
    ('Sverre Nypan', 'Noruega'),
    ('Fredrik Aursnes', 'Noruega'),
    ('Diogo Costa', 'Portugal'),
    ('José Sá', 'Portugal'),
    ('Rúben Dias', 'Portugal'),
    ('Gonçalo Inácio', 'Portugal'),
    ('António Silva', 'Portugal'),
    ('Nuno Mendes', 'Portugal'),
    ('João Neves', 'Portugal'),
    ('Vitinha', 'Portugal'),
    ('Bruno Fernandes', 'Portugal'),
    ('Bernardo Silva', 'Portugal'),
    ('Rafael Leão', 'Portugal'),
    ('Pedro Neto', 'Portugal'),
    ('Gonçalo Ramos', 'Portugal'),
    ('João Félix', 'Portugal'),
    ('Cristiano Ronaldo', 'Portugal'),
    ('Rúben Neves', 'Portugal'),
    ('Rodrigo Mora', 'Portugal'),
    ('Geovany Quenda', 'Portugal'),
    ('Francisco Conceição', 'Portugal'),
    ('Angus Gunn', 'Escocia'),
    ('Craig Gordon', 'Escocia'),
    ('Andrew Robertson', 'Escocia'),
    ('Kieran Tierney', 'Escocia'),
    ('Grant Hanley', 'Escocia'),
    ('Scott McTominay', 'Escocia'),
    ('John McGinn', 'Escocia'),
    ('Billy Gilmour', 'Escocia'),
    ('Kenny McLean', 'Escocia'),
    ('Ryan Christie', 'Escocia'),
    ('Che Adams', 'Escocia'),
    ('Lawrence Shankland', 'Escocia'),
    ('Ben Doak', 'Escocia'),
    ('Tommy Conway', 'Escocia'),
    ('Lewis Ferguson', 'Escocia'),
    ('Unai Simón', 'Espanha'),
    ('David Raya', 'Espanha'),
    ('Robin Le Normand', 'Espanha'),
    ('Pau Cubarsí', 'Espanha'),
    ('Dani Vivian', 'Espanha'),
    ('Marc Cucurella', 'Espanha'),
    ('Rodri', 'Espanha'),
    ('Pedri', 'Espanha'),
    ('Gavi', 'Espanha'),
    ('Martín Zubimendi', 'Espanha'),
    ('Fabián Ruiz', 'Espanha'),
    ('Dani Olmo', 'Espanha'),
    ('Nico Williams', 'Espanha'),
    ('Lamine Yamal', 'Espanha'),
    ('Álvaro Morata', 'Espanha'),
    ('Mikel Oyarzabal', 'Espanha'),
    ('Samu Omorodion', 'Espanha'),
    ('Fermín López', 'Espanha'),
    ('Aleix García', 'Espanha'),
    ('Robin Olsen', 'Suecia'),
    ('Victor Lindelöf', 'Suecia'),
    ('Isak Hien', 'Suecia'),
    ('Emil Holm', 'Suecia'),
    ('Ludwig Augustinsson', 'Suecia'),
    ('Hugo Larsson', 'Suecia'),
    ('Dejan Kulusevski', 'Suecia'),
    ('Emil Forsberg', 'Suecia'),
    ('Jesper Karlsson', 'Suecia'),
    ('Viktor Gyökeres', 'Suecia'),
    ('Alexander Isak', 'Suecia'),
    ('Anthony Elanga', 'Suecia'),
    ('Lucas Bergvall', 'Suecia'),
    ('Roony Bardghji', 'Suecia'),
    ('Samuel Dahl', 'Suecia'),
    ('Carl Starfelt', 'Suecia'),
    ('Gregor Kobel', 'Suica'),
    ('Yann Sommer', 'Suica'),
    ('Manuel Akanji', 'Suica'),
    ('Fabian Schär', 'Suica'),
    ('Ricardo Rodríguez', 'Suica'),
    ('Silvan Widmer', 'Suica'),
    ('Granit Xhaka', 'Suica'),
    ('Denis Zakaria', 'Suica'),
    ('Remo Freuler', 'Suica'),
    ('Ardon Jashari', 'Suica'),
    ('Dan Ndoye', 'Suica'),
    ('Ruben Vargas', 'Suica'),
    ('Breel Embolo', 'Suica'),
    ('Zeki Amdouni', 'Suica'),
    ('Michel Aebischer', 'Suica'),
    ('Leon Avdullahu', 'Suica'),
    ('Nico Elvedi', 'Suica'),
    ('Mert Günok', 'Turquia'),
    ('Altay Bayındır', 'Turquia'),
    ('Merih Demiral', 'Turquia'),
    ('Abdülkerim Bardakcı', 'Turquia'),
    ('Ferdi Kadıoğlu', 'Turquia'),
    ('Kaan Ayhan', 'Turquia'),
    ('Hakan Çalhanoğlu', 'Turquia'),
    ('Orkun Kökçü', 'Turquia'),
    ('Salih Özcan', 'Turquia'),
    ('Arda Güler', 'Turquia'),
    ('Kenan Yıldız', 'Turquia'),
    ('Kerem Aktürkoğlu', 'Turquia'),
    ('Barış Alper Yılmaz', 'Turquia'),
    ('Semih Kılıçsoy', 'Turquia'),
    ('Can Uzun', 'Turquia'),
    ('Yusuf Akçiçek', 'Turquia'),
    ('İsmail Yüksek', 'Turquia'),
]


# ── Cache persistente em disco (sobrevive a reinicializações do processo)
_CACHE_PATH = SCRIPT_DIR / ".bolao_cache.pkl"

def _compute_fingerprint(paths):
    parts = []
    for p in sorted(str(x) for x in paths):
        try:
            s = Path(p).stat()
            parts.append(f"{p}|{s.st_mtime_ns}|{s.st_size}")
        except Exception:
            parts.append(f"{p}|missing")
    return hashlib.md5("\n".join(parts).encode()).hexdigest()

@st.cache_resource(show_spinner=False)
def _disk_cache_get(fingerprint):
    """Lê do cache em disco se o fingerprint bater; retorna None caso contrário."""
    if not _CACHE_PATH.exists():
        return None
    try:
        with open(_CACHE_PATH, "rb") as _f:
            _c = pickle.load(_f)
        if _c.get("fp") == fingerprint:
            return _c["data"]
    except Exception:
        pass
    return None

# ══════════════════════════════════════════════════════════════════════
# LOGIN GATE
# ══════════════════════════════════════════════════════════════════════
#_PWD = os.getenv("APP_PASSWORD")
if "auth" not in st.session_state:
    st.session_state.auth = False
 
if not st.session_state.auth:
    fp_b64 = img_to_b64(FRONT_PAGE) if FRONT_PAGE else ""
 
    # ── Background image + chrome removal ─────────────────────────────
    st.markdown(f"""
    <style>
    section[data-testid="stSidebar"],#MainMenu,footer,header,
    [data-testid="stHeader"],[data-testid="stToolbar"]{{display:none!important;}}
    html,body{{overflow:hidden!important;height:100%!important;}}
    .stApp{{overflow:hidden!important;height:100vh!important;}}
    [data-testid="stMain"]{{overflow:hidden!important;}}
 
    /* Full-screen background from front_page.png */
    .stApp,[data-testid="stAppViewContainer"]{{
        background: url("data:image/png;base64,{fp_b64}") left center / cover no-repeat !important;
        min-height:100vh;
    }}
    [data-testid="stMain"]>div{{background:transparent!important;}}
 
    /* Remove all padding, full width */
    .block-container{{
        padding:0 44px!important; max-width:100%!important;
        background:transparent!important;
        min-height:100vh;
    }}
 
    /* Columns fill full viewport height */
    [data-testid="stHorizontalBlock"]{{
        min-height:100vh; gap:0!important;
        align-items:stretch;
    }}
    [data-testid="column"]{{min-height:100vh;}}
 
    /* Left spacer: transparent */
    [data-testid="column"]:nth-child(1){{
        background:transparent!important;
    }}
 
    /* Right panel: dark glass */
    [data-testid="column"]:nth-child(2){{
        background:rgba(6,14,26,.92)!important;
        backdrop-filter:blur(22px)!important;
        -webkit-backdrop-filter:blur(22px)!important;
        border-left:1px solid rgba(214,184,100,.22)!important;
        box-shadow:-16px 0 60px rgba(0,0,0,.6)!important;
        padding:0!important;
        display:flex!important;
        flex-direction:column!important;
        justify-content:flex-start!important;
        padding:0 44px!important;
    }}
 
    /* Input */
    [data-testid="column"]:nth-child(2) input[type="password"]{{
        background:rgba(255,255,255,.07)!important;
        border:1px solid rgba(214,184,100,.35)!important;
        border-radius:10px!important; color:#fff!important;
        font-size:.95rem!important; padding:13px 15px!important;
        transition:border-color .2s,box-shadow .2s;
    }}
    [data-testid="column"]:nth-child(2) input[type="password"]:focus{{
        border-color:#D6B864!important;
        box-shadow:0 0 0 3px rgba(214,184,100,.2)!important;
    }}
    [data-testid="column"]:nth-child(2) input::placeholder{{
        color:rgba(255,255,255,.28)!important;
    }}
    [data-testid="column"]:nth-child(2) label{{
        display:inline-flex!important;
        align-items:center!important;
        background:rgba(214,184,100,.15)!important;
        border:1px solid rgba(214,184,100,.4)!important;
        border-radius:24px!important;
        padding:5px 14px!important;
        color:#D6B864!important;
        font-size:.7rem!important;
        font-weight:700!important;
        text-transform:uppercase!important;
        letter-spacing:.8px!important;
        margin-bottom:8px!important;
        width:auto!important;
    }}
 
    /* Button */
    [data-testid="column"]:nth-child(2) .stButton>button{{
        width:100%;
        background:#ffffff!important;
        color:#0D2B40!important; border:none!important;
        border-radius:10px!important; font-weight:800!important;
        font-size:1rem!important; letter-spacing:.3px!important;
        padding:14px 0!important; margin-top:8px!important;
        box-shadow:0 2px 12px rgba(0,0,0,.25)!important;
        transition:all .25s ease!important;
    }}
    [data-testid="column"]:nth-child(2) .stButton>button:hover{{
        background:linear-gradient(135deg,#D6B864 0%,#b89640 100%)!important;
        color:#0D2B40!important;
        transform:translateY(-2px)!important;
        box-shadow:0 8px 28px rgba(214,184,100,.55)!important;
    }}
 
    /* Alert */
    [data-testid="column"]:nth-child(2) [data-testid="stAlert"]{{
        background:rgba(196,30,58,.18)!important;
        border:1px solid rgba(196,30,58,.45)!important;
        border-radius:8px!important;
        color:#FCA5A5!important;
    }}
    </style>
    """, unsafe_allow_html=True)
 
    # ── Two columns: left spacer | right login panel ──────────────────
    _, rcol = st.columns([58, 42])
    with rcol:
        # Spacer — pushes content to lower half of the panel
        st.markdown("<div style='height:42vh'></div>", unsafe_allow_html=True)
        # Badge
        st.markdown("""
        <div style="margin-top:0;margin-bottom:22px">
          <span style="display:inline-flex;align-items:center;gap:6px;
            background:rgba(214,184,100,.14);border:1px solid rgba(214,184,100,.4);
            border-radius:24px;padding:5px 14px;font-size:.7rem;font-weight:700;
            color:#D6B864;letter-spacing:.8px;text-transform:uppercase">
            ⚽ &nbsp;TURIM MFO &nbsp;·&nbsp; ACESSO RESTRITO
          </span>
        </div>
        <div style="font-size:2.1rem;font-weight:900;color:#fff;
                    line-height:1.15;margin-bottom:8px;letter-spacing:-.5px">
          Bolão<br><span style="color:#D6B864">Copa 2026</span>
        </div>
        <div style="font-size:.82rem;color:rgba(255,255,255,.45);margin-bottom:28px">
          EUA &nbsp;·&nbsp; México &nbsp;·&nbsp; Canadá &nbsp;·&nbsp; Jun–Jul 2026
        </div>
        <div style="height:1px;background:linear-gradient(90deg,rgba(214,184,100,.45),transparent);
                    margin-bottom:26px"></div>
        """, unsafe_allow_html=True)
 
        # Badge label rendered manually (native label hidden)
        st.markdown("""
        <div style="display:inline-flex;align-items:center;gap:6px;
                    background:rgba(214,184,100,.15);border:1px solid rgba(214,184,100,.4);
                    border-radius:24px;padding:5px 14px;
                    color:#D6B864;font-size:.7rem;font-weight:700;
                    text-transform:uppercase;letter-spacing:.8px;margin-bottom:8px">
          🔑 &nbsp;SENHA DO BOLÃO
        </div>
        """, unsafe_allow_html=True)
        st.markdown("""<style>
        div[data-testid="column"]:nth-child(2) div.stButton button:hover {
            background: linear-gradient(135deg,#D6B864 0%,#b89640 100%) !important;
            color: #0D2B40 !important;
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 28px rgba(214,184,100,.55) !important;
        }
        /* Hide form border/background */
        [data-testid="stForm"] {
            border: none !important;
            background: transparent !important;
            padding: 0 !important;
        }
        </style>""", unsafe_allow_html=True)
        with st.form("login_form", border=False):
            pwd = st.text_input(
                "Senha", type="password",
                placeholder="Digite a senha...",
                label_visibility="collapsed",
            )
            submitted = st.form_submit_button("Entrar →", width='stretch')
            if submitted:
                if pwd == _PWD:
                    st.session_state.auth = True
                    st.rerun()
                else:
                    st.error("Senha incorreta. Tente novamente.")
 
        st.markdown("""
        <div style="text-align:center;margin-top:30px;margin-bottom:12px;
                    font-size:.68rem;color:rgba(255,255,255,.22);letter-spacing:.6px">
          🐂 TURIM &nbsp;·&nbsp; TORI &nbsp;·&nbsp; RUMO AO HEXA!
        </div>
        """, unsafe_allow_html=True)
    st.stop()

# ══════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════
FLAGS={
    'Mexico':'🇲🇽','Africa do Sul':'🇿🇦','Coreia do Sul':'🇰🇷','Tchequia':'🇨🇿',
    'Canada':'🇨🇦','Bosnia e Herzegovina':'🇧🇦','Qatar':'🇶🇦','Suica':'🇨🇭',
    'Brasil':'🇧🇷','Marrocos':'🇲🇦','Haiti':'🇭🇹','Escocia':'🏴󠁧󠁢󠁳󠁣󠁴󠁿',
    'Estados Unidos':'🇺🇸','Paraguai':'🇵🇾','Australia':'🇦🇺','Turquia':'🇹🇷',
    'Alemanha':'🇩🇪','Curacao':'🇨🇼','Costa do Marfim':'🇨🇮','Equador':'🇪🇨',
    'Holanda':'🇳🇱','Japao':'🇯🇵','Suecia':'🇸🇪','Tunisia':'🇹🇳',
    'Belgica':'🇧🇪','Egito':'🇪🇬','Ira':'🇮🇷','Nova Zelandia':'🇳🇿',
    'Espanha':'🇪🇸','Cabo Verde':'🇨🇻','Arabia Saudita':'🇸🇦','Uruguai':'🇺🇾',
    'Franca':'🇫🇷','Senegal':'🇸🇳','Iraque':'🇮🇶','Noruega':'🇳🇴',
    'Argentina':'🇦🇷','Argelia':'🇩🇿','Austria':'🇦🇹','Jordania':'🇯🇴',
    'Portugal':'🇵🇹','RD Congo':'🇨🇩','Uzbequistao':'🇺🇿','Colombia':'🇨🇴',
    'Inglaterra':'🏴󠁧󠁢󠁥󠁮󠁧󠁿','Croacia':'🇭🇷','Gana':'🇬🇭','Panama':'🇵🇦',
}
def F(t): return FLAGS.get(str(t),'🌍') if t and str(t) not in ('?','') else ''

COUNTRY_ISO = {
    'Mexico':'mx','Africa do Sul':'za','Coreia do Sul':'kr','Tchequia':'cz',
    'Canada':'ca','Bosnia e Herzegovina':'ba','Qatar':'qa','Suica':'ch',
    'Brasil':'br','Marrocos':'ma','Haiti':'ht','Escocia':'gb-sct',
    'Estados Unidos':'us','Paraguai':'py','Australia':'au','Turquia':'tr',
    'Alemanha':'de','Curacao':'cw','Costa do Marfim':'ci','Equador':'ec',
    'Holanda':'nl','Japao':'jp','Suecia':'se','Tunisia':'tn',
    'Belgica':'be','Egito':'eg','Ira':'ir','Nova Zelandia':'nz',
    'Espanha':'es','Cabo Verde':'cv','Arabia Saudita':'sa','Uruguai':'uy',
    'Franca':'fr','Senegal':'sn','Iraque':'iq','Noruega':'no',
    'Argentina':'ar','Argelia':'dz','Austria':'at','Jordania':'jo',
    'Portugal':'pt','RD Congo':'cd','Uzbequistao':'uz','Colombia':'co',
    'Inglaterra':'gb-eng','Croacia':'hr','Gana':'gh','Panama':'pa',
}
def FI(t):
    """Flag img tag for HTML blocks (works on all platforms)."""
    if not t or str(t) in ('?',''):
        return ''
    code = COUNTRY_ISO.get(str(t),'')
    if not code:
        return ''
    return f'<img src="https://flagcdn.com/16x12/{code}.png" style="vertical-align:middle;margin-right:5px;border-radius:1px" loading="lazy">'

GRP_COLORS={
    'A':'#123A56','B':'#0D2B40','C':'#0D8587','D':'#1a6b6d',
    'E':'#B2584E','F':'#8B3D35','G':'#DC884A','H':'#a86028',
    'I':'#D6B864','J':'#b89640','K':'#0D2B40','L':'#123A56',
}
FIFA_RANKINGS={
    'Mexico':16,'Africa do Sul':63,'Coreia do Sul':23,'Tchequia':38,
    'Canada':44,'Bosnia e Herzegovina':55,'Qatar':38,'Suica':20,
    'Brasil':5,'Marrocos':14,'Haiti':86,'Escocia':27,
    'Estados Unidos':11,'Paraguai':51,'Australia':23,'Turquia':36,
    'Alemanha':12,'Curacao':83,'Costa do Marfim':41,'Equador':39,
    'Holanda':8,'Japao':15,'Suecia':26,'Tunisia':30,
    'Belgica':3,'Egito':36,'Ira':25,'Nova Zelandia':98,
    'Espanha':4,'Cabo Verde':88,'Arabia Saudita':57,'Uruguai':17,
    'Franca':2,'Senegal':21,'Iraque':65,'Noruega':28,
    'Argentina':1,'Argelia':32,'Austria':24,'Jordania':86,
    'Portugal':6,'RD Congo':57,'Uzbequistao':74,'Colombia':18,
    'Inglaterra':9,'Croacia':10,'Gana':52,'Panama':72,
}
GROUPS_DATA=OrderedDict([
    ('A',['Mexico','Africa do Sul','Coreia do Sul','Tchequia']),
    ('B',['Canada','Bosnia e Herzegovina','Qatar','Suica']),
    ('C',['Brasil','Marrocos','Haiti','Escocia']),
    ('D',['Estados Unidos','Paraguai','Australia','Turquia']),
    ('E',['Alemanha','Curacao','Costa do Marfim','Equador']),
    ('F',['Holanda','Japao','Suecia','Tunisia']),
    ('G',['Belgica','Egito','Ira','Nova Zelandia']),
    ('H',['Espanha','Cabo Verde','Arabia Saudita','Uruguai']),
    ('I',['Franca','Senegal','Iraque','Noruega']),
    ('J',['Argentina','Argelia','Austria','Jordania']),
    ('K',['Portugal','RD Congo','Uzbequistao','Colombia']),
    ('L',['Inglaterra','Croacia','Gana','Panama']),
])
GL=list(GROUPS_DATA.keys())
GROUP_FIXTURES=[
    (date(2026,6,11),'A','Mexico','Africa do Sul'),(date(2026,6,11),'A','Coreia do Sul','Tchequia'),
    (date(2026,6,18),'A','Tchequia','Africa do Sul'),(date(2026,6,18),'A','Mexico','Coreia do Sul'),
    (date(2026,6,24),'A','Tchequia','Mexico'),(date(2026,6,24),'A','Africa do Sul','Coreia do Sul'),
    (date(2026,6,12),'B','Canada','Bosnia e Herzegovina'),(date(2026,6,13),'B','Qatar','Suica'),
    (date(2026,6,18),'B','Suica','Bosnia e Herzegovina'),(date(2026,6,18),'B','Canada','Qatar'),
    (date(2026,6,24),'B','Suica','Canada'),(date(2026,6,24),'B','Bosnia e Herzegovina','Qatar'),
    (date(2026,6,13),'C','Brasil','Marrocos'),(date(2026,6,13),'C','Haiti','Escocia'),
    (date(2026,6,19),'C','Escocia','Marrocos'),(date(2026,6,19),'C','Brasil','Haiti'),
    (date(2026,6,24),'C','Escocia','Brasil'),(date(2026,6,24),'C','Marrocos','Haiti'),
    (date(2026,6,12),'D','Estados Unidos','Paraguai'),(date(2026,6,13),'D','Australia','Turquia'),
    (date(2026,6,19),'D','Estados Unidos','Australia'),(date(2026,6,19),'D','Turquia','Paraguai'),
    (date(2026,6,25),'D','Turquia','Estados Unidos'),(date(2026,6,25),'D','Paraguai','Australia'),
    (date(2026,6,14),'E','Alemanha','Curacao'),(date(2026,6,14),'E','Costa do Marfim','Equador'),
    (date(2026,6,20),'E','Alemanha','Costa do Marfim'),(date(2026,6,20),'E','Equador','Curacao'),
    (date(2026,6,25),'E','Equador','Alemanha'),(date(2026,6,25),'E','Curacao','Costa do Marfim'),
    (date(2026,6,14),'F','Holanda','Japao'),(date(2026,6,14),'F','Suecia','Tunisia'),
    (date(2026,6,20),'F','Holanda','Suecia'),(date(2026,6,20),'F','Tunisia','Japao'),
    (date(2026,6,25),'F','Japao','Suecia'),(date(2026,6,25),'F','Tunisia','Holanda'),
    (date(2026,6,15),'G','Belgica','Egito'),(date(2026,6,15),'G','Ira','Nova Zelandia'),
    (date(2026,6,21),'G','Belgica','Ira'),(date(2026,6,21),'G','Nova Zelandia','Egito'),
    (date(2026,6,26),'G','Egito','Ira'),(date(2026,6,26),'G','Nova Zelandia','Belgica'),
    (date(2026,6,15),'H','Espanha','Cabo Verde'),(date(2026,6,15),'H','Arabia Saudita','Uruguai'),
    (date(2026,6,21),'H','Espanha','Arabia Saudita'),(date(2026,6,21),'H','Uruguai','Cabo Verde'),
    (date(2026,6,26),'H','Cabo Verde','Arabia Saudita'),(date(2026,6,26),'H','Uruguai','Espanha'),
    (date(2026,6,16),'I','Franca','Senegal'),(date(2026,6,16),'I','Iraque','Noruega'),
    (date(2026,6,22),'I','Franca','Iraque'),(date(2026,6,22),'I','Noruega','Senegal'),
    (date(2026,6,26),'I','Noruega','Franca'),(date(2026,6,26),'I','Senegal','Iraque'),
    (date(2026,6,16),'J','Argentina','Argelia'),(date(2026,6,16),'J','Austria','Jordania'),
    (date(2026,6,22),'J','Argentina','Austria'),(date(2026,6,22),'J','Jordania','Argelia'),
    (date(2026,6,27),'J','Argelia','Austria'),(date(2026,6,27),'J','Jordania','Argentina'),
    (date(2026,6,17),'K','Portugal','RD Congo'),(date(2026,6,17),'K','Uzbequistao','Colombia'),
    (date(2026,6,23),'K','Portugal','Uzbequistao'),(date(2026,6,23),'K','Colombia','RD Congo'),
    (date(2026,6,27),'K','Colombia','Portugal'),(date(2026,6,27),'K','RD Congo','Uzbequistao'),
    (date(2026,6,17),'L','Inglaterra','Croacia'),(date(2026,6,17),'L','Gana','Panama'),
    (date(2026,6,23),'L','Inglaterra','Gana'),(date(2026,6,23),'L','Panama','Croacia'),
    (date(2026,6,27),'L','Panama','Inglaterra'),(date(2026,6,27),'L','Croacia','Gana'),
]
NG=len(GROUP_FIXTURES)
KO=[
    ('R32','M73',('r2','A'),('r2','B')),('R32','M74',('w1','E'),('t3',3)),
    ('R32','M75',('w1','F'),('r2','C')),('R32','M76',('w1','C'),('r2','F')),
    ('R32','M77',('w1','I'),('t3',5)),('R32','M78',('r2','E'),('r2','I')),
    ('R32','M79',('w1','A'),('t3',0)),('R32','M80',('w1','L'),('t3',7)),
    ('R32','M81',('w1','D'),('t3',2)),('R32','M82',('w1','G'),('t3',4)),
    ('R32','M83',('r2','K'),('r2','L')),('R32','M84',('w1','H'),('r2','J')),
    ('R32','M85',('w1','B'),('t3',1)),('R32','M86',('w1','J'),('r2','H')),
    ('R32','M87',('w1','K'),('t3',6)),('R32','M88',('r2','D'),('r2','G')),
    ('Oitavas','J89',('win',1),('win',4)),('Oitavas','J90',('win',0),('win',2)),
    ('Oitavas','J91',('win',3),('win',5)),('Oitavas','J92',('win',6),('win',7)),
    ('Oitavas','J93',('win',10),('win',11)),('Oitavas','J94',('win',8),('win',9)),
    ('Oitavas','J95',('win',13),('win',15)),('Oitavas','J96',('win',12),('win',14)),
    ('Quartas','J97',('win',16),('win',17)),('Quartas','J98',('win',20),('win',21)),
    ('Quartas','J99',('win',18),('win',19)),('Quartas','J100',('win',22),('win',23)),
    ('Semi','J101',('win',24),('win',25)),('Semi','J102',('win',26),('win',27)),
    ('3o Lugar','J103',('lose',28),('lose',29)),('Final','J104',('win',28),('win',29)),
]
NKO=len(KO)
KO_PTS={'R32':3,'Oitavas':5,'Quartas':8,'Semi':12,'3o Lugar':10,'Final':25}
MEDALS={1:'🥇',2:'🥈',3:'🥉'}
RND_ICO={'R32':'🔵','Oitavas':'🟢','Quartas':'🟡','Semi':'🔴','3o Lugar':'🟠','Final':'⭐'}

# constantes globais, ou logo antes da função compute_mc.
MC_N_SIMS   = 20000          # nº de simulações
MC_W_FIFA   = 0.5           # peso do ranking FIFA
MC_W_FORMA  = 0.25          # peso da forma na fase de grupos
MC_W_CROWD  = 0.20           # peso da "torcida" dos apostadores
MC_K        = 0.9           # inclinação da logística (quão decisiva é a força)
MC_SLOPE    = 1.00
MC_CLAMP    = 0.07           # piso/teto de probabilidade por jogo (zebra sempre possível)
MC_WIN_BONUS, MC_QUAL_COEF = 0.22, 0.60   # momentum: bônus por vitória e peso da qualidade do adversário
MC_KO_W = {'R32': 1.0, 'Oitavas': 1.25, 'Quartas': 1.6, 'Semi': 2.0, 'Final': 2.4}
MC_DEPTH = {'R32': 1, 'Oitavas': 2, 'Quartas': 3, 'Semi': 4, '3o Lugar': 4, 'Final': 5}
MC_FORM_MARGIN = 3.0    # saturação do placar (3 gols já é goleada)
MC_FORM_OPP_K  = 0.70   # peso da força do adversário (ranking FIFA)
_MC_ROUNDS = ['R32', 'Oitavas', 'Quartas', 'Semi', '3o Lugar', 'Final']
# Árvore do chaveamento (extraída do seu build_bracket; validada 100%)
_MC_OCT = [(1, 4), (0, 2), (3, 5), (6, 7), (10, 11), (8, 9), (13, 15), (12, 14)]
_MC_QF  = [(16, 17), (20, 21), (18, 19), (22, 23)]
_MC_SF  = [(24, 25), (26, 27)]
_MC_KOR_RND = {**{m: 'R32' for m in range(16)}, **{m: 'Oitavas' for m in range(16, 24)},
               **{m: 'Quartas' for m in range(24, 28)}, 28: 'Semi', 29: 'Semi',
               30: '3o Lugar', 31: 'Final'}


# ══════════════════════════════════════════════════════════════════════
# PONTUAÇÃO
# ══════════════════════════════════════════════════════════════════════
def sg(b1,b2,r1,r2):
    try: b1,b2,r1,r2=int(b1),int(b2),int(r1),int(r2)
    except: return 0
    if b1==r1 and b2==r2: return 5
    br=(b1>b2)-(b1<b2); rr=(r1>r2)-(r1<r2)
    if br==rr: return 3 if (b1-b2)==(r1-r2) else 2
    return 0

def calc_st(data):
    st={g:{t:{'pts':0,'played':0,'w':0,'d':0,'l':0,'gf':0,'ga':0} for t in ts}
        for g,ts in GROUPS_DATA.items()}
    for m,(_,g,t1,t2) in enumerate(GROUP_FIXTURES):
        if m not in data or data[m] is None: continue
        s1,s2=data[m]
        if s1 is None or s2 is None: continue
        try: s1,s2=int(s1),int(s2)
        except: continue
        for t,gf,ga in [(t1,s1,s2),(t2,s2,s1)]:
            st[g][t]['played']+=1; st[g][t]['gf']+=gf; st[g][t]['ga']+=ga
        if s1>s2:   st[g][t1]['pts']+=3;st[g][t1]['w']+=1;st[g][t2]['l']+=1
        elif s2>s1: st[g][t2]['pts']+=3;st[g][t2]['w']+=1;st[g][t1]['l']+=1
        else:
            st[g][t1]['pts']+=1;st[g][t1]['d']+=1
            st[g][t2]['pts']+=1;st[g][t2]['d']+=1
    return st

def sort_st(st):
    return {g:sorted(d.items(),key=lambda x:(-x[1]['pts'],-(x[1]['gf']-x[1]['ga']),-x[1]['gf'],FIFA_RANKINGS.get(x[0],150)))
            for g,d in st.items()}

def build_r32(ss,t495):
    g1={g:l[0][0] for g,l in ss.items() if l}
    g2={g:l[1][0] for g,l in ss.items() if len(l)>1}
    g3={g:l[2][0] for g,l in ss.items() if len(l)>2}
    thirds=sorted(
        [(g,-d['pts'],-(d['gf']-d['ga']),-d['gf'],FIFA_RANKINGS.get(g3.get(g,''),150))
         for g,l in ss.items() if len(l)>2 for _,d in [l[2]]],key=lambda x:x[1:])
    top8=[t[0] for t in thirds[:8]]; key=''.join(sorted(top8))
    asgn=t495.get(key,['?']*8)
    def gt3(i):
        a=str(asgn[i]) if i<len(asgn) else '?'
        return g3.get(a[1],'?') if len(a)>=2 else '?'
    def res(sp):
        if sp[0]=='w1': return g1.get(sp[1],'?')
        if sp[0]=='r2': return g2.get(sp[1],'?')
        if sp[0]=='t3': return gt3(sp[1])
        return '?'
    return {m:(res(KO[m][2]),res(KO[m][3])) for m in range(16)},g1,g2,g3,key

def build_bracket(r32,kor):
    tn=dict(r32); w={}
    def wof(m):
        if m not in kor: return None
        t1,t2=tn.get(m,(None,None)); a,b,pen=kor[m]
        if a is None or b is None: return None
        if a>b: return t1
        if b>a: return t2
        return t1 if pen=='S1' else (t2 if pen=='S2' else None)
    def lof(m):
        t1,t2=tn.get(m,(None,None)); wm=wof(m)
        return t2 if wm==t1 else (t1 if wm==t2 else None)
    for m in range(16): w[m]=wof(m)
    for i,(a,b) in enumerate([(1,4),(0,2),(3,5),(6,7),(10,11),(8,9),(13,15),(12,14)]):
        m=16+i; tn[m]=(w.get(a),w.get(b)); w[m]=wof(m)
    for i,(a,b) in enumerate([(16,17),(20,21),(18,19),(22,23)]):
        m=24+i; tn[m]=(w.get(a),w.get(b)); w[m]=wof(m)
    for i,(a,b) in enumerate([(24,25),(26,27)]):
        m=28+i; tn[m]=(w.get(a),w.get(b)); w[m]=wof(m)
    tn[30]=(lof(28),lof(29)); w[30]=wof(30)
    tn[31]=(w.get(28),w.get(29)); w[31]=wof(31)
    return w,tn

def sim_bet(gb,xm,t495):
    r32,*_=build_r32(sort_st(calc_st(gb)),t495)
    picks={}; bn=dict(r32)
    for m in range(16):
        t1,t2=r32.get(m,('?','?')); x=xm.get(m)
        if x=='t1': picks[m]=t1
        elif x=='t2': picks[m]=t2
    def nr(struct,base):
        for i,(p1,p2) in enumerate(struct):
            m=base+i; t1=picks.get(p1,'?'); t2=picks.get(p2,'?')
            bn[m]=(t1,t2); x=xm.get(m)
            if x=='t1': picks[m]=t1
            elif x=='t2': picks[m]=t2
    nr([(1,4),(0,2),(3,5),(6,7),(10,11),(8,9),(13,15),(12,14)],16)
    nr([(16,17),(20,21),(18,19),(22,23)],24)
    nr([(24,25),(26,27)],28)
    for mi in [30,31]:
        def gh(sp):
            ac,src=sp; t1,t2=bn.get(src,('?','?')); wp=picks.get(src,'?')
            return wp if ac=='win' else (t2 if wp==t1 else t1 if wp==t2 else '?')
        t1=gh(KO[mi][2]); t2=gh(KO[mi][3]); bn[mi]=(t1,t2)
        x=xm.get(mi)
        if x=='t1': picks[mi]=t1
        elif x=='t2': picks[mi]=t2
    return picks,bn

def picks_by_round(gb, xm, t495_):
    bp, _ = sim_bet(gb, xm, t495_)
    out: dict = {}
    for m, (rnd, *_) in enumerate(KO):
        p = bp.get(m)
        if p and p != "?":
            out.setdefault(rnd, set()).add(p)
    return out

def score_all(gb,xm,bb,gr,mmr,br,t495):
    gdet={}; gt=0
    for m,(r1,r2) in gr.items():
        bv=gb.get(m)
        if bv is None: gdet[m]=None; continue
        p=sg(bv[0],bv[1],r1,r2); gdet[m]=p; gt+=p
    ap=mp=0
    if br and bb:
        ar,mr_=br; ab,mb=bb
        if ab and ar and str(ab).strip().lower()==str(ar).strip().lower(): ap=10
        if mb and mr_ and str(mb).strip().lower()==str(mr_).strip().lower(): mp=5
    rw,rn={},{}
    if gr:
        r32r,*_=build_r32(sort_st(calc_st(gr)),t495)
        rw,rn=build_bracket(r32r,mmr)
    bp,_=sim_bet(gb,xm,t495)

    # Vencedores reais agrupados por rodada
    # Lógica v7c: pontua se a seleção escolhida avançou em QUALQUER
    # jogo daquela rodada, independente do confronto específico previsto.
    _rnd_midxs: dict = {}
    for _m, (_rnd, *_rest) in enumerate(KO):
        _rnd_midxs.setdefault(_rnd, []).append(_m)
    _round_winners: dict = {
        _rnd: {rw.get(_m) for _m in _midxs} - {None}
        for _rnd, _midxs in _rnd_midxs.items()
    }

    mdet={}; mt=0
    for m in range(NKO):
        rnd  = KO[m][0]
        pv   = KO_PTS[rnd]
        pick = bp.get(m)
        winners_this_rnd = _round_winners.get(rnd, set())
        if not winners_this_rnd:
            # Nenhum jogo da rodada finalizado → aguardando
            mdet[m] = None
        elif pick and pick != '?' and pick in winners_this_rnd:
            # Seleção escolhida avançou em qualquer jogo da rodada ✓
            mdet[m] = pv; mt += pv
        else:
            # Rodada iniciada mas seleção não avançou (ou não escolheu)
            mdet[m] = 0
    return dict(total=gt+ap+mp+mt,grupos=gt,bonus=ap+mp,mm=mt,
                art_pts=ap,mg_pts=mp,gdet=gdet,mdet=mdet)

def _safe_int(v):
    try:
        return int(v or 0)
    except Exception:
        return 0

def _acertos_5pts_grupos(sc):
    return sum(1 for p in (sc.get("gdet") or {}).values() if p == 5)

def _nome_alfabetico(nome):
    nome = str(nome or "").strip().casefold()
    nome = unicodedata.normalize("NFKD", nome)
    nome = "".join(ch for ch in nome if not unicodedata.combining(ch))
    return nome

def ranking_score_key(nome, sc):
    return (
        -_safe_int(sc.get("total")),       # Pontuação total
        -_safe_int(sc.get("mm")),          # 1º desempate: Mata-Mata
        -_acertos_5pts_grupos(sc),         # 2º desempate: acertos de 5 pts nos grupos
        -_safe_int(sc.get("grupos")),      # 3º desempate: pontos nos grupos
        _nome_alfabetico(nome),            # 4º desempate: ordem alfabética
    )

def ranking_bettor_key(bettor):
    nm = bettor[0]
    sc = bettor[4]
    return ranking_score_key(nm, sc)

# ══════════════════════════════════════════════════════════════════════
# FILE LOADING
# ══════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_t495(path):
    try:
        wb=load_workbook(path,data_only=True,read_only=True)
        if 'T495' not in wb.sheetnames: return {}
        ws=wb['T495']; t={}
        for row in ws.iter_rows(min_row=3,values_only=True):
            if row[0] and isinstance(row[0],str) and len(row[0])==8:
                t[row[0]]=[str(v) if v else '?' for v in row[1:9]]
        return t
    except: return {}

@st.cache_data(show_spinner=False)
def load_gab(path):
    try:
        wb=load_workbook(path,data_only=True,read_only=True); sh=wb.sheetnames
        gr={}
        if 'Jogos - Grupos' in sh:
            ws=wb['Jogos - Grupos']
            for m in range(72):
                r1=ws.cell(row=3+m,column=6).value; r2=ws.cell(row=3+m,column=7).value
                if r1 is not None and r2 is not None:
                    try: gr[m]=(int(r1),int(r2))
                    except: pass
        mmr={}
        if 'Mata-Mata - Jogos' in sh:
            ws=wb['Mata-Mata - Jogos']
            for m in range(32):
                g1v=ws.cell(row=4+m,column=6).value; g2v=ws.cell(row=4+m,column=7).value
                pen=ws.cell(row=4+m,column=8).value
                if g1v is not None and g2v is not None:
                    try: mmr[m]=(int(g1v),int(g2v),pen)
                    except: pass
        br=None
        if 'Apostas - Bonus' in sh:
            ws=wb['Apostas - Bonus']
            a=ws.cell(row=4,column=2).value; mg=ws.cell(row=4,column=3).value
            if a or mg: br=(a,mg)
        return gr,mmr,br
    except Exception as e:
        st.error(f"Erro gabarito: {e}"); return {},{},None

@st.cache_data(show_spinner=False)
def load_part(path):
    try:
        wb=load_workbook(path,data_only=True,read_only=True); sh=wb.sheetnames
        gb={}
        if 'Apostas - Grupos' in sh:
            ws=wb['Apostas - Grupos']
            for m in range(72):
                v1=ws.cell(row=5,column=2+m*2).value; v2=ws.cell(row=5,column=3+m*2).value
                if v1 is not None and v2 is not None:
                    try: gb[m]=(int(v1),int(v2))
                    except: pass
        bb=(None,None)
        if 'Apostas - Bonus' in sh:
            ws=wb['Apostas - Bonus']
            bb=(ws.cell(row=5,column=2).value,ws.cell(row=5,column=3).value)
        xm={}
        if 'Apostas - Mata-Mata' in sh:
            ws=wb['Apostas - Mata-Mata']
            for m in range(32):
                v1=ws.cell(row=6,column=2+m*2).value; v2=ws.cell(row=6,column=3+m*2).value
                if v1 and str(v1).strip().upper()=='X': xm[m]='t1'
                elif v2 and str(v2).strip().upper()=='X': xm[m]='t2'
        return gb,bb,xm
    except Exception:
        return {},(None,None),{}
    
@st.cache_data(show_spinner=False)
def detect():
    gab_files = sorted(GABARITO_DIR.glob("Bolao_Copa2026_TurimMFO_*.xlsx"))
    part_files = sorted(APOSTAS_DIR.glob("Bolao_Copa2026_TurimMFO_*.xlsx"))

    gabs = []
    parts = []

    for f in gab_files:
        base = f.name
        nm = re.sub(r'(?i)^Bolao_Copa2026_TurimMFO_', '', base).replace('.xlsx', '').replace('_', ' ').strip()
        gabs.append((nm, str(f)))

    for f in part_files:
        base = f.name
        nm = re.sub(r'(?i)^Bolao_Copa2026_TurimMFO_', '', base).replace('.xlsx', '').replace('_', ' ').strip()
        parts.append((nm, str(f)))

    return gabs, parts

@st.cache_data(show_spinner=False)
def load_consolidada(path):
    """Lê o arquivo consolidado e retorna {nm: {gb, bb, xm}}."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sh = wb.sheetnames
    except Exception as e:
        st.error(f"Erro ao ler arquivo consolidado: {e}")
        return {}

    result = {}

    if "Apostas - Grupos" in sh:
        ws = wb["Apostas - Grupos"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            nm = str(row[0])
            gb = {}
            for m in range(72):
                ci = 1 + m * 2
                v1 = row[ci] if ci < len(row) else None
                v2 = row[ci + 1] if ci + 1 < len(row) else None
                if v1 is not None and v2 is not None:
                    try: gb[m] = (int(v1), int(v2))
                    except: pass
            result.setdefault(nm, {"gb": {}, "bb": (None, None), "xm": {}})["gb"] = gb

    if "Apostas - Bonus" in sh:
        ws = wb["Apostas - Bonus"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            nm = str(row[0])
            result.setdefault(nm, {"gb": {}, "bb": (None, None), "xm": {}})["bb"] = (
                row[1] if len(row) > 1 else None,
                row[2] if len(row) > 2 else None,
            )

    if "Apostas - Mata-Mata" in sh:
        ws = wb["Apostas - Mata-Mata"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            nm = str(row[0])
            xm = {}
            for m in range(32):
                ci = 1 + m * 2
                v1 = row[ci] if ci < len(row) else None
                v2 = row[ci + 1] if ci + 1 < len(row) else None
                if v1 and str(v1).strip().upper() == "X": xm[m] = "t1"
                elif v2 and str(v2).strip().upper() == "X": xm[m] = "t2"
            result.setdefault(nm, {"gb": {}, "bb": (None, None), "xm": {}})["xm"] = xm

    return result

@st.cache_resource(show_spinner=False)  
def load_all_data_consolidated(gab_path, consol_path):
    """Versão rápida: lê uma única planilha consolidada."""
    t495      = load_t495(gab_path)
    gr,mmr,br = load_gab(gab_path)
    consol    = load_consolidada(consol_path)
    bettors   = []
    for nm, d in consol.items():
        gb, bb, xm = d["gb"], d["bb"], d["xm"]
        sc   = score_all(gb, xm, bb, gr, mmr, br, t495)
        prnd = picks_by_round(gb, xm, t495)
        bettors.append((nm, gb, bb, xm, sc, prnd))
    bettors.sort(key=ranking_bettor_key)
    return t495, gr, mmr, br, bettors

@st.cache_data(show_spinner=False)
def load_all_data(gab_path, parts_tuple):
    t495      = load_t495(gab_path)
    gr,mmr,br = load_gab(gab_path)

    def _load_one(item):
        nm, fp = item
        gb, bb, xm = load_part(fp)
        sc   = score_all(gb, xm, bb, gr, mmr, br, t495)
        prnd = picks_by_round(gb, xm, t495)
        return (nm, gb, bb, xm, sc, prnd)

    n_workers = min(16, len(parts_tuple) or 1)
    with ThreadPoolExecutor(max_workers=n_workers) as ex:
        bettors = list(ex.map(_load_one, parts_tuple))

    bettors.sort(key=ranking_bettor_key)
    return t495, gr, mmr, br, bettors

@st.cache_resource(show_spinner=False)
def compute_mc(fingerprint, _bettors, n_sims=MC_N_SIMS, seed=42):
    """Monte Carlo do mata-mata. Cacheado por fingerprint (recomputa quando
    os dados mudam). Respeita o gabarito: R32 vem do build_r32 real e jogos já
    disputados (mmr) ficam travados; só os jogos que faltam são sorteados."""
    import math
    import numpy as np
    import random as _random
 
    _ss = sort_st(calc_st(gr))
    r32, _g1, _g2, _g3, _key = build_r32(_ss, t495)
    w_real, tn_real = build_bracket(r32, mmr)
    teams = sorted({t for m in range(16) for t in r32[m] if t and t != '?'})
    if len(teams) < 2:
        return None
 
    # forma na fase de grupos AJUSTADA pela força do adversário (ranking FIFA):
    # cada jogo conta o quanto a seleção rendeu ACIMA do esperado contra aquele rival.
    def _f_fifa(x):
        return -math.log(FIFA_RANKINGS.get(x, 150))
    form_acc = {t: 0.0 for t in teams}; form_n = {t: 0 for t in teams}
    for _m, sgr in gr.items():
        if _m >= len(GROUP_FIXTURES):
            continue
        _dt, _grp, t1, t2 = GROUP_FIXTURES[_m]
        g1v, g2v = sgr[0], sgr[1]
        for _me, _opp, _gfor, _gag in ((t1, t2, g1v, g2v), (t2, t1, g2v, g1v)):
            if _me not in form_acc:
                continue
            _dom = 0.5 + 0.5 * math.tanh((_gfor - _gag) / MC_FORM_MARGIN)              # dominância no jogo [0..1]
            _exp = 1.0 / (1.0 + math.exp(-MC_FORM_OPP_K * (_f_fifa(_me) - _f_fifa(_opp))))  # esperado p/ esse rival
            form_acc[_me] += (_dom - _exp)                                            # acima/abaixo do esperado
            form_n[_me] += 1
    form_raw = {t: (form_acc[t] / form_n[t]) if form_n[t] else 0.0 for t in teams}
 
    nb = len(_bettors)
    crowd = {t: 0.0 for t in teams}
    for pos, bt in enumerate(_bettors, 1):
        w = 0.5 + 0.5 * (nb - pos) / max(1, nb - 1)
        best = {}
        for r in _MC_ROUNDS:
            d = MC_DEPTH[r]
            for t in bt[5].get(r, ()):
                if t in crowd and d > best.get(t, 0):
                    best[t] = d
        for t, d in best.items():
            crowd[t] += w * d
 
    def _z(d):
        vals = [v for v in d.values() if v is not None]
        mu = sum(vals) / len(vals)
        sd = (sum((v - mu) ** 2 for v in vals) / len(vals)) ** 0.5 or 1.0
        return {t: (d[t] - mu) / sd for t in d}
 
    zf = _z({t: -math.log(FIFA_RANKINGS.get(t, 150)) for t in teams})
    zg = _z(form_raw)
    zc = _z(crowd)
    S_base = _z({t: MC_W_FIFA * zf[t] + MC_W_FORMA * zg[t] + MC_W_CROWD * zc[t] for t in teams})
 
    # Boost CONSISTENTE: toda vitória (real OU simulada) dá boost ao vencedor nas
    # rodadas seguintes, escalado pela força-base do adversário batido. Assim, time
    # que passou de verdade e time que passou na simulação chegam à rodada seguinte
    # em pé de igualdade (corrige a assimetria do estado misto). É aplicado dentro
    # do laço, por simulação, no dicionário `bo`.
    def _wp(a, b, bo):
        d = (S_base.get(a, 0.0) + bo.get(a, 0.0)) - (S_base.get(b, 0.0) + bo.get(b, 0.0))
        p = 1.0 / (1.0 + math.exp(-MC_SLOPE * d))
        return min(1 - MC_CLAMP, max(MC_CLAMP, p))
 
    def _resolve(m, t1, t2, rnd, bo):
        w = locked[m] if m in locked else (t1 if rr.random() < _wp(t1, t2, bo) else t2)
        if m < 30:                       # vencedor ainda terá próxima rodada -> ganha boost
            l = t2 if w == t1 else t1
            bo[w] = bo.get(w, 0.0) + MC_WIN_BONUS * MC_KO_W[rnd] * (1 + MC_QUAL_COEF * max(0.0, S_base.get(l, 0.0)))
        return w
 
    locked = {m: w_real[m] for m in range(32) if m in mmr and w_real.get(m)}
 
    idx = {t: i for i, t in enumerate(teams)}; T = len(idx)
    Pmat = {r: np.zeros((nb, T)) for r in _MC_ROUNDS}
    for b, bt in enumerate(_bettors):
        for r in _MC_ROUNDS:
            for t in bt[5].get(r, ()):
                if t in idx:
                    Pmat[r][b, idx[t]] = 1.0
    fixed = np.array([bt[4].get('grupos', 0) + bt[4].get('bonus', 0) for bt in _bettors], dtype=float)
 
    rr = _random.Random(seed)
    totals = np.zeros((n_sims, nb))
    champ = {}; finalist = {}
    reach = {k: {} for k in ['Oitavas', 'Quartas', 'Semi']}
    third = {}
    mw = {m: {} for m in range(32)}
    for s in range(n_sims):
        win = {}; pair = {}; bo = {}
        for m in range(16):
            t1, t2 = r32[m]; pair[m] = (t1, t2)
            win[m] = _resolve(m, t1, t2, 'R32', bo)
        for i, (a, b) in enumerate(_MC_OCT):
            m = 16 + i; t1, t2 = win[a], win[b]; pair[m] = (t1, t2)
            win[m] = _resolve(m, t1, t2, 'Oitavas', bo)
        for i, (a, b) in enumerate(_MC_QF):
            m = 24 + i; t1, t2 = win[a], win[b]; pair[m] = (t1, t2)
            win[m] = _resolve(m, t1, t2, 'Quartas', bo)
        for i, (a, b) in enumerate(_MC_SF):
            m = 28 + i; t1, t2 = win[a], win[b]; pair[m] = (t1, t2)
            win[m] = _resolve(m, t1, t2, 'Semi', bo)
        l28 = pair[28][0] if win[28] == pair[28][1] else pair[28][1]
        l29 = pair[29][0] if win[29] == pair[29][1] else pair[29][1]
        pair[30] = (l28, l29)
        win[30] = _resolve(30, l28, l29, 'Final', bo)
        t1, t2 = win[28], win[29]; pair[31] = (t1, t2)
        win[31] = _resolve(31, t1, t2, 'Final', bo)
 
        adv = {'R32': [win[m] for m in range(16)],
               'Oitavas': [win[m] for m in range(16, 24)],
               'Quartas': [win[m] for m in range(24, 28)],
               'Semi': [win[28], win[29]],
               'Final': [win[31]], '3o Lugar': [win[30]]}
        tot = fixed.copy()
        for r in _MC_ROUNDS:
            ii = [idx[t] for t in adv[r] if t in idx]
            if ii:
                tot = tot + KO_PTS[r] * Pmat[r][:, ii].sum(axis=1)
        totals[s] = tot
 
        champ[win[31]] = champ.get(win[31], 0) + 1
        third[win[30]] = third.get(win[30], 0) + 1
        finalist[win[28]] = finalist.get(win[28], 0) + 1
        finalist[win[29]] = finalist.get(win[29], 0) + 1
        for t in adv['Oitavas']: reach['Oitavas'][t] = reach['Oitavas'].get(t, 0) + 1
        for t in adv['Quartas']: reach['Quartas'][t] = reach['Quartas'].get(t, 0) + 1
        for t in adv['Semi']:    reach['Semi'][t] = reach['Semi'].get(t, 0) + 1
        for m in range(32):
            mw[m][win[m]] = mw[m].get(win[m], 0) + 1
 
    order = np.argsort(-totals, axis=1)
    win_count = np.bincount(order[:, 0], minlength=nb)
    pot = nb * 50.0; prizes = [0.70 * pot, 0.20 * pot, 0.10 * pot]
    prize = np.zeros(nb); podium = np.zeros(nb)
    for k in range(min(3, nb)):
        col = order[:, k]
        prize += np.bincount(col, minlength=nb) * prizes[k]
        podium += np.bincount(col, minlength=nb)
    p10, p50, p90 = np.percentile(totals, [10, 50, 90], axis=0)
    mean = totals.mean(axis=0)
 
    real_adv = {'R32': {w_real[m] for m in range(16) if w_real.get(m)},
                'Oitavas': {w_real[m] for m in range(16, 24) if w_real.get(m)},
                'Quartas': {w_real[m] for m in range(24, 28) if w_real.get(m)},
                'Semi': {w_real[m] for m in (28, 29) if w_real.get(m)},
                'Final': {w_real[31]} if w_real.get(31) else set(),
                '3o Lugar': {w_real[30]} if w_real.get(30) else set()}
    mm_ok = True
    for bt in _bettors:
        mine = sum(KO_PTS[r] * len(set(bt[5].get(r, set())) & real_adv[r]) for r in _MC_ROUNDS)
        if mine != bt[4].get('mm', mine):
            mm_ok = False
            break
 
    return dict(
        n=n_sims, teams=teams, S=S_base, r32=r32, locked=locked, w_real=w_real, tn_real=tn_real,
        champ=champ, finalist=finalist, reach=reach, third=third, mw=mw,
        names=[b[0] for b in _bettors],
        win_pct=(win_count / n_sims), podium_pct=(podium / n_sims),
        exp_pts=mean, p10=p10, p90=p90, exp_prize=(prize / n_sims),
        mm_ok=mm_ok,
    )
 
@st.cache_resource(show_spinner=False)
def compute_stats(fingerprint, _bettors):
    """Agrega consenso do bolão UMA vez (cacheado por fingerprint).
    `_bettors` tem prefixo _ para o Streamlit não tentar fazer hash dele;
    a chave de cache é só o `fingerprint` (string)."""
    n = len(_bettors)
 
    # ── Consenso de placares (fase de grupos) ──
    group_dist: dict = {}                 # m -> Counter{(g1,g2): qtd}
    for _nm, _gb, *_ in _bettors:
        for _m, _bv in _gb.items():
            if _bv is None:
                continue
            try:
                _key = (int(_bv[0]), int(_bv[1]))
            except Exception:
                continue
            group_dist.setdefault(_m, Counter())[_key] += 1
 
    # Conformidade = média da popularidade dos placares de cada apostador
    conformity: dict = {}
    for _nm, _gb, *_ in _bettors:
        _pops = []
        for _m, _bv in _gb.items():
            _c = group_dist.get(_m)
            if not _c:
                continue
            try:
                _key = (int(_bv[0]), int(_bv[1]))
            except Exception:
                continue
            _tot = sum(_c.values())
            if _tot:
                _pops.append(_c.get(_key, 0) / _tot)
        conformity[_nm] = (sum(_pops) / len(_pops)) if _pops else 0.0
 
    # Ranking de ousadia: menor conformidade = mais ousado = 1º
    _order = sorted(conformity.items(), key=lambda x: x[1])
    conformity_rank = {nm: (i + 1, n) for i, (nm, _v) in enumerate(_order)}
 
    # ── Consenso do mata-mata por fase ──
    rnd_slots: dict = {}
    for _m, (_rnd, *_r) in enumerate(KO):
        rnd_slots[_rnd] = rnd_slots.get(_rnd, 0) + 1
    ko_dist = {rnd: Counter() for rnd in rnd_slots}
    for _nm, _gb, _bb, _xm, _sc, _prnd in _bettors:
        for _rnd, _teams in _prnd.items():
            for _t in _teams:
                if _t and _t != '?':
                    ko_dist[_rnd][_t] += 1
 
    # ── Bônus: distribuições ──
    art_dist = Counter()
    mg_dist = Counter()
    champ_dist = Counter()
    for _nm, _gb, _bb, _xm, _sc, _prnd in _bettors:
        _a = _bb[0] if _bb else None
        _mg = _bb[1] if _bb else None
        if _a:
            art_dist[str(_a).strip()] += 1
        if _mg:
            mg_dist[str(_mg).strip()] += 1
        for _t in _prnd.get('Final', set()):
            if _t and _t != '?':
                champ_dist[_t] += 1
 
    return dict(
        n=n, group_dist=group_dist, conformity=conformity,
        conformity_rank=conformity_rank, rnd_slots=rnd_slots, ko_dist=ko_dist,
        art_dist=art_dist, mg_dist=mg_dist, champ_dist=champ_dist,
    )
# ══════════════════════════════════════════════════════════════════════
# SIDEBAR  — nativa, sem CSS override
# ══════════════════════════════════════════════════════════════════════
with st.sidebar:
    # Logos centrados
    _, lc, _ = st.columns([1,3,1])
    with lc:
        if LOGO_TURIM_AZUL: st.image(LOGO_TURIM_AZUL, width=130)
        else: st.markdown("### 🐂 TURIM")
    _, lc2, _ = st.columns([1,5,1])
    with lc2:
        if LOGO_TORI: st.image(LOGO_TORI, width=160)
    st.markdown("---")
    st.markdown("### ⚙️ Configuração")
    st.caption(f"📋 Gabarito: `{GABARITO_DIR}`")
    st.caption(f"👥 Apostas: `{APOSTAS_DIR}`")

    gabs, parts = detect()

    st.markdown("---")
    if not gabs:
        st.warning("⚠️ Sem gabarito.\nNomeie:\n`Bolao_Copa2026_TurimMFO_Master.xlsx`")
        gab_path = None
    else:
        gsel = st.selectbox("📋 Gabarito", [n for n,_ in gabs])
        gab_path = next(p for n,p in gabs if n==gsel)
        st.success(f"✅ {gsel}")

    st.markdown("---")
    if CONSOLIDADA_PATH.exists():
        st.success("📦 Modo consolidado ativo")
        st.caption(f"`{CONSOLIDADA_PATH.name}`")
        st.caption("Execute `consolidar_apostas.py` para atualizar.")
    elif parts:
        st.markdown(f"**👥 {len(parts)} participante(s):**")
        for nm,_ in parts:
            st.markdown(f"  · {nm}")
    else:
        st.warning("⚠️ Sem participantes.\nGere a planilha consolidada ou coloque apostas em `apostas/`.")

    st.markdown("---")
    if st.button("🔄 Recarregar dados", width='stretch'):
        st.cache_data.clear()
        st.cache_resource.clear()
        _CACHE_PATH.unlink(missing_ok=True)
        st.rerun()

    # Mascotes no rodapé da sidebar
    if MASCOTES:
        st.markdown("---")
        st.image(MASCOTES, caption="🐂 Rumo ao Hexa! 🏆", width='stretch')

    st.markdown("---")
    st.caption("Bolão Copa 2026 · Turim MFO · v5.0")

# ══════════════════════════════════════════════════════════════════════
# HERO HEADER
# ══════════════════════════════════════════════════════════════════════
_fav_b64 = img_to_b64(FAVICON) if FAVICON else ""
_hero_icon = (
    f'<img src="data:image/png;base64,{_fav_b64}" style="width:6rem;height:6rem;object-fit:contain">'
    if _fav_b64 else '<span style="font-size:2.5rem">⚽</span>'
)
st.markdown(f"""
<div class="hero">
  <div style="display:flex;align-items:center;gap:16px">
    <div>{_hero_icon}</div>
    <div>
      <div class="hero-title">Bolão Copa do Mundo 2026</div>
      <div class="hero-sub">Dashboard Oficial · Turim MFO</div>
      <div class="hero-badge">🇨🇦 Canadá &nbsp;·&nbsp; 🇲🇽 México &nbsp;·&nbsp; 🇺🇸 EUA &nbsp;·&nbsp; Junho–Julho 2026</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if not gab_path or (not CONSOLIDADA_PATH.exists() and not parts):
    st.info("👈 Configure a pasta na barra lateral. Execute `consolidar_apostas.py` para gerar a planilha consolidada.")
    st.stop()

# ── Carregar dados — usa arquivo consolidado (1 arquivo) se disponível
if CONSOLIDADA_PATH.exists():
    _all_paths = [gab_path, str(CONSOLIDADA_PATH)]
else:
    _all_paths = [gab_path] + [fp for _, fp in parts]

_fp = _compute_fingerprint(_all_paths)
_disk_data = _disk_cache_get(_fp)
if _disk_data is not None:
    t495, gr, mmr, br, bettors = _disk_data
else:
    with st.spinner("Carregando dados..."):
        if CONSOLIDADA_PATH.exists():
            t495, gr, mmr, br, bettors = load_all_data_consolidated(gab_path, str(CONSOLIDADA_PATH))
        else:
            t495, gr, mmr, br, bettors = load_all_data(gab_path, tuple(parts))
    try:
        with open(_CACHE_PATH, "wb") as _cf:
            pickle.dump({"fp": _fp, "data": (t495, gr, mmr, br, bettors)}, _cf,
                        protocol=pickle.HIGHEST_PROTOCOL)
    except Exception:
        pass

maxp = max((b[4]['total'] for b in bettors), default=1) or 1
stats = compute_stats(_fp, bettors)

rw,rn = {},{}
if gr:
    r32r,*_ = build_r32(sort_st(calc_st(gr)), t495)
    rw,rn   = build_bracket(r32r, mmr)

rnd_ms = OrderedDict()
for m,(rnd,*_) in enumerate(KO): rnd_ms.setdefault(rnd,[]).append(m)

# ── Métricas
c5 = st.columns(5)
for col,(val,lbl) in zip(c5,[
    (f"{len(gr)}/72","Jogos Grupos"), (f"{len(mmr)}/32","Jogos MM"),
    (str(sum(a+b for a,b in gr.values())),"Gols"),
    (str(bettors[0][4]['total']) if bettors else "0","Líder (pts)"),
    (str(len(bettors)),"Participantes"),
]):
    col.markdown(
        f'<div class="mc"><div class="mc-v">{val}</div>'
        f'<div class="mc-l">{lbl}</div></div>',
        unsafe_allow_html=True)

st.markdown("")

# ══════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════
_tab_labels = ["🏆 Ranking","⚽ Grupos","🗓️ Mata-Mata","👤 Por Apostador"]
if MOSTRAR_SIMULACAO:
    _tab_labels.append("🔮 Simulação")
_tabs = st.tabs(_tab_labels)
T1, T2, T3, T4 = _tabs[:4]
T5 = _tabs[4] if MOSTRAR_SIMULACAO else None

# ── TAB 1: RANKING ───────────────────────────────────────────────────
with T1:
    n_parts    = len(bettors)
    prize_pool = n_parts * 50
    prize_pct  = {1: .70, 2: .20, 3: .10}

    # ── Filtro de participantes
    _all_rk_names = [b[0] for b in bettors]
    _rc1, _rc2 = st.columns([1, 5])
    with _rc1:
        if st.button("✖ Limpar", width='stretch', key="rk_clear_ms"):
            st.session_state["rk_ms_sel"] = []
            st.rerun()
    with _rc2:
        if "rk_ms_sel" not in st.session_state:
            st.session_state["rk_ms_sel"] = []
        _rk_valid = [n for n in st.session_state["rk_ms_sel"] if n in _all_rk_names]
        _rk_chosen = st.multiselect(
            "Filtrar:",
            options=_all_rk_names,
            default=_rk_valid,
            key="rk_ms",
            placeholder="Selecione participantes para filtrar...",
            label_visibility="collapsed",
        )
        st.session_state["rk_ms_sel"] = _rk_chosen

    if "show_all_ranking" not in st.session_state:
        st.session_state["show_all_ranking"] = False
    _show_all = st.session_state["show_all_ranking"]

    if _rk_chosen:
        _display = [b for b in bettors if b[0] in set(_rk_chosen)]
        _use_expand = False
    else:
        _display  = bettors if _show_all else bettors[:10]
        _use_expand = True
    cA,cB = st.columns([3,2], gap="large")
    with cA:
        st.markdown('<div class="sh">🏆 Classificação Geral</div>', unsafe_allow_html=True)
        for pos,(nm,_,_,_,sc,_) in enumerate(_display,1):
            real_pos = next(i for i,(b,*_) in enumerate(bettors,1) if b==nm)
            cls   = {1:'rc1',2:'rc2',3:'rc3'}.get(real_pos,'rcN')
            medal = MEDALS.get(real_pos, f"{real_pos}°")
            pct   = int(sc['total']/maxp*100)
            prize_badge = ""
            if real_pos in prize_pct:
                val = prize_pool * prize_pct[real_pos]
                prize_badge = (f'<span style="background:rgba(214,184,100,.18);'
                    f'border:1px solid rgba(214,184,100,.4);border-radius:12px;'
                    f'padding:1px 8px;font-size:.65rem;font-weight:700;'
                    f'color:#D6B864;margin-left:6px">R$ {val:,.0f}</span>')
            st.markdown(f"""<div class="rc {cls}">
              <div style="font-size:1.35rem;width:32px;text-align:center;flex-shrink:0">{medal}</div>
              <div style="flex:1;min-width:0">
                <div class="rc-name">{nm}{prize_badge}</div>
                <div class="rc-sub">Grupos {sc['grupos']} · Bônus {sc['bonus']} · MM {sc['mm']}</div>
                <div class="bar-bg"><div class="bar-fg" style="width:{pct}%"></div></div>
              </div>
              <div style="text-align:right;flex-shrink:0">
                <div class="rc-pts">{sc['total']}</div>
                <div class="rc-pl">pts</div>
              </div>
            </div>""", unsafe_allow_html=True)

        if _use_expand:
            if not _show_all and len(bettors) > 10:
                if st.button(f"👇 Ver todos os {len(bettors)} participantes",
                                 width='stretch'):
                    st.session_state["show_all_ranking"] = True
                    st.rerun()
            elif _show_all and len(bettors) > 10:
                if st.button("👆 Mostrar apenas top 10",
                                 width='stretch'):
                    st.session_state["show_all_ranking"] = False
                    st.rerun()

    with cB:
        st.markdown('<div class="sh">📊 Pontuação por Fase</div>', unsafe_allow_html=True)
        fig = go.Figure()
        for lbl,vals,color in [
            ('Grupos', [b[4]['grupos'] for b in _display], '#123A56'),
            ('Bônus',  [b[4]['bonus']  for b in _display], '#0D8587'),
            ('MM',     [b[4]['mm']     for b in _display], '#B2584E'),
        ]:
            fig.add_trace(go.Bar(
                name=lbl, y=[b[0] for b in _display], x=vals, orientation='h',
                marker_color=color, text=vals, textposition='inside',
                insidetextanchor='middle', textfont=dict(size=10, color='#fff')))
        fig.update_layout(
            barmode='stack', height=max(300, len(_display) * 28), margin=dict(l=0,r=5,t=5,b=5),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            legend=dict(orientation='h', y=1.06, x=0, font=dict(size=10)),
            xaxis=dict(gridcolor='rgba(0,0,0,.08)', tickfont=dict(size=10)),
            yaxis=dict(tickfont=dict(size=10), autorange='reversed'))
        st.plotly_chart(fig, width='stretch', config={'displayModeBar':False})

        if bettors:
            nm0,_,_,_,sc0,_ = bettors[0]
            st.markdown(f"""<div class="mc" style="text-align:left;padding:14px 18px">
              <div style="font-size:.65rem;color:#5C5F62;text-transform:uppercase;letter-spacing:1px">Líder atual</div>
              <div style="font-size:1.35rem;font-weight:800;margin:3px 0">🥇 {nm0}</div>
              <div class="mc-v" style="font-size:1.9rem;text-align:left;color:#D6B864">{sc0['total']} pts</div>
              <div style="font-size:.73rem;color:#5C5F62;margin-top:3px">
                G {sc0['grupos']} · B {sc0['bonus']} · MM {sc0['mm']}</div>
            </div>""", unsafe_allow_html=True)

    # ── EVOLUÇÃO TEMPORAL ────────────────────────────────────────────────
    st.markdown('<div class="sh">📈 Evolução de Pontuação</div>', unsafe_allow_html=True)

    # ── KO dates (R32→Final) ──────────────────────────────────────────────
    KO_DATES = [
        # R32 — 28 Jun–3 Jul
        date(2026,6,28),date(2026,6,28),date(2026,6,28),date(2026,6,28),
        date(2026,6,29),date(2026,6,29),date(2026,6,29),date(2026,6,29),
        date(2026,6,30),date(2026,6,30),date(2026,7,1), date(2026,7,1),
        date(2026,7,2), date(2026,7,2), date(2026,7,3), date(2026,7,3),
        # Oitavas — 4–7 Jul
        date(2026,7,4), date(2026,7,4), date(2026,7,5), date(2026,7,5),
        date(2026,7,6), date(2026,7,6), date(2026,7,7), date(2026,7,7),
        # Quartas — 9–11 Jul
        date(2026,7,9), date(2026,7,10),date(2026,7,11),date(2026,7,11),
        # Semis — 14–15 Jul
        date(2026,7,14),date(2026,7,15),
        # 3o Lugar, Final
        date(2026,7,18),date(2026,7,19),
    ]

    # Phase zones for background highlighting
    PHASE_ZONES = [
        ("Grupos",   date(2026,6,11), date(2026,6,27), "rgba(18,58,86,.12)"),
        ("R32",      date(2026,6,28), date(2026,7,3),  "rgba(13,133,135,.12)"),
        ("Oitavas",  date(2026,7,4),  date(2026,7,7),  "rgba(214,184,100,.10)"),
        ("Quartas",  date(2026,7,9),  date(2026,7,11), "rgba(220,136,74,.12)"),
        ("Semis",    date(2026,7,14), date(2026,7,15), "rgba(178,88,78,.12)"),
        ("3°/Final", date(2026,7,18), date(2026,7,19), "rgba(107,56,143,.12)"),
    ]

    CHART_COLORS = [
        '#D6B864','#0D8587','#B2584E','#2563EB',
        '#DC884A','#7B3FA0','#22C55E','#F97316',
        '#EF4444','#8B5CF6','#06B6D4','#EC4899',
    ]

    # ── Build timeline: {bettor: {date: pts_that_day}} ───────────────────
    # Group matches → date from GROUP_FIXTURES
    # KO matches    → date from KO_DATES
    # Bonus         → injected at Copa kick-off date (Jun 11)
    COPA_START = date(2026, 6, 11)

    all_names = [b[0] for b in bettors]

    # Participant filter
    fc1, fc2 = st.columns([1, 5], vertical_alignment="bottom")
    with fc1:
        if st.button("👥 Todos", width='stretch'):
            st.session_state["sel_bettors"] = all_names
            st.session_state["ms_bettors"] = all_names
            st.rerun()

    with fc2:
        if "sel_bettors" not in st.session_state:
                    st.session_state["sel_bettors"] = []
        valid_sel = [n for n in st.session_state["sel_bettors"] if n in all_names]
        chosen = st.multiselect(
                    "Participantes:",
                    options=all_names,
                    default=valid_sel,
                    key="ms_bettors",
                    placeholder="Selecione participantes para comparar...",
                )
        st.session_state["sel_bettors"] = chosen

    active_bettors = [b for b in bettors if b[0] in st.session_state["sel_bettors"]]

    # Chart mode
    chart_mode = st.radio(
        "Modo:",
        ["📈 Evolução Acumulada", "📊 Pontuação por Fase"],
        horizontal=True,
        label_visibility="collapsed",
    )

    if chart_mode == "📈 Evolução Acumulada":

        # Build sorted list of all match dates (group + KO)
        all_match_dates = sorted(set(
            [gf[0] for gf in GROUP_FIXTURES] + KO_DATES
        ))

        fig_evo = go.Figure()

        # Phase zone fills with top-aligned label annotations
        ZONE_FONT_COLOR = "rgba(200,220,240,.70)"
        for ph_name, d0, d1, color in PHASE_ZONES:
            fig_evo.add_vrect(
                x0=str(d0), x1=str(d1),
                fillcolor=color, layer="below",
                line=dict(width=1, color=color, dash="dot"),
            )
            # Place label at mid-x of zone using add_annotation
            # We'll use xref="x" with the first date of each zone as anchor
            fig_evo.add_annotation(
                x=str(d0), xref="x", y=1.0, yref="paper",
                text=f"<b>{ph_name}</b>",
                showarrow=False, xanchor="left", yanchor="top",
                font=dict(size=9, color=ZONE_FONT_COLOR),
                bgcolor="rgba(0,0,0,0)", borderpad=3,
            )

        for idx, (nm, gb, _, xm, sc, _) in enumerate(active_bettors):
            # Points per date
            pts_by_date: dict = {
                date(2026, 7, 19): sc["art_pts"],   # Final — J104
                date(2026, 6, 27): sc["mg_pts"],    # Último dia dos grupos
            }
            for m, (d_, g, t1, t2) in enumerate(GROUP_FIXTURES):
                p = sc["gdet"].get(m)
                if p is not None and p > 0:
                    pts_by_date[d_] = pts_by_date.get(d_, 0) + p
            for m, ko_date in enumerate(KO_DATES):
                p = sc["mdet"].get(m)
                if p is not None and p > 0:
                    pts_by_date[ko_date] = pts_by_date.get(ko_date, 0) + p

            # Build cumulative series over all match dates
            dates_x, cumul_y = [], []
            running = 0
            for d in all_match_dates:
                running += pts_by_date.get(d, 0)
                dates_x.append(str(d))
                cumul_y.append(running)

            color = CHART_COLORS[idx % len(CHART_COLORS)]
            first = nm
            fig_evo.add_trace(go.Scatter(
                x=dates_x, y=cumul_y,
                mode="lines+markers",
                name=first,
                line=dict(color=color, width=2.5),
                marker=dict(size=6, color=color,
                            line=dict(color="white", width=1.5)),
                hovertemplate=f"<b>{nm}</b><br>%{{x}}: <b>%{{y}} pts</b><extra></extra>",
            ))

        fig_evo.update_layout(
            height=400, margin=dict(l=0, r=10, t=36, b=10),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            legend=dict(orientation="h", y=1.08, x=0,
                        font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
            xaxis=dict(
                type="category",
                gridcolor="rgba(128,128,128,.10)",
                tickfont=dict(size=10),
                tickangle=-35,
                showline=True, linecolor="rgba(128,128,128,.2)",
                title=dict(text="Data do jogo", font=dict(size=11)),
            ),
            yaxis=dict(
                gridcolor="rgba(128,128,128,.12)",
                tickfont=dict(size=11),
                title=dict(text="Pts acumulados", font=dict(size=11)),
                rangemode="tozero",
            ),
            hovermode="x unified",
        )
        # Phase legend strip
        legend_html = (
            '<div style="display:flex;flex-wrap:wrap;align-items:center;'
            'gap:0;margin:-4px 0 6px;padding:6px 10px;'
            'border-radius:8px;background:rgba(128,128,128,.06)">' +
            f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:.78rem;font-weight:600;white-space:nowrap"><span style="display:inline-block;width:13px;height:13px;border-radius:3px;background:#123A56;opacity:.75;flex-shrink:0"></span>Grupos</span>'f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:.78rem;font-weight:600;white-space:nowrap"><span style="display:inline-block;width:13px;height:13px;border-radius:3px;background:#0D8587;opacity:.75;flex-shrink:0"></span>R32</span>'f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:.78rem;font-weight:600;white-space:nowrap"><span style="display:inline-block;width:13px;height:13px;border-radius:3px;background:#D6B864;opacity:.75;flex-shrink:0"></span>Oitavas</span>'f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:.78rem;font-weight:600;white-space:nowrap"><span style="display:inline-block;width:13px;height:13px;border-radius:3px;background:#DC884A;opacity:.75;flex-shrink:0"></span>Quartas</span>'f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:.78rem;font-weight:600;white-space:nowrap"><span style="display:inline-block;width:13px;height:13px;border-radius:3px;background:#B2584E;opacity:.75;flex-shrink:0"></span>Semis</span>'f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:.78rem;font-weight:600;white-space:nowrap"><span style="display:inline-block;width:13px;height:13px;border-radius:3px;background:#7B3FA0;opacity:.75;flex-shrink:0"></span>3°/Final</span>' +
            '</div>'
        )
        st.markdown(legend_html, unsafe_allow_html=True)
        st.plotly_chart(fig_evo, width='stretch', config={"displayModeBar": False})

    else:  # Pontuação por fase
        PHASES_ORDER = ["Grupos","Bônus","R32","Oitavas","Quartas","Semi","3o Lugar","Final"]

        def phase_pts(sc, phase):
            if phase == "Grupos": return sc["grupos"]
            if phase == "Bônus":  return sc["bonus"]
            midxs = [m for m,(rnd,*_) in enumerate(KO) if rnd==phase]
            return sum(sc["mdet"].get(m,0) or 0 for m in midxs)

        sel_phase = st.select_slider(
            "Fase:", options=PHASES_ORDER, value="Grupos",
            label_visibility="collapsed",
        )
        phase_data = sorted(
            [(nm, phase_pts(sc, sel_phase)) for nm,_,_,_,sc,_ in active_bettors],
            key=lambda x: -x[1],
        )
        bar_names  = [p[0] for p in phase_data]
        bar_vals   = [p[1] for p in phase_data]
        bar_colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(phase_data))]

        fig_bar = go.Figure(go.Bar(
            x=bar_names, y=bar_vals, text=bar_vals,
            textposition="outside",
            marker_color=bar_colors,
            hovertext=[p[0] for p in phase_data],
            hovertemplate="<b>%{hovertext}</b><br>%{y} pts<extra></extra>",
        ))
        fig_bar.update_layout(
            height=340, margin=dict(l=0, r=10, t=40, b=10),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            title=dict(text=f"Pontuação — fase: <b>{sel_phase}</b>",
                       font=dict(size=13), x=0),
            xaxis=dict(gridcolor="rgba(128,128,128,.10)", tickfont=dict(size=11)),
            yaxis=dict(gridcolor="rgba(128,128,128,.12)", tickfont=dict(size=11),
                       rangemode="tozero"),
        )
        st.plotly_chart(fig_bar, width='stretch', config={"displayModeBar": False})


# ── TAB 2: GRUPOS ────────────────────────────────────────────────────
with T2:
    st.markdown('<div class="sh">⚽ Classificação nos Grupos</div>', unsafe_allow_html=True)
    mode = st.radio("Baseado em:", ["📋 Resultados Reais", "👤 Apostas de Participante", "📊 Visão Consolidada"],
                    horizontal=True, label_visibility="collapsed")

    DOT   = {1:'#22C55E',2:'#86EFAC',3:'#FB923C',4:'#F87171'}
    RCLS  = {1:'row-q1',2:'row-q2',3:'row-q3',4:'row-q4'}

    if mode.startswith("📊"):
        # ── Consolidated: all bettors' predictions vs real results, game by game ──
        st.markdown('<div style="font-size:.8rem;opacity:.6;margin-bottom:8px">Apostas dos participantes selecionados por jogo, comparadas com o resultado real.</div>', unsafe_allow_html=True)

        # Participant filter
        _cons_all_names = [b[0] for b in bettors]
        _ca1, _ca2 = st.columns([1, 4], vertical_alignment="bottom")
        with _ca1:
            if st.button("👥 Todos", width='stretch', key="cons_todos"):
                st.session_state["cons_sel"] = _cons_all_names
                st.session_state["cons_ms"]  = _cons_all_names
                st.rerun()
        with _ca2:
            if "cons_sel" not in st.session_state:
                st.session_state["cons_sel"] = []
            _cons_valid = [n for n in st.session_state["cons_sel"] if n in _cons_all_names]
            _cons_chosen = st.multiselect(
                "Participantes:",
                options=_cons_all_names,
                default=_cons_valid,
                key="cons_ms",
                placeholder="Selecione participantes...",
            )
            st.session_state["cons_sel"] = _cons_chosen
        _cons_bettors = [b for b in bettors if b[0] in st.session_state["cons_sel"]] or bettors

        sel_grp = st.selectbox("Grupo:", GL, key="cons_grp")

        # Build columns for selected bettors
        _cb_names = [b[0] for b in _cons_bettors]

        # Compute sticky "Jogo" column width based on longest match name in this group
        _max_jogo_chars = max(
            (len(t1) + 3 + len(t2))  # 3 = len(" × ")
            for _, g_, t1, t2 in GROUP_FIXTURES
            if g_ == sel_grp
        ) if any(g_ == sel_grp for _, g_, _, _ in GROUP_FIXTURES) else 20
        _jogo_col_w = max(130, _max_jogo_chars * 7 + 50)  # ~7px/char + 2 flags + padding

        _STICKY_TH_G = (f"position:sticky;left:0;z-index:2;background:#0D2B40;"
                        f"text-align:left;white-space:nowrap;"
                        f"width:{_jogo_col_w}px;min-width:{_jogo_col_w}px")
        hdr_cells = "".join(
            f'<th style="white-space:nowrap;text-align:center;min-width:54px">{n}</th>'
            for n in _cb_names
        )
        hdr_html = (f'<th style="{_STICKY_TH_G}">Jogo</th>'
                    f'<th style="white-space:nowrap;min-width:46px">Data</th>'
                    f'<th style="white-space:nowrap;min-width:54px">Real</th>'
                    f'{hdr_cells}')
        rows_html = ""
        for m,(gdate,g,t1,t2) in enumerate(GROUP_FIXTURES):
            if g != sel_grp: continue
            real = gr.get(m)
            rs   = f"<b>{real[0]}–{real[1]}</b>" if real else "<span style='opacity:.4'>⏳</span>"
            ds   = gdate.strftime("%d/%m")
            cells = ""
            for nm_,bgb_,_,_,bsc_,_ in _cons_bettors:
                bet_ = bgb_.get(m)
                pts_ = bsc_['gdet'].get(m)
                if bet_ is None:
                    cells += '<td style="opacity:.3;text-align:center">—</td>'
                    continue
                bs_ = f"{bet_[0]}–{bet_[1]}"
                color_ = ('#22C55E' if pts_==5 else '#5EEAD4' if pts_==3 else
                          '#FB923C' if pts_==2 else '#F87171' if pts_==0 else 'inherit')
                cells += f'<td style="color:{color_};text-align:center">{bs_}</td>'
            _td_jogo = (f'<td style="position:sticky;left:0;z-index:1;'
                        f'background:var(--background-color,white);'
                        f'box-shadow:2px 0 4px rgba(0,0,0,.08);'
                        f'width:{_jogo_col_w}px;min-width:{_jogo_col_w}px;'
                        f'white-space:nowrap;font-size:.75rem">'
                        f'{FI(t1)}{t1} × {FI(t2)}{t2}</td>')
            rows_html += (f'<tr>{_td_jogo}'
                          f'<td style="opacity:.55;font-size:.72rem;text-align:center">{ds}</td>'
                          f'<td style="text-align:center">{rs}</td>{cells}</tr>')
        st.markdown(
            f'<div style="overflow-x:auto;-webkit-overflow-scrolling:touch;border-radius:8px">'
            f'<table class="mm-tbl" style="width:auto;table-layout:auto">'
            f'<thead><tr>{hdr_html}</tr></thead><tbody>{rows_html}</tbody>'
            f'</table></div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="font-size:.72rem;opacity:.5;margin-top:6px">'
            '<span style="color:#22C55E">■</span> Placar exato (5pts) &nbsp;'
            '<span style="color:#5EEAD4">■</span> Saldo exato (3pts) &nbsp;'
            '<span style="color:#FB923C">■</span> Vencedor certo (2pts) &nbsp;'
            '<span style="color:#F87171">■</span> Errou (0pts)</div>',
            unsafe_allow_html=True,
        )
    else:
        if mode.startswith("📋"):
            data_src = sort_st(calc_st(gr)) if gr else {
                g:[(t,{'pts':0,'played':0,'w':0,'d':0,'l':0,'gf':0,'ga':0}) for t in ts]
                for g,ts in GROUPS_DATA.items()}
            if not gr: st.info("Nenhum resultado preenchido no gabarito ainda.")
        else:
            sel_b = st.selectbox("Apostador", [b[0] for b in bettors], key='grp_pick')
            bd    = next(b for b in bettors if b[0]==sel_b)
            data_src = sort_st(calc_st(bd[1]))

        for row_g in [GL[i:i+3] for i in range(0,12,3)]:
            cs = st.columns(3, gap="small")
            for c,grp in zip(cs,row_g):
                with c:
                    color = GRP_COLORS.get(grp,'#123A56')
                    rows  = ""
                    for pos,(team,d) in enumerate(data_src.get(grp,[]),1):
                        gd  = d['gf']-d['ga']; pl = d['played']
                        pct = int(d['pts']/(pl*3)*100) if pl else 0
                        gdc = '#22C55E' if gd>0 else '#EF4444' if gd<0 else 'inherit'
                        gds = ('+' if gd>0 else '')+str(gd)
                        rows += f"""<tr>
                          <td><span class="dot" style="background:{DOT.get(pos,'#888')}"></span>{pos}</td>
                          <td class="nm">{FI(team)}{team}</td>
                          <td><b>{d['pts']}</b></td><td>{pl}</td>
                          <td>{d['w']}</td><td>{d['d']}</td><td>{d['l']}</td>
                          <td>{d['gf']}</td><td>{d['ga']}</td>
                          <td style="color:{gdc}">{gds}</td><td>{pct}%</td>
                        </tr>"""
                    st.markdown(f"""<div class="gb">
                      <div class="gb-hdr" style="color:{color}">Grupo {grp}</div>
                      <table class="gt">
                        <tr><th>#</th><th style="text-align:left">Seleção</th>
                            <th>P</th><th>J</th><th>V</th><th>E</th><th>D</th>
                            <th>GP</th><th>GC</th><th>SG</th><th>%</th></tr>
                        {rows}
                      </table>
                    </div>""", unsafe_allow_html=True)

# ── TAB 3: MATA-MATA ─────────────────────────────────────────────────
with T3:
    st.markdown('<div class="sh">🗓️ Mata-Mata</div>', unsafe_allow_html=True)

    mm_mode = st.radio(
        "Visualização:",
        ["📋 Resultados Reais", "🗓️ Seleções por Fase"],
        horizontal=True, label_visibility="collapsed",
    )

    # Pre-compute real winners per round (used in both views)
    real_by_rnd: dict = {
        rnd: {rw.get(m) for m in midxs} - {None}
        for rnd, midxs in rnd_ms.items()
    }

    if mm_mode.startswith("📋"):
        # ── Resultados Reais ─────────────────────────────────────────────
        if not mmr:
            st.info("Nenhum resultado de Mata-Mata preenchido no gabarito ainda.")
        else:
            for rnd, mlist in rnd_ms.items():
                _ico_r = RND_ICO.get(rnd, "⚪")
                _pv_r  = KO_PTS[rnd]
                _has   = any(m in mmr for m in mlist)
                with st.expander(f"{_ico_r} {rnd} — {_pv_r} pts/seleção", expanded=(_has and rnd == "R32")):
                    if not _has:
                        st.markdown('<div style="opacity:.45;font-size:.82rem">⏳ Jogos ainda não realizados.</div>',
                                    unsafe_allow_html=True)
                        continue
                    _html_r = ""
                    for m in mlist:
                        t1r, t2r = rn.get(m, ('?','?'))
                        sc_r     = mmr.get(m)
                        ko_date  = KO_DATES[m] if m < len(KO_DATES) else None
                        ds_r     = ko_date.strftime("%d/%m") if ko_date else "—"
                        if sc_r:
                            g1v, g2v, pen = sc_r
                            score_str = f"<b>{g1v}–{g2v}</b>"
                            if pen:
                                score_str += f' <span style="font-size:.68rem;opacity:.6">({pen})</span>'
                            winner = rw.get(m)
                            t1_sty = "font-weight:700;color:#22C55E" if winner == t1r else ""
                            t2_sty = "font-weight:700;color:#22C55E" if winner == t2r else ""
                            _html_r += (
                                f'<div class="mr">'
                                f'<div style="min-width:40px;font-size:.68rem;opacity:.5">{ds_r}</div>'
                                f'<div class="mr-t" style="display:flex;align-items:center;gap:6px">'
                                f'<span style="{t1_sty}">{FI(t1r)}{t1r}</span>'
                                f'<span style="opacity:.4;font-size:.75rem">vs</span>'
                                f'<span style="{t2_sty}">{FI(t2r)}{t2r}</span>'
                                f'</div>'
                                f'<span class="mr-s">{score_str}</span>'
                                f'</div>'
                            )
                        else:
                            _html_r += (
                                f'<div class="mr" style="opacity:.5">'
                                f'<div style="min-width:40px;font-size:.68rem;opacity:.5">{ds_r}</div>'
                                f'<div class="mr-t">{FI(t1r)}{t1r} <span style="opacity:.4">vs</span> {FI(t2r)}{t2r}</div>'
                                f'<span class="mr-s">⏳</span>'
                                f'</div>'
                            )
                    st.markdown(_html_r, unsafe_allow_html=True)

    else:
        # ── Seleções por Fase ────────────────────────────────────────────
        btr_all = [(nm, prnd, sc) for nm, gb, bb, xm, sc, prnd in bettors]

        _all_btr_names = [nm for nm, _, _ in btr_all]
        _f1, _f2 = st.columns([1, 5], vertical_alignment="bottom")
        with _f1:
            if st.button("👥 Todos", width='stretch', key="mm_todos"):
                st.session_state["mm_sel_bettors"] = _all_btr_names
                st.session_state["mm_ms_bettors"]  = _all_btr_names
                st.rerun()
        with _f2:
            if "mm_sel_bettors" not in st.session_state:
                st.session_state["mm_sel_bettors"] = []
            _valid_mm = [n for n in st.session_state["mm_sel_bettors"] if n in _all_btr_names]
            _chosen_mm = st.multiselect(
                "Participantes:",
                options=_all_btr_names,
                default=_valid_mm,
                key="mm_ms_bettors",
                placeholder="Selecione participantes...",
            )
            st.session_state["mm_sel_bettors"] = _chosen_mm

        btr = [b for b in btr_all if b[0] in st.session_state["mm_sel_bettors"]]
        if not btr:
            btr = btr_all

        # ── Legenda ───────────────────────────────────────────────────
        st.markdown("""
        <div style="display:flex;flex-wrap:wrap;gap:16px;align-items:center;
                    padding:7px 14px;border-radius:8px;
                    background:rgba(128,128,128,.06);
                    border:1px solid rgba(128,128,128,.1);
                    font-size:.75rem;margin-bottom:4px">
          <span style="opacity:.55;font-size:.68rem;font-weight:700;
                       text-transform:uppercase;letter-spacing:.8px">Legenda:</span>
          <span>✅ Apostou &amp; avançou</span>
          <span>❌ Apostou mas não avançou</span>
          <span>🔵 Avançou mas não apostou</span>
          <span>⏳ Aguardando resultado</span>
          <span style="opacity:.45">· Não apostou</span>
        </div>
        """, unsafe_allow_html=True)

        _STICKY_TH = ("position:sticky;left:0;z-index:2;background:#0D2B40;"
                      "min-width:160px;text-align:left")
        _STICKY_TD_HIT  = ("position:sticky;left:0;z-index:1;background:var(--background-color,white);"
                           "min-width:160px;font-weight:700;color:#0D8587;"
                           "border-right:2px solid rgba(13,133,135,.3)")
        _STICKY_TD_MISS = ("position:sticky;left:0;z-index:1;background:var(--background-color,white);"
                           "min-width:160px;opacity:.6;"
                           "border-right:2px solid rgba(128,128,128,.15)")
        _STICKY_FOOT    = ("position:sticky;left:0;z-index:1;background:#0D2B40;"
                           "color:white;font-size:.7rem;font-weight:700;min-width:160px")

        for rnd, mlist in rnd_ms.items():
            pv       = KO_PTS[rnd]
            real_set = real_by_rnd.get(rnd, set())
            started  = bool(real_set)
            ico      = RND_ICO.get(rnd, "⚪")

            with st.expander(f"{ico} {rnd} — {pv} pts por seleção que avançar", expanded=(rnd=="R32")):
                if started:
                    real_pills = "".join(
                        f'<span class="pill pill-real">{FI(t)}{t}</span>'
                        for t in sorted(real_set, key=str)
                    )
                    _real_lbl = {'3o Lugar': '🥉 3ª Colocada', 'Final': '🏆 Campeã'}.get(rnd, '✅ Avançaram')
                    st.markdown(
                        f'<div style="margin-bottom:10px">'
                        f'<div class="rnd-section-lbl">{_real_lbl} ({len(real_set)})</div>'
                        f'<div>{real_pills}</div></div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="opacity:.5;font-size:.82rem;margin-bottom:10px">⏳ Rodada ainda não iniciou.</div>',
                        unsafe_allow_html=True,
                    )

                all_teams = set(real_set)
                for _, prnd, _ in btr:
                    all_teams |= prnd.get(rnd, set())
                all_teams -= {"?", None}
                sorted_teams = (
                    sorted(all_teams & real_set, key=str) +
                    sorted(all_teams - real_set,  key=str)
                )

                _th_cells = "".join(
                    f'<th style="text-align:center;white-space:nowrap;min-width:52px">'
                    f'{nm}</th>'
                    for nm, _, _ in btr
                )
                _header = f'<tr><th style="{_STICKY_TH}">Seleção</th>{_th_cells}</tr>'

                _body = ""
                for team in sorted_teams:
                    in_real = team in real_set
                    _td_sty = _STICKY_TD_HIT if in_real else _STICKY_TD_MISS
                    _td_team = f'<td style="{_td_sty};font-size:.8rem">{FI(team)}{team}</td>'
                    _tds = ""
                    for _, prnd, _ in btr:
                        picked = team in prnd.get(rnd, set())
                        if picked and in_real and started: ic = "✅" # apostou E avançou
                        elif picked and not started:        ic = "⏳" # apostou, mas rodada ainda não começou
                        elif picked:                        ic = "❌" # apostou mas NÃO avançou
                        elif in_real and started:           ic = "🔵" # avançou mas o apostador NÃO apostou
                        else:                               ic = '<span style="opacity:.25">·</span>' # nem apostou nem avançou
                        _tds += f'<td style="text-align:center;font-size:.85rem">{ic}</td>'
                    _row_bg = "rgba(13,133,135,.06)" if in_real else ""
                    _body  += f'<tr style="background:{_row_bg}">{_td_team}{_tds}</tr>'

                _ft_cells = ""
                for _, prnd, sc in btr:
                    my = prnd.get(rnd, set())
                    ok = len(my & real_set) if started else "—"
                    pts_rnd = sum(sc["mdet"].get(m, 0) or 0 for m in mlist)
                    _ft_cells += (
                        f'<td style="text-align:center;border-top:2px solid rgba(214,184,100,.3)">'
                        f'<div style="font-size:.85rem;font-weight:800;color:#D6B864">{pts_rnd}</div>'
                        f'<div style="font-size:.62rem;opacity:.55">{ok}/{len(my)}</div></td>'
                    )
                _foot = (f'<tr><td style="{_STICKY_FOOT};border-top:2px solid rgba(214,184,100,.3)">'
                         f'PTS / ACERTOS</td>{_ft_cells}</tr>')

                st.markdown(
                    f'<div style="overflow-x:auto;-webkit-overflow-scrolling:touch;'
                    f'border-radius:8px;margin-top:8px">'
                    f'<table class="mm-tbl" style="width:auto;table-layout:auto">'
                    f'<thead>{_header}</thead>'
                    f'<tbody>{_body}</tbody>'
                    f'<tfoot>{_foot}</tfoot>'
                    f'</table></div>',
                    unsafe_allow_html=True,
                )

    # ── Per-participant bracket view ─────────────────────────────────────
    st.markdown('<div class="sh">🔍 Chaveamento por Participante</div>', unsafe_allow_html=True)
    sel_b_mm = st.selectbox("Apostador", [b[0] for b in bettors], key="mm_part_sel")
    _bd_mm = next(b for b in bettors if b[0] == sel_b_mm)
    _, _bgb_mm, _, _bxm_mm, _bsc_mm,_ = _bd_mm
    _picks_mm, _bn_mm = sim_bet(_bgb_mm, _bxm_mm, t495)

    # Group picks by round
    _part_prnd: dict = {}
    for _m, (_rnd, *_) in enumerate(KO):
        _p = _picks_mm.get(_m)
        if _p and _p != "?":
            _part_prnd.setdefault(_rnd, []).append((_m, _p))

    # Bracket-style columns: one per phase
    _phases_order = ['R32','Oitavas','Quartas','Semi','3o Lugar','Final']
    _phase_cols = st.columns(len(_phases_order), gap="small")
    for _ci, _ph in enumerate(_phases_order):
        with _phase_cols[_ci]:
            _pv = KO_PTS[_ph]
            _ico = RND_ICO.get(_ph, '⚪')
            st.markdown(
                f'<div style="font-size:.7rem;font-weight:800;text-align:center;'
                f'opacity:.7;margin-bottom:6px">{_ico} {_ph}<br>'
                f'<span style="font-weight:400;opacity:.6">{_pv}pts/sel</span></div>',
                unsafe_allow_html=True,
            )
            _real_ph = real_by_rnd.get(_ph, set())
            _part_picks_ph = _part_prnd.get(_ph, [])
            if not _part_picks_ph:
                st.markdown('<div style="opacity:.3;font-size:.75rem;text-align:center">—</div>', unsafe_allow_html=True)
                continue
            for _m_idx, _team in _part_picks_ph:
                _t1, _t2 = _bn_mm.get(_m_idx, ('?','?'))
                _in_real = _team in _real_ph
                _started = bool(_real_ph)
                if not _started:
                    _bg, _fc = "rgba(128,128,128,.1)", "inherit"
                    _status = "⏳"
                elif _in_real:
                    _bg, _fc = "rgba(13,133,135,.18)", "#0D8587"
                    _status = "✅"
                else:
                    _bg, _fc = "rgba(178,88,78,.15)", "#B2584E"
                    _status = "❌"
                # show both matchup teams + highlight predicted winner
                _opp = _t2 if _team == _t1 else _t1
                st.markdown(
                    f'<div style="background:{_bg};border-radius:7px;padding:5px 7px;'
                    f'margin-bottom:4px;font-size:.73rem">'
                    f'<div style="font-weight:700;color:{_fc}">{_status} {FI(_team)}{_team}</div>'
                    f'<div style="opacity:.45;font-size:.67rem">vs {FI(_opp)}{_opp}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            _pts_ph = sum(_bsc_mm["mdet"].get(_m_idx,0) or 0 for _m_idx,_ in _part_picks_ph)
            _ok_ph  = sum(1 for _,_t in _part_picks_ph if _t in _real_ph) if _real_ph else 0
            st.markdown(
                f'<div style="text-align:center;margin-top:4px;font-size:.72rem;'
                f'font-weight:700;color:#D6B864">{_pts_ph} pts'
                f'<span style="font-weight:400;opacity:.55"> ({_ok_ph}/{len(_part_picks_ph)})</span></div>',
                unsafe_allow_html=True,
            )

    # ── Overall MM ranking chart ──────────────────────────────────────────
    st.markdown('<div class="sh">📊 Ranking Mata-Mata</div>', unsafe_allow_html=True)
    mmdf = pd.DataFrame([(b[0], b[4]["mm"]) for b in bettors], columns=["Apostador","Pts"])
    fig2 = px.bar(mmdf, x="Apostador", y="Pts", text="Pts",
                  color="Pts", color_continuous_scale=["#B2584E","#DC884A","#D6B864"])
    fig2.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=0,r=0,t=5,b=5), coloraxis_showscale=False, height=230,
        xaxis=dict(gridcolor="rgba(128,128,128,.12)"),
        yaxis=dict(gridcolor="rgba(128,128,128,.12)"),
    )
    fig2.update_traces(textposition="outside")
    st.plotly_chart(fig2, width='stretch', config={"displayModeBar":False})

# ── TAB 4: POR APOSTADOR ─────────────────────────────────────────────
with T4:
    sel = st.selectbox("👤 Apostador", [b[0] for b in bettors])
    bnm,bgb,bbb,bxm,bsc,bprnd = next(b for b in bettors if b[0]==sel)
    pos   = next(i for i,(nm,*_) in enumerate(bettors,1) if nm==bnm)
    medal = MEDALS.get(pos, f"{pos}°")

    st.markdown(f"""
    <div class="hero" style="padding:18px 26px;margin-bottom:12px">
      <span style="font-size:1.5rem">{medal}</span>
      <span class="hero-title" style="font-size:1.45rem;margin-left:10px">{bnm}</span>
      <div class="hero-sub" style="margin-top:6px">
        Total: <b>{bsc['total']} pts</b> &nbsp;·&nbsp;
        Grupos: <b>{bsc['grupos']}</b> &nbsp;·&nbsp;
        Bônus: <b>{bsc['bonus']}</b> &nbsp;·&nbsp;
        Mata-Mata: <b>{bsc['mm']}</b>
      </div>
    </div>""", unsafe_allow_html=True)

    _sub_resumo, _sub_stats = st.tabs(["📋 Resumo", "📊 Estatísticas"])
 
    # ════════════════════════════════════════════════════════════════
    # SUB-ABA: RESUMO (conteúdo original da aba Por Apostador)
    # ════════════════════════════════════════════════════════════════
    with _sub_resumo:
        # Bônus
        st.markdown('<div class="sh">🎁 Bônus Pré-Copa</div>', unsafe_allow_html=True)
        ab,mb = bbb; ar,mr_ = (br or (None,None))
        for col,(lbl,bv,rv,pts) in zip(st.columns(2),[
            ("⚽ Artilheiro", ab, ar, bsc['art_pts']),
            ("🏅 Melhor Campanha", mb, mr_, bsc['mg_pts']),
        ]):
            ok  = bv and rv and str(bv).strip().lower()==str(rv).strip().lower()
            ico = "✅" if ok else ("❌" if rv else "⏳")
            bvs = f"{F(bv)} {bv}" if bv else "—"
            rvs = f"{F(rv)} {rv}" if rv else "Aguardando"
            col.markdown(f"""<div class="bc">
            <div class="bc-lbl">{lbl}</div>
            <div class="bc-ico">{ico}</div>
            <div class="bc-bet">Apostou: {bvs}</div>
            <div class="bc-real">Real: {rvs}</div>
            <div class="bc-pts">{pts} pts</div>
            </div>""", unsafe_allow_html=True)

        # Grupos
        st.markdown('<div class="sh">⚽ Fase de Grupos — Jogo a Jogo</div>', unsafe_allow_html=True)
        for row_g in [GL[i:i+3] for i in range(0,12,3)]:
            gcols = st.columns(3)
            for gcol,grp in zip(gcols,row_g):
                with gcol:
                    color = GRP_COLORS.get(grp,'#123A56')
                    st.markdown(
                        f'<div style="font-weight:800;color:{color};font-size:.87rem;margin-bottom:5px">Grupo {grp}</div>',
                        unsafe_allow_html=True)
                    html = ""
                    for m,(gdate,g,t1,t2) in enumerate(GROUP_FIXTURES):
                        if g!=grp: continue
                        pts  = bsc['gdet'].get(m); bet = bgb.get(m); real = gr.get(m)
                        bs   = f"{bet[0]}–{bet[1]}" if bet else "—"
                        rs   = f"{real[0]}–{real[1]}" if real else "⏳"
                        bdg  = ('<span class="b5">5</span>' if pts==5 else
                                '<span class="b3">3</span>' if pts==3 else
                                '<span class="b2">2</span>' if pts==2 else
                                '<span class="b0">0</span>' if pts==0 else
                                '<span class="bN">–</span>')
                        ds = gdate.strftime("%d/%m")
                        html += f"""<div class="mr">
                        <div class="mr-t">{FI(t1)}{t1}<br>{FI(t2)}{t2}</div>
                        <span class="mr-s"><span style="opacity:.5;font-size:.7rem">{ds}</span> {bs} → {rs}</span>
                        {bdg}
                        </div>"""
                    st.markdown(html, unsafe_allow_html=True)

        # Mata-mata
        # Mata-mata por apostador
        # ── Mata-Mata: seleções por fase ──────────────────────────────────────
        st.markdown('<div class="sh">🏆 Mata-Mata — Seleções por Fase</div>', unsafe_allow_html=True)

        bpicks_b, bnames_b = sim_bet(bgb, bxm, t495)
        b_prnd: dict = {}
        for m, (rnd, *_) in enumerate(KO):
            p = bpicks_b.get(m)
            if p and p != "?":
                b_prnd.setdefault(rnd, set()).add(p)

        # Reuse real_by_rnd from T3 scope (same module)

        for rnd, mlist in rnd_ms.items():
            pv         = KO_PTS[rnd]
            real_set   = real_by_rnd.get(rnd, set())
            my_picks   = b_prnd.get(rnd, set())
            started    = bool(real_set)
            pts_earned = sum(bsc["mdet"].get(m,0) or 0 for m in mlist)
            correct    = len(my_picks & real_set) if started else 0
            missed     = my_picks - real_set         # picked but didn't advance
            not_picked = real_set - my_picks         # advanced but not picked
            ico        = RND_ICO.get(rnd,"⚪")

            pts_label = (f"+{pts_earned} pts" if pts_earned > 0 else
                        ("⏳ aguardando" if not started else "0 pts"))

            with st.expander(
                f"{ico} {rnd}  ({pv} pts/seleção) — {pts_label}",
                expanded=(rnd in ["R32","Oitavas"]),
            ):
                col_p, col_r = st.columns(2, gap="large")

                with col_p:
                    _pick_lbl = {
                        '3o Lugar': '🥉 Apostou que seria a 3ª Colocada',
                        'Final':    '🏆 Apostou que seria a Campeã',
                    }.get(rnd, f'🎯 Apostou que avançariam')
                    st.markdown(
                        f'<div class="rnd-section-lbl">{_pick_lbl} ({len(my_picks)})</div>',
                        unsafe_allow_html=True,
                    )
                    if my_picks:
                        html_p = ""
                        for team in sorted(my_picks, key=str):
                            in_real = team in real_set
                            if not started:
                                cls, ico_t = "pill-wait", "⏳"
                            elif in_real:
                                cls, ico_t = "pill-hit",  "✅"
                            else:
                                cls, ico_t = "pill-miss", "❌"
                            html_p += f'<span class="pill {cls}">{ico_t} {FI(team)}{team}</span>'
                        st.markdown(f'<div style="line-height:2">{html_p}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="opacity:.45;font-size:.82rem">Nenhuma seleção escolhida.</div>',
                                    unsafe_allow_html=True)

                with col_r:
                    st.markdown(
                        f'<div class="rnd-section-lbl">✅ Avançaram de verdade ({len(real_set)})</div>',
                        unsafe_allow_html=True,
                    )
                    if real_set:
                        html_r = ""
                        for team in sorted(real_set, key=str):
                            picked_it = team in my_picks
                            cls_r  = "pill-hit"  if picked_it else "pill-real"
                            ico_r  = "🎯 " if picked_it else ""
                            html_r += f'<span class="pill {cls_r}">{ico_r}{FI(team)}{team}</span>'
                        st.markdown(f'<div style="line-height:2">{html_r}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="opacity:.45;font-size:.82rem">Aguardando resultados...</div>',
                                    unsafe_allow_html=True)

                # Score strip
                if started:
                    miss_str = (", ".join(f"{F(t)}{t}" for t in sorted(missed, key=str))
                                if missed else "—")
                    unexp_str = (", ".join(f"{F(t)}{t}" for t in sorted(not_picked, key=str))
                                if not_picked else "—")
                    st.markdown(
                        f'<div class="score-strip">'
                        f'<span>✅ <b style="color:#22C55E">{correct}</b> acertos</span>'
                        f'<span>❌ <b style="color:#EF4444">{len(missed)}</b> erros</span>'
                        f'<span style="opacity:.6">Não apostou em: {unexp_str}</span>'
                        f'<span style="margin-left:auto;font-weight:800;color:#D6B864">{pts_earned} pts</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

    # ════════════════════════════════════════════════════════════════
    # SUB-ABA: ESTATÍSTICAS
    # ════════════════════════════════════════════════════════════════
    with _sub_stats:
 
        # Helpers locais ------------------------------------------------
        def _mc(col, v, l):
            col.markdown(
                f'<div class="mc"><div class="mc-v">{v}</div>'
                f'<div class="mc-l">{l}</div></div>', unsafe_allow_html=True)
 
        def _cbar(flag_html, label, count, total, color="#0D8587", mark=False):
            pct = (count / total * 100) if total else 0
            edge = "outline:2px solid #D6B864;outline-offset:-2px;" if mark else ""
            return (
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">'
                f'<div style="width:150px;font-size:.78rem;white-space:nowrap;'
                f'overflow:hidden;text-overflow:ellipsis">{flag_html}{label}</div>'
                f'<div style="flex:1;background:rgba(128,128,128,.12);border-radius:5px;height:18px">'
                f'<div style="width:{pct:.0f}%;background:{color};height:100%;'
                f'border-radius:5px;min-width:3px;{edge}"></div></div>'
                f'<div style="width:70px;text-align:right;font-size:.72rem;font-weight:700">'
                f'{pct:.0f}% <span style="opacity:.5;font-weight:400">({count})</span></div>'
                f'</div>')
 
        def _placar(t): return f"{t[0]}×{t[1]}"
 
        def _vbar(label, valstr, frac, color):
            w = max(3.0, min(100.0, frac * 100))
            return (
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:5px">'
                f'<div style="width:150px;font-size:.78rem">{label}</div>'
                f'<div style="flex:1;background:rgba(128,128,128,.12);border-radius:5px;height:20px">'
                f'<div style="width:{w:.0f}%;background:{color};height:100%;border-radius:5px"></div></div>'
                f'<div style="width:74px;text-align:right;font-weight:700;font-size:.82rem">{valstr}</div>'
                f'</div>')
 
        def _sent(p):
            return 0 if p[0] == p[1] else (1 if p[0] > p[1] else -1)
 
        def _classify(my, cons):
            # classifica a divergência do SEU palpite vs o consenso (mais comum)
            if my == cons:
                return ('Igual', '🟢', '#22C55E')
            a, b = _sent(my), _sent(cons)
            if a == b:
                return ('Mesmo sentido', '🟡', '#C9A227')      # mesmo resultado, placar difere
            if a == 0 or b == 0:
                return ('Sentido adjacente', '🟠', '#E08A3C')  # um cravou empate, o outro vitória
            return ('Sentido oposto', '🔴', '#B2584E')          # um cravou mandante, o outro visitante
 
        _n = stats["n"]
 
        # ── FEATURE 1: divergência nos placares de grupos ─────────────
        st.markdown('<div class="sh">📐 Seus placares vs o consenso do bolão</div>',
                    unsafe_allow_html=True)
 
        _gd = stats["group_dist"]
        _rows = []  # (m, t1, t2, mypick, mode, mode_pct, my_pct, total, my_n)
        for _m,(_dt,_g,_t1,_t2) in enumerate(GROUP_FIXTURES):
            _bv = bgb.get(_m)
            _c  = _gd.get(_m)
            if _bv is None or not _c:
                continue
            try: _mykey = (int(_bv[0]), int(_bv[1]))
            except Exception: continue
            _tot = sum(_c.values())
            _mode, _mode_n = _c.most_common(1)[0]
            _my_n = _c.get(_mykey, 0)
            _rows.append((_m, _t1, _t2, _mykey, _mode,
                          _mode_n/_tot, _my_n/_tot, _tot, _my_n))
 
        if not _rows:
            st.info("Sem palpites de grupos suficientes para comparar ainda.")
        else:
            _n_main   = sum(1 for r in _rows if r[3] == r[4])
            _n_unique = sum(1 for r in _rows if r[8] == 1)
            _rank, _rn = stats["conformity_rank"].get(bnm, (0, 0))
 
            _mcols = st.columns(3)
            _mc(_mcols[0], f"{_n_main}/{len(_rows)}", "Colocou o placar + popular")
            _mc(_mcols[1], f"{_n_unique}", "Palpites únicos (só você)")
            _mc(_mcols[2], f"{_rank}º<span style='font-size:.9rem'> / {_rn}</span>",
                "Ranking de ousadia")
 
            # Você vs média do bolão (alinhamento de placar exato)
            _my_conf  = stats["conformity"].get(bnm, 0.0)
            _avg_conf = (sum(stats["conformity"].values()) / len(stats["conformity"])
                         if stats["conformity"] else 0.0)
            _ax = max(_my_conf, _avg_conf, 0.01)
            st.markdown(
                '<div style="font-size:.78rem;font-weight:700;margin:12px 0 2px">'
                '🧭 Alinhamento com o bolão</div>'
                '<div style="font-size:.72rem;opacity:.6;margin-bottom:6px">'
                '% médio do bolão que colocou o <b>placar exato</b> igual ao seu, jogo a jogo. '
                'Quanto maior, mais você joga "na manada".</div>',
                unsafe_allow_html=True)
            st.markdown(
                _vbar("<b>Você</b>", f'{_my_conf*100:.0f}% <span style="opacity:.5;font-weight:400">({round(_my_conf*_n)})</span>', _my_conf/_ax, "#D6B864") +
                _vbar("Média do bolão", f'{_avg_conf*100:.0f}% <span style="opacity:.5;font-weight:400">({round(_avg_conf*_n)})</span>', _avg_conf/_ax, "#7F7F7F"),
                unsafe_allow_html=True)
 
            # Pódio de ousadia (menor conformidade = mais ousado)
            _bold = sorted(stats["conformity"].items(), key=lambda x: x[1])[:3]
            _podium = " &nbsp; ".join(
                f'{m_} <b>{nm_}</b> <span style="opacity:.5">({cf*100:.0f}%)</span>'
                for (nm_, cf), m_ in zip(_bold, ["🥇","🥈","🥉"]))
            st.markdown(
                f'<div style="margin:10px 0;padding:8px 12px;border-radius:8px;'
                f'background:rgba(128,128,128,.06);font-size:.82rem">'
                f'<span style="opacity:.6;font-size:.7rem;font-weight:700;text-transform:uppercase;'
                f'letter-spacing:.6px">🎲 Mais ousados do bolão</span><br>{_podium}</div>',
                unsafe_allow_html=True)
 
            # Listas: mais ousados / mais alinhados
            _l1, _l2 = st.columns(2, gap="large")
            with _l1:
                st.markdown('<div style="font-size:.82rem;font-weight:700;margin-bottom:4px">'
                            '🎲 Seus palpites mais ousados</div>', unsafe_allow_html=True)
                for r in sorted(_rows, key=lambda x: x[6])[:6]:
                    _m,_t1,_t2,_mk,_md,_mdp,_myp,_tot,_myn = r
                    st.markdown(
                        f'<div style="font-size:.76rem;padding:5px 0;'
                        f'border-bottom:1px solid rgba(128,128,128,.08)">'
                        f'{FI(_t1)}{_t1} <span style="opacity:.4">×</span> {FI(_t2)}{_t2}<br>'
                        f'<span style="opacity:.85">você <b>{_placar(_mk)}</b> '
                        f'<span style="color:#B2584E">({_myp*100:.0f}%)</span></span> '
                        f'<span style="opacity:.5">· comum {_placar(_md)} ({_mdp*100:.0f}%)</span>'
                        f'</div>', unsafe_allow_html=True)
            with _l2:
                st.markdown('<div style="font-size:.82rem;font-weight:700;margin-bottom:4px">'
                            '🐑 Onde você seguiu a manada</div>', unsafe_allow_html=True)
                for r in sorted(_rows, key=lambda x: -x[6])[:6]:
                    _m,_t1,_t2,_mk,_md,_mdp,_myp,_tot,_myn = r
                    st.markdown(
                        f'<div style="font-size:.76rem;padding:5px 0;'
                        f'border-bottom:1px solid rgba(128,128,128,.08)">'
                        f'{FI(_t1)}{_t1} <span style="opacity:.4">×</span> {FI(_t2)}{_t2}<br>'
                        f'<span style="opacity:.85">você <b>{_placar(_mk)}</b> '
                        f'<span style="color:#0D8587">({_myp*100:.0f}%)</span></span>'
                        f'</div>', unsafe_allow_html=True)
 
 
            # Tabela cronológica completa: seu placar vs consenso (sentido + magnitude)
            _cats = {'Igual': 0, 'Mesmo sentido': 0, 'Sentido adjacente': 0, 'Sentido oposto': 0}
            for r in _rows:
                _cats[_classify(r[3], r[4])[0]] += 1
            st.markdown(
                f'<div style="font-size:.78rem;font-weight:700;margin:14px 0 4px">'
                f'📅 Jogo a jogo — seu placar vs o consenso</div>'
                f'<div style="font-size:.74rem;margin-bottom:6px">'
                f'🟢 <b>{_cats["Igual"]}</b> igual &nbsp;·&nbsp; '
                f'🟡 <b>{_cats["Mesmo sentido"]}</b> mesmo sentido &nbsp;·&nbsp; '
                f'🟠 <b>{_cats["Sentido adjacente"]}</b> adjacente &nbsp;·&nbsp; '
                f'🔴 <b>{_cats["Sentido oposto"]}</b> oposto</div>',
                unsafe_allow_html=True)
            with st.expander(f"Ver todos os {len(_rows)} jogos (ordem cronológica)", expanded=False):
                _html = ('<div style="display:flex;gap:6px;padding:3px 8px;font-size:.66rem;'
                         'opacity:.5;font-weight:700;text-transform:uppercase;letter-spacing:.4px">'
                         '<span style="width:42px">Data</span>'
                         '<span style="flex:1">Jogo</span>'
                         '<span style="width:60px;text-align:center">Você</span>'
                         '<span style="width:90px;text-align:center">Consenso</span>'
                         '<span style="width:150px;text-align:right">Divergência vs Consenso</span>'
                         '<span style="width:56px;text-align:center">Placar Real</span>'
                         '<span style="width:48px;text-align:center">Seus Pontos</span>'
                         '<span style="width:56px;text-align:center">Média Bolão</span>'
                         '<span style="width:66px;text-align:center">Δ vs méd</span></div>')
                
                # média de pontos por jogo (entre quem apostou e o jogo já pontuou)
                _pavg = {}
                for _b in bettors:
                    for _mm, _pv in _b[4].get('gdet', {}).items():
                        if _pv is not None:
                            _ac = _pavg.setdefault(_mm, [0, 0]); _ac[0] += _pv; _ac[1] += 1

                for r in sorted(_rows, key=lambda x: (GROUP_FIXTURES[x[0]][0], x[0])):
                    _m, _t1, _t2, _mk, _md, _mdp, _myp, _tot, _myn = r
                    _cat, _ic, _col = _classify(_mk, _md)
                    _dist = abs(_mk[0] - _md[0]) + abs(_mk[1] - _md[1])
                    _dstr = "" if _cat == 'Igual' else f' <span style="opacity:.55">·Δ{_dist}</span>'
                    _ds = GROUP_FIXTURES[_m][0].strftime("%d/%m")
                    _rp = gr.get(_m)
                    _sp = bsc['gdet'].get(_m, 0)
                    _am = (_pavg[_m][0] / _pavg[_m][1]) if _m in _pavg else None
                    if _rp is None or _am is None:
                        _real_s, _sua_s, _med_s = "⏳", "—", "—"
                        _dlt_html = '<span style="opacity:.35">—</span>'
                    else:
                        _real_s, _sua_s, _med_s = _placar(_rp), f"{_sp}", f"{_am:.1f}"
                        _d = _sp - _am
                        if _d > 0:
                            _bg = f"rgba(34,197,94,{0.12 + min(1.0, _d/5.0)*0.45:.2f})"; _dsg = f"+{_d:.1f}"
                        elif _d < 0:
                            _bg = f"rgba(239,68,68,{0.12 + min(1.0, -_d/5.0)*0.45:.2f})"; _dsg = f"{_d:.1f}"
                        else:
                            _bg = "rgba(128,128,128,.10)"; _dsg = "0"
                        _dlt_html = (f'<span style="display:inline-block;min-width:42px;padding:1px 6px;'
                                    f'border-radius:6px;background:{_bg};font-weight:700">{_dsg}</span>')

                    _html += (
                        f'<div style="display:flex;align-items:center;gap:6px;padding:5px 8px;'
                        f'border-left:3px solid {_col};border-bottom:1px solid rgba(128,128,128,.06);'
                        f'font-size:.75rem">'
                        f'<span style="width:42px;opacity:.5">{_ds}</span>'
                        f'<span style="flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;'
                        f'white-space:nowrap">{FI(_t1)}{_t1} <span style="opacity:.4">×</span> {FI(_t2)}{_t2}</span>'
                        f'<span style="width:60px;text-align:center"><b>{_placar(_mk)}</b></span>'
                        f'<span style="width:90px;text-align:center;opacity:.78">{_placar(_md)} '
                        f'<span style="opacity:.55;font-size:.68rem">{_mdp*100:.0f}%</span></span>'
                        f'<span style="width:150px;text-align:right;color:{_col}">{_ic} {_cat}{_dstr}</span>'
                        f'<span style="width:56px;text-align:center;opacity:.85">{_real_s}</span>'
                        f'<span style="width:48px;text-align:center;font-weight:700">{_sua_s}</span>'
                        f'<span style="width:56px;text-align:center;opacity:.7">{_med_s}</span>'
                        f'<span style="width:66px;text-align:center">{_dlt_html}</span>'
                        f'</div>')
                st.markdown(_html, unsafe_allow_html=True)
 
        # ── Perfil de gols: você vs média do bolão ────────────────────
        st.markdown('<div class="sh">⚽ Seu perfil de gols</div>', unsafe_allow_html=True)
        st.caption("Quantos gols seus palpites da fase de grupos grupos somam, comparado ao resto do bolão.")
        _gt = {}
        _gc = {}
        for _b in bettors:
            _s = 0
            _k = 0
            for _p in _b[1].values():
                try:
                    _s += int(_p[0]) + int(_p[1])
                    _k += 1
                except Exception:
                    pass
            _gt[_b[0]] = _s
            _gc[_b[0]] = _k
        _my_gt = _gt.get(bnm, 0)
        _my_gc = _gc.get(bnm, 0)
        _my_gpg = (_my_gt / _my_gc) if _my_gc else 0.0
        _gpgs = [(_gt[k] / _gc[k]) for k in _gt if _gc[k]]
        _avg_gpg = (sum(_gpgs) / len(_gpgs)) if _gpgs else 0.0
        _avg_gt = (sum(_gt.values()) / len(_gt)) if _gt else 0.0
        _rank_g = sorted(((k, (_gt[k] / _gc[k]) if _gc[k] else 0) for k in _gt), key=lambda x: -x[1])
        _grk = next((i + 1 for i, (k, _v) in enumerate(_rank_g) if k == bnm), 0)
 
        _axg = max(_my_gpg, _avg_gpg, 0.01)
        st.markdown('<div style="font-size:.78rem;font-weight:700;margin:4px 0 4px">Gols por jogo</div>',
                    unsafe_allow_html=True)
        st.markdown(
            _vbar("<b>Você</b>", f"{_my_gpg:.2f}", _my_gpg / _axg, "#D6B864") +
            _vbar("Média do bolão", f"{_avg_gpg:.2f}", _avg_gpg / _axg, "#7F7F7F"),
            unsafe_allow_html=True)
        _axt = max(_my_gt, _avg_gt, 1)
        st.markdown('<div style="font-size:.78rem;font-weight:700;margin:10px 0 4px">Total de gols previstos</div>',
                    unsafe_allow_html=True)
        st.markdown(
            _vbar("<b>Você</b>", f"{_my_gt:.0f}", _my_gt / _axt, "#0D8587") +
            _vbar("Média do bolão", f"{_avg_gt:.0f}", _avg_gt / _axt, "#7F7F7F"),
            unsafe_allow_html=True)
        _vsig = "acima da" if _my_gpg > _avg_gpg else ("abaixo da" if _my_gpg < _avg_gpg else "na")
        st.markdown(
            f'<div style="margin-top:8px;padding:8px 12px;border-radius:8px;'
            f'background:rgba(128,128,128,.06);font-size:.82rem">'
            f'Você prevê <b>{_my_gpg:.2f}</b> gols/jogo — <b>{_vsig}</b> média '
            f'(<b>{_avg_gpg:.2f}</b>). É o <b>{_grk}º</b> mais "goleador" de {len(_gt)}.</div>',
            unsafe_allow_html=True)
 
 
        # ── FEATURE 2: consenso do mata-mata por fase ─────────────────
        st.markdown('<div class="sh">🗺️ Para onde o bolão acha que cada seleção avança</div>',
                    unsafe_allow_html=True)
        st.caption("As seleções que mais gente colocou para avançar em cada fase. "
                   "Suas escolhas aparecem destacadas abaixo de cada gráfico.")
 
        _ROUNDS_F2 = [
            ('R32',     '🔵', 'Avançam às Oitavas'),
            ('Oitavas', '🟢', 'Avançam às Quartas'),
            ('Quartas', '🟡', 'Avançam às Semis'),
            ('Semi',    '🔴', 'Finalistas'),
        ]
        for _rk, _ico, _lbl in _ROUNDS_F2:
            _slots = stats["rnd_slots"].get(_rk, 0)
            _dist  = stats["ko_dist"].get(_rk, Counter())
            _top   = _dist.most_common(_slots)
            _top_set = {t for t, _ in _top}
            _mypicks = bprnd.get(_rk, set())
            with st.expander(f"{_ico} {_rk} — {_lbl} (top {_slots})",
                             expanded=(_rk == 'R32')):
                if not _top:
                    st.markdown('<div style="opacity:.45;font-size:.82rem">Sem palpites ainda.</div>',
                                unsafe_allow_html=True)
                else:
                    _bars = "".join(
                        _cbar(FI(t), t, c, _n, mark=(t in _mypicks)) for t, c in _top)
                    st.markdown(_bars, unsafe_allow_html=True)
                if _mypicks:
                    _chips = ""
                    for t in sorted(_mypicks, key=str):
                        _inc = t in _top_set
                        _bg  = "rgba(13,133,135,.16)" if _inc else "rgba(178,88,78,.14)"
                        _ic  = "✅" if _inc else "⚠️"
                        _pt = (_dist.get(t, 0) / _n * 100) if _n else 0
                        _chips += (f'<span title="{_pt:.0f}% do bolão colocou esta seleção" '
                                   f'style="display:inline-block;background:{_bg};'
                                   f'border-radius:14px;padding:2px 10px;margin:3px 4px 0 0;cursor:help;'
                                   f'font-size:.74rem">{_ic} {FI(t)}{t}</span>')
                    st.markdown(
                        f'<div style="margin-top:8px;border-top:1px solid rgba(128,128,128,.12);'
                        f'padding-top:6px"><span style="font-size:.7rem;opacity:.6;'
                        f'font-weight:700">SEUS PALPITES (✅ no consenso · ⚠️ fora):</span><br>'
                        f'{_chips}</div>', unsafe_allow_html=True)
 
        # ── FEATURES 3/4/5: bônus — você vs o bolão (pizzas) ──────────
        st.markdown('<div class="sh">🥧 Apostas — você vs o bolão</div>',
                    unsafe_allow_html=True)
 
        def _pie(col, title, dist, mypick, slug, flag=False):
            with col:
                st.markdown(f'<div style="font-weight:800;font-size:.9rem;text-align:center">'
                            f'{title}</div>', unsafe_allow_html=True)
                _mp = str(mypick).strip() if mypick else None
                if _mp:
                    _cnt = dist.get(_mp, 0)
                    _pct = (_cnt / _n * 100) if _n else 0
                    _fl  = F(_mp) + " " if flag else ""
                    st.markdown(
                        f'<div style="text-align:center;font-size:.8rem;margin:2px 0 6px">'
                        f'Você: <b>{_fl}{_mp}</b><br>'
                        f'<span style="color:#D6B864;font-weight:700">{_cnt} de {_n} '
                        f'({_pct:.0f}%)</span> colocaram igual</div>',
                        unsafe_allow_html=True)
                else:
                    st.markdown('<div style="text-align:center;font-size:.8rem;opacity:.5;'
                                'margin:2px 0 6px">Você não apostou</div>',
                                unsafe_allow_html=True)
                if dist:
                    _pal = ['#0D8587', '#D6B864', '#B2584E', '#123A56', '#5B8C5A', '#C9772E',
                            '#7B5EA7', '#3A7CA5', '#E0A458', '#8B4049', '#4C9F70', '#A0572E',
                            '#6A8EAE', '#C97B84', '#566B3F', '#9C6B30']
                    _pairs  = sorted(dist.items(), key=lambda x: -x[1])
                    _labels = [p[0] for p in _pairs]
                    _vals   = [p[1] for p in _pairs]
                    if len(_labels) <= len(_pal):
                        _colors = _pal[:len(_labels)]
                    else:
                        _colors = [f"hsl({(i * 137.508) % 360:.0f},62%,{45 if i % 2 == 0 else 60}%)"
                                   for i in range(len(_labels))]
                    _pull   = [0.14 if (_mp and l == _mp) else 0 for l in _labels]
                    _fig = go.Figure(go.Pie(
                        labels=_labels, values=_vals, pull=_pull, sort=False,
                        textinfo='none',
                        hovertemplate="%{label}<br>%{value} aposta(s) · %{percent}<extra></extra>",
                        marker=dict(colors=_colors, line=dict(color='rgba(0,0,0,.15)', width=1)),
                    ))
                    _fig.update_layout(height=230, margin=dict(l=4, r=4, t=4, b=4),
                                       showlegend=False, paper_bgcolor="rgba(0,0,0,0)")
                    st.plotly_chart(_fig, width='stretch',
                                    config={"displayModeBar": False}, key=f"pie_{slug}")
                    _leg = ('<div style="display:flex;flex-wrap:wrap;gap:2px 10px;'
                            'justify-content:center;margin-top:2px">')
                    for _lb, _vv, _cc in zip(_labels, _vals, _colors):
                        _pc = (_vv / _n * 100) if _n else 0
                        _b  = 'font-weight:700;' if (_mp and _lb == _mp) else ''
                        _fl = (F(_lb) + ' ') if flag else ''
                        _leg += (f'<span style="display:inline-flex;align-items:center;gap:4px;'
                                 f'font-size:.7rem;line-height:1.6;{_b}">'
                                 f'<span style="width:9px;height:9px;border-radius:2px;flex:none;'
                                 f'background:{_cc}"></span>{_fl}{_lb} '
                                 f'<span style="opacity:.55">{_pc:.0f}%</span></span>')
                    _leg += '</div>'
                    st.markdown(_leg, unsafe_allow_html=True)

                else:
                    st.markdown('<div style="text-align:center;opacity:.4;font-size:.78rem">'
                                'sem apostas</div>', unsafe_allow_html=True)
 
        _pcols = st.columns(3, gap="medium")
        _champ_pick = next(iter(bprnd.get('Final', set())), None)
        _pie(_pcols[0], "⚽ Artilheiro",     stats["art_dist"],   (bbb[0] if bbb else None), "art")
        _pie(_pcols[1], "🏅 Melhor Seleção Fase de Grupos", stats["mg_dist"],    (bbb[1] if bbb else None), "mg",   flag=True)
        _pie(_pcols[2], "🏆 Campeã",          stats["champ_dist"], _champ_pick,               "champ",flag=True)
 

# ══════════════════════════════════════════════════════════════════════
# TAB 5: SIMULAÇÃO
# ══════════════════════════════════════════════════════════════════════
if MOSTRAR_SIMULACAO:
    with T5:
        # ── Init state ────────────────────────────────────────────────
        for _k, _d in [("sim_rv", 0), ("sim_gr", {}), ("sim_mm", {}),
                        ("sim_res", None), ("sim_view", "grupos"),
                        ("sim_sel", None), ("sim_art", None)]:
            if _k not in st.session_state:
                st.session_state[_k] = _d
        _rv = st.session_state["sim_rv"]

        # ── View toggle + clear ───────────────────────────────────────
        _tv1, _tv2, _tvmc, _tv_sp, _tv3 = st.columns([2, 2, 2, 2, 1])
        with _tv1:
            if st.button("⚽ Fase de Grupos", width='stretch',
                         type="primary" if st.session_state["sim_view"] == "grupos" else "secondary",
                         key="sim_btn_grp"):
                st.session_state["sim_view"] = "grupos"; st.rerun()
        with _tv2:
            if st.button("🗓️ Mata-Mata", width='stretch',
                         type="primary" if st.session_state["sim_view"] == "mata-mata" else "secondary",
                         key="sim_btn_mm"):
                st.session_state["sim_view"] = "mata-mata"; st.rerun()
        with _tv3:
            if st.button("🗑 Apagar Simulação", width='stretch', key="sim_btn_clr"):
                st.session_state["sim_rv"]  += 1
                st.session_state["sim_gr"]   = {}
                st.session_state["sim_mm"]   = {}
                st.session_state["sim_res"]  = None
                st.session_state["sim_sel"]  = None
                st.session_state["sim_art"]  = None
                st.session_state["mc_run"]    = False
                st.rerun()
        st.markdown("---")

        with _tvmc:
            if st.button("🎲 Monte Carlo", width='stretch',
                        type="primary" if st.session_state["sim_view"] == "montecarlo" else "secondary",
                        key="sim_btn_mc"):
                st.session_state["sim_view"] = "montecarlo"; st.rerun()

        # ── Pre-compute bracket state (needed for both views) ─────────
        _mgr_cur = {**st.session_state["sim_gr"]}
        for _m, _v in gr.items():
            _mgr_cur[_m] = _v

        if _mgr_cur:
            _sim_r32r, *_ = build_r32(sort_st(calc_st(_mgr_cur)), t495)
        else:
            _sim_r32r = {_mi: ('?', '?') for _mi in range(16)}

        _sim_mmr_cur = dict(mmr)
        for _m, _ch in st.session_state["sim_mm"].items():
            if _m in mmr:
                continue

            if _ch == "t1":
                _sim_mmr_cur[_m] = (1, 0, None)
            elif _ch == "t2":
                _sim_mmr_cur[_m] = (0, 1, None)

        _, _sim_tn = build_bracket(_sim_r32r, _sim_mmr_cur)

        def _sim_w(m):
            t1, t2 = _sim_tn.get(m, (None, None))
            if m not in _sim_mmr_cur:
                return None
            r = _sim_mmr_cur[m]
            if r[0] > r[1]: return t1
            if r[1] > r[0]: return t2
            return t1 if r[2] == 'S1' else t2

        # Bracket column match ordering (derived from build_bracket pairings)
        _r32L = [0, 2, 1, 4, 8, 9, 10, 11]
        _octL = [17, 16, 21, 20]
        _qfL  = [24, 25]
        _sfL  = [28]
        _r32R = [3, 5, 6, 7, 13, 15, 12, 14]
        _octR = [18, 19, 22, 23]
        _qfR  = [26, 27]
        _sfR  = [29]

        _OPEN_MM = "— Em aberto —"

        def _mm_key(m):
            return f"mm_{m}_v{_rv}"

        def _future_matches_to_clear(m):
            if m in (_r32L + _r32R):
                return list(range(16, 32))

            if m in (_octL + _octR):
                return list(range(24, 32))

            if m in [24, 25, 26, 27]:
                return [28, 29, 30, 31]

            if m in [28, 29]:
                return [30, 31]

            return []

        def _clear_future_choices(m):
            for fm in _future_matches_to_clear(m):
                st.session_state["sim_mm"].pop(fm, None)
                st.session_state.pop(_mm_key(fm), None)

        def _set_mm_choice(m, widget_key):
            old = st.session_state["sim_mm"].get(m)
            val = st.session_state.get(widget_key, "open")

            if val == "open":
                st.session_state["sim_mm"].pop(m, None)
                new = None
            else:
                st.session_state["sim_mm"][m] = val
                new = val

            if old != new:
                _clear_future_choices(m)

        # ══ GRUPOS VIEW ═══════════════════════════════════════════════
        if st.session_state["sim_view"] == "grupos":
            _sg_sel = st.radio("", GL, horizontal=True,
                               label_visibility="collapsed", key="sim_grp_sel")
            _sg_games = [(m, gd, t1, t2)
                         for m, (gd, g, t1, t2) in enumerate(GROUP_FIXTURES)
                         if g == _sg_sel]
            #_live_st = sort_st(calc_st(_mgr_cur))
            #_grp_st  = _live_st.get(_sg_sel, [])

            _gc1, _gc2 = st.columns([5, 4])


            with _gc2:
                st.markdown(
                    f'<div class="sh">⚽ Jogos — Grupo {_sg_sel}</div>',
                    unsafe_allow_html=True)
                for _m, _gd, _t1, _t2 in _sg_games:
                    _ds = _gd.strftime("%d/%m")
                    if _m in gr:
                        _r = gr[_m]
                        st.markdown(
                            f'<div class="mr" style="opacity:.7">'
                            f'<div style="font-size:.68rem;opacity:.5;min-width:30px">{_ds}</div>'
                            f'<div class="mr-t">{FI(_t1)}{_t1}</div>'
                            f'<div class="mr-s">{_r[0]}–{_r[1]}</div>'
                            f'<div class="mr-t" style="text-align:right">{FI(_t2)}{_t2}</div>'
                            f'<span style="font-size:.62rem;opacity:.4;margin-left:4px">🔒</span>'
                            f'</div>',
                            unsafe_allow_html=True)
                        continue
                    _cur = st.session_state["sim_gr"].get(_m)
                    _t1v = "" if _cur is None else str(_cur[0])
                    _t2v = "" if _cur is None else str(_cur[1])
                    _mc1, _mc2, _mc3, _mc4, _mc5 = st.columns([0.6, 2.7, 0.6, 0.6, 2.7])
                    _mc1.markdown(
                        f'<div style="font-size:.65rem;opacity:.5;padding-top:7px;text-align:right">{_ds}</div>',
                        unsafe_allow_html=True)
                    _mc2.markdown(
                        f'<div style="font-size:.82rem;font-weight:700;text-align:right;'
                        f'padding-top:5px">{FI(_t1)}{_t1}</div>',
                        unsafe_allow_html=True)
                    _r1 = _mc3.text_input("", value=_t1v, max_chars=2,
                                          key=f"sg_{_m}_1_v{_rv}",
                                          label_visibility="collapsed",
                                          placeholder="–")
                    _r2 = _mc4.text_input("", value=_t2v, max_chars=2,
                                          key=f"sg_{_m}_2_v{_rv}",
                                          label_visibility="collapsed",
                                          placeholder="–")
                    _mc5.markdown(
                        f'<div style="font-size:.82rem;font-weight:700;padding-top:5px">'
                        f'{FI(_t2)}{_t2}</div>',
                        unsafe_allow_html=True)
                    
                    _x1 = (_r1 or "").strip(); _x2 = (_r2 or "").strip()
                    _s1 = int(_x1) if (_x1.isdigit() and int(_x1) <= 20) else None
                    _s2 = int(_x2) if (_x2.isdigit() and int(_x2) <= 20) else None
                    if _s1 is not None and _s2 is not None:
                        st.session_state["sim_gr"][_m] = (_s1, _s2)
                    else:
                        st.session_state["sim_gr"].pop(_m, None)

            _mgr_now = {**st.session_state["sim_gr"]}
            for _m, _v in gr.items():
                _mgr_now[_m] = _v
            _grp_st = sort_st(calc_st(_mgr_now)).get(_sg_sel, [])

            with _gc1:
                st.markdown(
                    f'<div class="sh">📊 Classificação — Grupo {_sg_sel}</div>',
                    unsafe_allow_html=True)
                _st_h = ("<table class='gt'><thead><tr>"
                         "<th></th><th style='text-align:left;min-width:110px'>Seleção</th>"
                         "<th>P</th><th>J</th><th>V</th><th>E</th><th>D</th>"
                         "<th>GP</th><th>GC</th><th>SG</th></tr></thead><tbody>")
                _st_r = ""
                for _pos, (_team, _td) in enumerate(_grp_st, 1):
                    _bg  = "rgba(13,133,135,.10)" if _pos <= 2 else ""
                    _fc  = "#22C55E" if _pos <= 2 else "#FB923C" if _pos == 3 else "inherit"
                    _st_r += (
                        f"<tr style='background:{_bg}'>"
                        f"<td style='font-weight:700;color:{_fc}'>{_pos}</td>"
                        f"<td style='text-align:left'>{FI(_team)}{_team}</td>"
                        f"<td style='font-weight:700'>{_td['pts']}</td>"
                        f"<td>{_td['played']}</td><td>{_td['w']}</td>"
                        f"<td>{_td['d']}</td><td>{_td['l']}</td>"
                        f"<td>{_td['gf']}</td><td>{_td['ga']}</td>"
                        f"<td>{_td['gf']-_td['ga']:+d}</td></tr>")
                st.markdown(f"{_st_h}{_st_r}</tbody></table>",
                            unsafe_allow_html=True)

        # ══ MATA-MATA VIEW ════════════════════════════════════════════
        elif st.session_state["sim_view"] == "mata-mata":
            _BH = 480

            def _tp(team, m):
                if not team or team == '?':
                    return '<div style="font-size:.68rem;opacity:.28;padding:1px 3px">?</div>'
                w = _sim_w(m)
                if w == team:
                    sty = "font-weight:700;color:#22C55E"
                elif w and w != team:
                    sty = "opacity:.25;text-decoration:line-through"
                else:
                    sty = "opacity:.85"
                short = team[:11] + "…" if len(team) > 11 else team
                return (f'<div style="font-size:.68rem;padding:1px 3px;'
                        f'white-space:nowrap;{sty}">{FI(team)}{short}</div>')

            def _mc(m):
                t1, t2 = _sim_tn.get(m, ('?', '?'))
                t1 = t1 or '?'; t2 = t2 or '?'
                return (f'<div style="border:1px solid rgba(128,128,128,.18);'
                        f'border-radius:3px;background:rgba(18,58,86,.04);'
                        f'margin:1px 0;padding:1px 1px">'
                        f'{_tp(t1, m)}'
                        f'<div style="font-size:.55rem;opacity:.25;text-align:center">vs</div>'
                        f'{_tp(t2, m)}'
                        f'</div>')

            def _bcol(mlist):
                return (f'<div style="display:flex;flex-direction:column;'
                        f'justify-content:space-around;flex:1;min-width:95px;'
                        f'height:{_BH}px">'
                        + ''.join(_mc(m) for m in mlist) + '</div>')

            _fw = _sim_w(31) or '?'
            _center = (
                f'<div style="display:flex;flex-direction:column;justify-content:center;'
                f'align-items:center;min-width:80px;height:{_BH}px;gap:3px;text-align:center">'
                f'<div style="font-size:1.3rem">🏆</div>'
                f'<div style="font-size:.72rem;font-weight:700;color:#D6B864">'
                f'{FI(_fw)}{_fw}</div>'
                f'<div style="font-size:.58rem;opacity:.45;margin-top:8px">Final</div>'
                f'{_mc(31)}'
                f'<div style="font-size:.58rem;opacity:.35;margin-top:5px">3º Lugar</div>'
                f'{_mc(30)}'
                f'</div>')

            st.markdown(
                f'<div style="display:flex;align-items:stretch;width:100%;'
                f'overflow-x:auto;gap:2px;margin:6px 0">'
                f'{_bcol(_r32L)}{_bcol(_octL)}{_bcol(_qfL)}{_bcol(_sfL)}'
                f'{_center}'
                f'{_bcol(_sfR)}{_bcol(_qfR)}{_bcol(_octR)}{_bcol(_r32R)}'
                f'</div>',
                unsafe_allow_html=True)

            st.caption("Selecione os vencedores abaixo — o chaveamento atualiza automaticamente.")
            _KO_SIM = [
                ('R32',     '🔵 Rodada de 32',       _r32L + _r32R),
                ('Oitavas', '🟢 Oitavas de Final',   _octL + _octR),
                ('Quartas', '🟡 Quartas de Final',   [24, 25, 26, 27]),
                ('Semi',    '🔴 Semifinais',          [28, 29]),
                ('3o Lugar','🟠 3º Lugar',            [30]),
                ('Final',   '⭐ Final',              [31]),
            ]
            for _rk, _rl, _rm in _KO_SIM:
                with st.expander(_rl, expanded=(_rk == 'R32')):
                    _nc = 4 if len(_rm) >= 4 else 2 if len(_rm) >= 2 else 1
                    _mmcols = st.columns(_nc)
                    for _ci, _m in enumerate(_rm):
                        with _mmcols[_ci % _nc]:
                            _t1k, _t2k = _sim_tn.get(_m, ('?', '?'))
                            _t1k = _t1k or '?'; _t2k = _t2k or '?'
                            if _m in mmr:
                                _rr = mmr[_m]
                                _wk = (_t1k if _rr[0] > _rr[1]
                                       else (_t2k if _rr[1] > _rr[0]
                                             else (_t1k if _rr[2] == 'S1' else _t2k)))
                                st.markdown(
                                    f'<div style="font-size:.75rem;padding:3px 0">'
                                    f'{KO[_m][1]}: 🔒 <b>{_wk}</b></div>',
                                    unsafe_allow_html=True)
                            elif _t1k == '?' or _t2k == '?':
                                st.markdown(
                                    f'<div style="font-size:.72rem;opacity:.35;padding:3px 0">'
                                    f'{KO[_m][1]}: ⏳</div>',
                                    unsafe_allow_html=True)
                            else:
                                _saved = st.session_state["sim_mm"].get(_m, "open")

                                if _saved not in ("open", "t1", "t2"):
                                    _saved = "open"

                                _opts = ["open", "t1", "t2"]
                                _idx = _opts.index(_saved)
                                _wkey = _mm_key(_m)

                                st.radio(
                                    f"{KO[_m][1]}",
                                    options=_opts,
                                    index=_idx,
                                    key=_wkey,
                                    format_func=lambda x, a=_t1k, b=_t2k: (
                                        _OPEN_MM if x == "open" else a if x == "t1" else b
                                    ),
                                    on_change=_set_mm_choice,
                                    args=(_m, _wkey),
                                    horizontal=False,
                                )

        # ══ MONTECARLO VIEW ════════════════════════════════════════════
        elif st.session_state["sim_view"] == "montecarlo":
            _groups_done = len(gr) >= len(GROUP_FIXTURES)
            if not _groups_done:
                st.info("🎲 A simulação de Monte Carlo fica disponível quando a "
                        "fase de grupos terminar (aí os 32 classificados e os "
                        "confrontos da rodada de 32 estão definidos).")
            else:
                st.markdown(
                    '<div style="font-size:.86rem;opacity:.75;margin-bottom:8px">'
                    'Sorteia o mata-mata milhares de vezes a partir de um modelo de '
                    'força (ranking FIFA + forma nos grupos + opinião dos apostadores + '
                    'momentum das vitórias reais no mata-mata), respeitando sempre o '
                    'gabarito. Jogos já disputados ficam travados.</div>',
                    unsafe_allow_html=True)
 
                # ── botão Rodar / Limpar (alterna a cada clique) ──
                _mc_on = st.session_state.get("mc_run", False)
                if st.button(("🧹 Limpar simulação" if _mc_on
                              else f"🎲 Rodar simulação ({MC_N_SIMS:,} cenários)".replace(",", ".")),
                             width='stretch', type="primary", key="mc_run_btn"):
                    st.session_state["mc_run"] = not _mc_on
                    st.rerun()
 
                if st.session_state.get("mc_run"):
                    with st.spinner(f"Simulando {MC_N_SIMS:,} cenários…".replace(",", ".")):
                        _R = compute_mc(_fp, bettors)
                    if not _R:
                        st.warning("Não foi possível montar o chaveamento — confira se os "
                                   "grupos estão completos.")
                    else:
                        if not _R.get("mm_ok", True):
                            st.warning("⚠️ A pontuação simulada do mata-mata divergiu do "
                                       "placar atual do app — os números podem estar levemente "
                                       "deslocados. (Me avise se aparecer.)")
 
                        _BH = 470
                        def _short(t):
                            t = t or '?'
                            return (t[:11] + "…") if len(t) > 11 else t
 
                        def _cell_lines(lines):
                            inner = ""
                            for team, pct, kind in lines:
                                if kind == 'win':          # vencedor de jogo já decidido
                                    inner += (f'<div style="font-size:.66rem;font-weight:700;'
                                              f'color:#22C55E;white-space:nowrap;padding:1px 3px">'
                                              f'🔒 {FI(team)}{_short(team)}</div>')
                                elif kind == 'lose':       # perdedor de jogo já decidido
                                    inner += (f'<div style="font-size:.66rem;white-space:nowrap;'
                                              f'padding:1px 3px;opacity:.38;'
                                              f'text-decoration:line-through">'
                                              f'{FI(team)}{_short(team)}</div>')
                                else:                      # jogo em aberto (probabilístico)
                                    fav = (kind == 'fav') or (pct is not None and pct >= 0.5)
                                    if fav:                # quem avança: dourado escuro, opacidade cheia
                                        inner += (f'<div style="font-size:.66rem;white-space:nowrap;'
                                                  f'padding:1px 3px;font-weight:700;color:#8C6B1A">'
                                                  f'{FI(team)}{_short(team)} '
                                                  f'<span style="opacity:.75;font-size:.6rem">'
                                                  f'{(pct or 0)*100:.0f}%</span></div>')
                                    else:                  # o outro time, mais apagado (como antes)
                                        op = 0.35 + 0.6 * (pct or 0)
                                        inner += (f'<div style="font-size:.66rem;white-space:nowrap;'
                                                  f'padding:1px 3px;opacity:{op:.2f}">'
                                                  f'{FI(team)}{_short(team)} '
                                                  f'<span style="opacity:.7;font-size:.6rem">'
                                                  f'{(pct or 0)*100:.0f}%</span></div>')
                            return (f'<div style="border:1px solid rgba(128,128,128,.18);'
                                    f'border-radius:3px;background:rgba(18,58,86,.04);'
                                    f'margin:1px 0;padding:1px 1px">{inner}</div>')
 
                        def _pair_of(m):
                            return _R['r32'][m] if m < 16 else _R['tn_real'].get(m)
 
                        def _mc_cell(m):
                            dist = _R['mw'].get(m, {}); tot = sum(dist.values()) or 1
                            # jogo já decidido (resultado real) -> mostra OS DOIS times
                            if m in _R['locked']:
                                wn = _R['locked'][m]
                                _pp = _pair_of(m)
                                if _pp and _pp[0] and _pp[1] and '?' not in _pp:
                                    return _cell_lines([(t, None, 'win' if t == wn else 'lose')
                                                        for t in _pp])
                                return _cell_lines([(wn, None, 'win')])
                            # R32 ainda em aberto -> os dois times com probabilidade
                            if m < 16:
                                t1, t2 = _R['r32'][m]
                                return _cell_lines([(t1, dist.get(t1, 0) / tot, 'r32'),
                                                    (t2, dist.get(t2, 0) / tot, 'r32')])
                            # fases adiante em aberto -> favorito a avançar
                            top = sorted(dist.items(), key=lambda x: -x[1])[:1]
                            return _cell_lines([(t, c / tot, 'fav') for t, c in top]
                                               or [('?', 0, 'fav')])
 
                        def _mc_col(ms):
                            return (f'<div style="display:flex;flex-direction:column;'
                                    f'justify-content:space-around;flex:1;min-width:104px;'
                                    f'height:{_BH}px">' + ''.join(_mc_cell(m) for m in ms) + '</div>')
 
                        _champ_top = sorted(_R['champ'].items(), key=lambda x: -x[1])
                        _fav_t, _fav_c = (_champ_top[0] if _champ_top else ('?', 0))
                        _center = (
                            f'<div style="display:flex;flex-direction:column;justify-content:center;'
                            f'align-items:center;min-width:96px;height:{_BH}px;gap:3px;text-align:center">'
                            f'<div style="font-size:1.3rem">🏆</div>'
                            f'<div style="font-size:.7rem;font-weight:700;color:#D6B864">'
                            f'{FI(_fav_t)}{_short(_fav_t)}</div>'
                            f'<div style="font-size:.62rem;opacity:.6">{_fav_c/_R["n"]*100:.0f}% campeã</div>'
                            f'<div style="font-size:.55rem;opacity:.4;margin-top:8px">Final</div>'
                            f'{_mc_cell(31)}'
                            f'<div style="font-size:.55rem;opacity:.35;margin-top:4px">3º Lugar</div>'
                            f'{_mc_cell(30)}</div>')
                        st.markdown(
                            f'<div style="display:flex;align-items:stretch;width:100%;'
                            f'overflow-x:auto;gap:2px;margin:6px 0">'
                            f'{_mc_col(_r32L)}{_mc_col(_octL)}{_mc_col(_qfL)}{_mc_col(_sfL)}'
                            f'{_center}'
                            f'{_mc_col(_sfR)}{_mc_col(_qfR)}{_mc_col(_octR)}{_mc_col(_r32R)}'
                            f'</div>', unsafe_allow_html=True)
                        st.caption("Cada confronto mostra os dois times; nos jogos ainda em aberto, "
                                   "a % é a chance de avançar. 🔒 = resultado real já definido.")
 
                        # ---- Corrida pela taça ----
                        st.markdown('<div class="sh">🏆 Corrida pela taça</div>', unsafe_allow_html=True)
                        _mx = (_champ_top[0][1] / _R['n']) if _champ_top else 1
                        _bars = ""
                        for t, c in _champ_top[:12]:
                            p = c / _R['n']
                            _bars += (
                                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">'
                                f'<div style="width:150px;font-size:.78rem;white-space:nowrap;overflow:hidden;'
                                f'text-overflow:ellipsis">{FI(t)}{t}</div>'
                                f'<div style="flex:1;background:rgba(128,128,128,.12);border-radius:5px;height:18px">'
                                f'<div style="width:{p/_mx*100:.0f}%;background:#D6B864;height:100%;'
                                f'border-radius:5px;min-width:3px"></div></div>'
                                f'<div style="width:52px;text-align:right;font-size:.75rem;font-weight:700">'
                                f'{p*100:.1f}%</div></div>')
                        st.markdown(_bars, unsafe_allow_html=True)
 
                        # ---- Ranking dos apostadores (se seguir o Monte Carlo) ----
                        st.markdown('<div class="sh">👤 Ranking esperado dos apostadores</div>',
                                    unsafe_allow_html=True)
                        st.caption("Ordenado pela chance de ganhar o bolão. "
                                   "Pontos com faixa provável (p10–p90) e prêmio esperado no rateio 70/20/10.")
                        _np = __import__("numpy")
                        _ordb = _np.argsort(-_R['win_pct'])
                        _hc = st.columns([0.5, 3, 1.2, 1.2, 2, 1.4])
                        for _c, _h in zip(_hc, ["#", "Participante", "🏆 Vitória", "🥇🥈🥉 Pódio",
                                                "Pontos (p10–p90)", "Prêmio méd."]):
                            _c.markdown(f'<div style="font-size:.68rem;font-weight:700;opacity:.5;'
                                        f'text-transform:uppercase;letter-spacing:.6px">{_h}</div>',
                                        unsafe_allow_html=True)
                        for _rk, _bi in enumerate(_ordb, 1):
                            _bi = int(_bi)
                            _rc = st.columns([0.5, 3, 1.2, 1.2, 2, 1.4])
                            _rc[0].markdown(f'<div style="font-size:.8rem;opacity:.5">{MEDALS.get(_rk, _rk)}</div>',
                                            unsafe_allow_html=True)
                            _rc[1].markdown(f'<div style="font-size:.82rem;font-weight:600">{_R["names"][_bi]}</div>',
                                            unsafe_allow_html=True)
                            _rc[2].markdown(f'<div style="font-size:.82rem;font-weight:800;color:#D6B864">'
                                            f'{_R["win_pct"][_bi]*100:.1f}%</div>', unsafe_allow_html=True)
                            _rc[3].markdown(f'<div style="font-size:.82rem">{_R["podium_pct"][_bi]*100:.1f}%</div>',
                                            unsafe_allow_html=True)
                            _rc[4].markdown(f'<div style="font-size:.8rem">{_R["exp_pts"][_bi]:.0f} '
                                            f'<span style="opacity:.5;font-size:.72rem">'
                                            f'({_R["p10"][_bi]:.0f}–{_R["p90"][_bi]:.0f})</span></div>',
                                            unsafe_allow_html=True)
                            _rc[5].markdown(f'<div style="font-size:.8rem;font-weight:700">'
                                            f'R$ {_R["exp_prize"][_bi]:,.0f}</div>'.replace(",", "."),
                                            unsafe_allow_html=True)

        if st.session_state["sim_view"] != "montecarlo":
        # ══ Bônus simulado ════════════════════════════════════════════
            with st.expander("🎯 Simular Bônus (Melhor Seleção & Artilheiro)", expanded=False):
                _all_teams_sim = sorted({t for g_data in GROUPS_DATA.values() for t in g_data})
                _bon_c1, _bon_c2 = st.columns(2)

                with _bon_c1:

                    st.markdown(
                        '<div style="font-size:.8rem;font-weight:700;opacity:.6;'
                        'text-transform:uppercase;letter-spacing:.7px;margin-bottom:4px">'
                        '🏅 Melhor Seleção da Fase de Grupos</div>',
                        unsafe_allow_html=True)

                    if br and br[1]:
                        st.session_state["sim_sel"] = None
                        st.markdown(
                            f'<div style="margin-top:6px;font-size:.9rem;font-weight:600">'
                            f'🔒 {FI(br[1])}{br[1]}</div>'
                            f'<div style="font-size:.7rem;opacity:.45;margin-top:2px">'
                            f'definido no gabarito</div>',
                            unsafe_allow_html=True)
                    else:

                        _sel_opts = ["(não simular)"] + _all_teams_sim
                        _cur_sel  = st.session_state.get("sim_sel")
                        _sel_idx  = (_sel_opts.index(_cur_sel)
                                    if _cur_sel in _sel_opts else 0)
                        _sel_choice = st.selectbox(
                            "Melhor Seleção",
                            options=_sel_opts,
                            index=_sel_idx,
                            key=f"sim_sel_box_v{_rv}",
                            label_visibility="collapsed",
                        )
                        if _sel_choice != "(não simular)":
                            st.session_state["sim_sel"] = _sel_choice
                            st.markdown(
                                f'<div style="margin-top:6px;font-size:.9rem;font-weight:600">'
                                f'{FI(_sel_choice)}{_sel_choice}</div>',
                                unsafe_allow_html=True)
                        else:
                            st.session_state["sim_sel"] = None
                            st.markdown(
                                '<div style="margin-top:6px;font-size:.8rem;opacity:.4">—</div>',
                                unsafe_allow_html=True)

                with _bon_c2:
                    st.markdown(
                        '<div style="font-size:.8rem;font-weight:700;opacity:.6;'
                        'text-transform:uppercase;letter-spacing:.7px;margin-bottom:4px">'
                        '⚽ Artilheiro do Torneio</div>',
                        unsafe_allow_html=True)
                    
                    if br and br[0]:
                        st.session_state["sim_art"] = None
                        _alk = next((c for j, c in JOGADORES_ARTILHEIRO if j == br[0]), None)
                        st.markdown(
                            f'<div style="margin-top:6px;font-size:.9rem;font-weight:600">'
                            f'🔒 {FI(_alk) if _alk else ""}{br[0]}</div>'
                            f'<div style="font-size:.7rem;opacity:.45;margin-top:2px">'
                            f'definido no gabarito</div>',
                            unsafe_allow_html=True)

                    elif JOGADORES_ARTILHEIRO:
                        _art_names = [j for j, _ in JOGADORES_ARTILHEIRO]
                        _art_opts  = ["(não simular)"] + _art_names
                        _cur_art   = st.session_state.get("sim_art")
                        _art_idx   = (_art_opts.index(_cur_art)
                                    if _cur_art in _art_opts else 0)
                        _art_choice = st.selectbox(
                            "Artilheiro",
                            options=_art_opts,
                            index=_art_idx,
                            key=f"sim_art_box_v{_rv}",
                            label_visibility="collapsed",
                        )
                        if _art_choice != "(não simular)":
                            st.session_state["sim_art"] = _art_choice
                            _art_country = next(
                                (c for j, c in JOGADORES_ARTILHEIRO if j == _art_choice), None)
                            st.markdown(
                                f'<div style="margin-top:6px;font-size:.9rem;font-weight:600">'
                                f'{FI(_art_country) if _art_country else ""}{_art_choice}</div>',
                                unsafe_allow_html=True)
                        else:
                            st.session_state["sim_art"] = None
                            st.markdown(
                                '<div style="margin-top:6px;font-size:.8rem;opacity:.4">—</div>',
                                unsafe_allow_html=True)
                    else:
                        st.markdown(
                            '<div style="font-size:.82rem;opacity:.5;padding:4px 0">'
                            'Lista de artilheiros ainda não configurada.<br>'
                            'Adicione jogadores em <code>JOGADORES_ARTILHEIRO</code>.</div>',
                            unsafe_allow_html=True)

            # ══ Simular button (ambas as views) ═══════════════════════════
            st.markdown("---")
            if st.button("▶ Calcular Ranking Simulado", width='stretch', key="sim_calc"):
                _mgr_c = dict(gr)
                for _m, _vv in st.session_state["sim_gr"].items():
                    if _m not in _mgr_c:
                        _mgr_c[_m] = _vv
                _mmmr_c = dict(mmr)

                for _m, _ch in st.session_state["sim_mm"].items():
                    if _m in _mmmr_c:
                        continue

                    if _ch == "t1":
                        _mmmr_c[_m] = (1, 0, None)
                    elif _ch == "t2":
                        _mmmr_c[_m] = (0, 1, None)
                # Bônus simulado: sobrepõe apenas o que o usuário escolheu
                _sim_br = list(br) if br else [None, None]
                if st.session_state.get("sim_art"):
                    _sim_br[0] = st.session_state["sim_art"]
                if st.session_state.get("sim_sel"):
                    _sim_br[1] = st.session_state["sim_sel"]
                _sim_br = tuple(_sim_br)
                _sim_r = []
                for _nm, _gb, _bb, _xm, _rsc, _ in bettors:
                    _sc2 = score_all(_gb, _xm, _bb, _mgr_c, _mmmr_c, _sim_br, t495)
                    _d2  = _sc2["total"] - _rsc["total"]
                    _sim_r.append((_nm, _sc2["total"], _rsc["total"], _d2, _sc2))

                # Ordena o ranking simulado pelos mesmos critérios oficiais
                _sim_r.sort(key=lambda x: ranking_score_key(x[0], x[4]))

                # Mantém o formato antigo que a tela já espera:
                # nome, pontuação simulada, pontuação atual, delta
                st.session_state["sim_res"] = [
                    (_nm, _sp, _rp, _dlt)
                    for _nm, _sp, _rp, _dlt, _sc2 in _sim_r
                ]

            # ══ Ranking resultado ═════════════════════════════════════════
            if st.session_state.get("sim_res"):
                st.markdown('<div class="sh">🏆 Ranking Simulado</div>',
                            unsafe_allow_html=True)
                _hc = st.columns([0.5, 4, 1.5, 1.5, 1.5])
                for _c, _h in zip(_hc, ["#", "Participante", "Simulado", "Atual", "Δ"]):
                    _c.markdown(
                        f'<div style="font-size:.7rem;font-weight:700;opacity:.5;'
                        f'text-transform:uppercase;letter-spacing:.8px">{_h}</div>',
                        unsafe_allow_html=True)
                for _rk, (_nm, _sp, _rp, _dlt) in enumerate(st.session_state["sim_res"], 1):
                    _dc = "#22C55E" if _dlt > 0 else "#F87171" if _dlt < 0 else "inherit"
                    _ds2 = f"+{_dlt}" if _dlt > 0 else str(_dlt)
                    _rc2 = st.columns([0.5, 4, 1.5, 1.5, 1.5])
                    _rc2[0].markdown(
                        f'<div style="font-size:.82rem;opacity:.5">{MEDALS.get(_rk, _rk)}</div>',
                        unsafe_allow_html=True)
                    _rc2[1].markdown(
                        f'<div style="font-size:.85rem;font-weight:600">{_nm}</div>',
                        unsafe_allow_html=True)
                    _rc2[2].markdown(
                        f'<div style="font-size:.9rem;font-weight:900;color:#D6B864">{_sp}</div>',
                        unsafe_allow_html=True)
                    _rc2[3].markdown(
                        f'<div style="font-size:.85rem;opacity:.55">{_rp}</div>',
                        unsafe_allow_html=True)
                    _rc2[4].markdown(
                        f'<div style="font-size:.85rem;font-weight:700;color:{_dc}">{_ds2}</div>',
                        unsafe_allow_html=True)

# ── Footer
st.markdown("---")
ft_cols = st.columns([1,3,1])
with ft_cols[0]:
    if LOGO_TORI:
        st.image(LOGO_TORI, width=100)
with ft_cols[1]:
    st.markdown(
        "<div style='text-align:center;color:#5C5F62;font-size:.8rem;padding:.5rem 0'>"
        "Bolão Copa do Mundo 2026 · Turim MFO &nbsp;·&nbsp; "
        "Desenvolvido Em Casa &nbsp;·&nbsp; 🐂 Rumo ao Hexa!</div>",
        unsafe_allow_html=True)
with ft_cols[2]:
    if LOGO_TURIM_AZUL:
        st.image(LOGO_TURIM_AZUL, width=100)