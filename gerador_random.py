from pathlib import Path
from copy import copy
import random
from openpyxl import load_workbook

template_path = Path(r"C:\Users\pbrandao\Projetos\bolao-copa-2026\apostas\Bolao_Copa2026_TurimMFO_Pedro Brandão.xlsx")
output_dir = Path(r"C:\Users\pbrandao\Projetos\bolao-copa-2026\apostas")

nomes = [
    "Joao Silva", "Maria Souza", "Carlos Lima", "Ana Costa", "Bruno Rocha",
    "Lucas Pereira", "Fernanda Alves", "Rafael Gomes", "Juliana Martins", "Pedro Santos",
    "Mariana Ferreira", "Gustavo Ribeiro", "Camila Barros", "Thiago Almeida", "Beatriz Nunes",
    "Rodrigo Teixeira", "Larissa Carvalho", "Felipe Mendes", "Amanda Castro", "Daniel Moreira",
    "Patricia Lopes", "Eduardo Araujo", "Renata Cardoso", "Marcelo Duarte", "Isabela Freitas",
    "Vinicius Oliveira", "Carolina Batista", "Andre Farias", "Leticia Ramos", "Henrique Moura",
    "Natalia Correia", "Leonardo Vieira", "Bianca Monteiro", "Diego Campos", "Sofia Reis",
    "Ricardo Neves", "Gabriela Tavares", "Matheus Cunha", "Luiza Borges", "Caio Fonseca",
    "Clara Pires", "Igor Machado", "Vanessa Andrade", "Murilo Brito", "Tatiana Guedes",
    "Arthur Melo", "Helena Diniz", "Vitor Sales", "Priscila Xavier", "Samuel Dias",
]

# Duplica até chegar em 100 nomes únicos
nomes_100 = []
i = 1
while len(nomes_100) < 100:
    for nome in nomes:
        nomes_100.append(f"{nome} {i}")
        if len(nomes_100) == 100:
            break
    i += 1

for nome in nomes_100:
    wb = load_workbook(template_path)
    ws = wb["Apostas - Grupos"]

    # Nome do participante na célula A5
    ws["A5"] = nome

    # B até EO = colunas 2 até 145 = 144 valores
    for col in range(2, 146):
        ws.cell(row=5, column=col).value = random.randint(0, 5)

    output_name = f"Bolao_Copa2026_TurimMFO_{nome}.xlsx"
    output_path = output_dir / output_name

    wb.save(output_path)
    wb.close()

print("100 planilhas criadas com sucesso.")