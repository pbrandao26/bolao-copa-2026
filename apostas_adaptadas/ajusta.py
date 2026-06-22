#!/usr/bin/env python3
"""
============================================================================
AJUSTA — Bolão Copa 2026 Turim MFO
============================================================================
Popula a pasta `apostas_adaptadas/` a partir de `apostas/`:

  • Corrige a aba T495 de CADA planilha de participante, gravando os
    confrontos certos (da tabela `matchups_r32_copa_2026.xlsx`,
    aba 'R32 Matchups Certo').

  • Para os participantes NÃO afetados (chave caiu numa linha que já
    estava certa): nada mais muda além da T495 ficar idêntica à correta
    (valores iguais → sem efeito visível).

  • Para os AFETADOS: pinta de VINHO ESCURO, na aba 'Apostas - Mata-Mata',
    a célula do NOME da seleção adversária (3º colocado) de cada confronto
    da R32 que mudou. A célula é uma fórmula que lê a T495, então após a
    correção ela recalcula sozinha para o adversário certo ao abrir no Excel.

Como roda
---------
    python ajusta.py
(de dentro da pasta `apostas_adaptadas/`, com `apostas/` e
 `matchups_r32_copa_2026.xlsx` acessíveis pelos caminhos abaixo)

Ajuste PASTA_ENTRADA / PASTA_SAIDA / PLANILHA_CORRETA se necessário.
============================================================================
"""

import os
import glob
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── CONFIG ──────────────────────────────────────────────────────────────
# Por padrão assume estrutura:  repo/apostas/*.xlsx
#                               repo/apostas_adaptadas/ajusta.py  (este)
#                               repo/matchups_r32_copa_2026.xlsx
PASTA_ENTRADA    = "../apostas"
PASTA_SAIDA      = "."                       # a própria apostas_adaptadas
PLANILHA_CORRETA = "../apostas/matchups_r32_copa_2026.xlsx"
ABA_CORRETA      = "R32 Matchups Certo"

ABA_T495   = "T495"            # aba a corrigir (nome real procurado case-insensitive)
ABA_CALC   = "Calc"
CELULA_CHAVE = "B88"
ABA_MM     = "Apostas - Mata-Mata"
MM_ROW_SEL = 5                 # linha dos nomes das seleções

# Vinho escuro
COR_VINHO = "FF0000"           # ARGB sem alpha; PatternFill adiciona FF

# Posição na T495 (0..7, ordem das colunas 1A,1B,1D,1E,1G,1I,1K,1L)
# → coluna (1-based) do 3º colocado na aba Apostas - Mata-Mata.
# (mapeado a partir da fórmula VLOOKUP(...'T495'...) de cada jogo)
POS_TO_MM_COL = {
    0: 15,   # 1A → M79 → col O
    1: 27,   # 1B → M85 → col AA
    2: 19,   # 1D → M81 → col S
    3: 5,    # 1E → M74 → col E
    4: 21,   # 1G → M82 → col U
    5: 11,   # 1I → M77 → col K
    6: 31,   # 1K → M87 → col AE
    7: 17,   # 1L → M80 → col Q
}
# Rótulo do jogo (só para log)
POS_TO_GAME = {0: "M79", 1: "M85", 2: "M81", 3: "M74",
               4: "M82", 5: "M77", 6: "M87", 7: "M80"}
COLUNAS_CONFRONTO = ["1A", "1B", "1D", "1E", "1G", "1I", "1K", "1L"]


# ── CARREGAR TABELA CORRETA ────────────────────────────────────────────
def carregar_t495_correta(caminho, aba):
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb[aba]
    tabela = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        chave = str(row[0]).strip()
        if len(chave) == 8:
            tabela[chave] = [str(v).strip() if v is not None else "?" for v in row[1:9]]
    return tabela


def achar_aba(wb, nome_alvo):
    """Acha aba pelo nome (case-insensitive, ignora espaços)."""
    alvo = nome_alvo.strip().upper()
    return next((s for s in wb.sheetnames if s.strip().upper() == alvo), None)


# ── PROCESSAR UMA PLANILHA ─────────────────────────────────────────────
def processar(caminho_in, caminho_out, t495_correta, fill):
    nome_arq = os.path.basename(caminho_in)
    nome = (nome_arq.replace("Bolao_Copa2026_TurimMFO_", "")
                    .replace(".xlsx", "").replace("_", " ").strip())

    # data_only=False → preserva fórmulas e formatação ao salvar
    wb = load_workbook(caminho_in, data_only=False)

    aba_t495 = achar_aba(wb, ABA_T495)
    if not aba_t495:
        return {"nome": nome, "status": "SEM_T495"}

    # ── 1. Ler a chave do participante (precisa do valor calculado).
    # Como abrimos sem data_only, B88 pode vir como fórmula. Reabrimos
    # só pra ler o valor calculado da chave.
    chave = None
    try:
        wbv = load_workbook(caminho_in, data_only=True, read_only=True)
        ab_calc_v = achar_aba(wbv, ABA_CALC)
        if ab_calc_v:
            chave = wbv[ab_calc_v][CELULA_CHAVE].value
        wbv.close()
    except Exception:
        pass
    chave = str(chave).strip() if chave else None

    # ── 2. Corrigir a aba T495 inteira (sobrescreve valores literais).
    ws_t = wb[aba_t495]
    # Descobrir a linha de início dos dados (primeira linha cuja col A tem 8 letras)
    _linha_ini = None
    for r in range(1, ws_t.max_row + 1):
        v = ws_t.cell(row=r, column=1).value
        if v and len(str(v).strip()) == 8 and str(v).strip().isalpha():
            _linha_ini = r
            break
    if _linha_ini is None:
        return {"nome": nome, "status": "T495_VAZIA"}

    n_corrigidas = 0
    for r in range(_linha_ini, ws_t.max_row + 1):
        k = ws_t.cell(row=r, column=1).value
        if not k:
            continue
        k = str(k).strip()
        certo = t495_correta.get(k)
        if not certo:
            continue
        for i, val in enumerate(certo):           # colunas B..I (2..9)
            cell = ws_t.cell(row=r, column=2 + i)
            if str(cell.value).strip() != val:
                cell.value = val
                n_corrigidas += 1

    # ── 3. Determinar se ESTE participante foi afetado e pintar.
    mudou = []
    if chave and len(chave) == 8 and chave in t495_correta:
        # linha errada = o que a planilha tinha ANTES (lemos de uma cópia
        # limpa? não — já corrigimos em memória. Mas a tabela correta é a
        # referência; o "errado" é a tabela embutida original).
        # Para saber se mudou, comparamos a linha ORIGINAL da T495 do arquivo
        # com a correta. Reabrimos os valores originais:
        try:
            wbo = load_workbook(caminho_in, data_only=True, read_only=True)
            ab_t_o = achar_aba(wbo, ABA_T495)
            linha_errada = None
            if ab_t_o:
                wso = wbo[ab_t_o]
                for row in wso.iter_rows(min_row=1, values_only=True):
                    if row and row[0] and str(row[0]).strip() == chave:
                        linha_errada = [str(v).strip() if v is not None else "?"
                                        for v in row[1:9]]
                        break
            wbo.close()
        except Exception:
            linha_errada = None

        certa = t495_correta[chave]
        if linha_errada:
            for pos, (ve, vc) in enumerate(zip(linha_errada, certa)):
                if ve != vc:
                    mudou.append(pos)

    # Pintar as células do 3º colocado dos jogos que mudaram
    if mudou:
        aba_mm = achar_aba(wb, ABA_MM)
        if aba_mm:
            ws_mm = wb[aba_mm]
            for pos in mudou:
                col = POS_TO_MM_COL.get(pos)
                if col:
                    ws_mm.cell(row=MM_ROW_SEL, column=col).fill = fill

    # ── 4. Salvar
    wb.save(caminho_out)

    status = "AFETADO" if mudou else "OK"
    return {"nome": nome, "status": status, "chave": chave,
            "n_corrigidas": n_corrigidas,
            "jogos": [POS_TO_GAME[p] for p in mudou]}


# ── MAIN ────────────────────────────────────────────────────────────────
def main():
    print("=" * 72)
    print("AJUSTA — corrige T495 e pinta confrontos alterados")
    print("=" * 72)

    if not os.path.isdir(PASTA_ENTRADA):
        print(f"❌ Pasta de entrada não encontrada: {PASTA_ENTRADA}")
        return
    os.makedirs(PASTA_SAIDA, exist_ok=True)

    t495_correta = carregar_t495_correta(PLANILHA_CORRETA, ABA_CORRETA)
    print(f"Tabela correta: {len(t495_correta)} chaves.\n")

    fill = PatternFill(start_color="FF" + COR_VINHO,
                       end_color="FF" + COR_VINHO, fill_type="solid")

    arquivos = sorted(glob.glob(os.path.join(PASTA_ENTRADA, "*.xlsx")))
    arquivos = [a for a in arquivos if not os.path.basename(a).startswith("~$")]
    # ignora a consolidada (não é planilha de participante individual)
    arquivos = [a for a in arquivos if "Consolidada" not in os.path.basename(a)]
    print(f"{len(arquivos)} planilha(s) de participante.\n")

    afetados, ok, problemas = [], [], []
    for caminho in arquivos:
        nome_arq = os.path.basename(caminho)
        destino = os.path.join(PASTA_SAIDA, nome_arq)
        try:
            r = processar(caminho, destino, t495_correta, fill)
        except Exception as e:
            problemas.append({"nome": nome_arq, "status": "ERRO", "detalhe": str(e)})
            continue
        if r["status"] == "AFETADO":
            afetados.append(r)
        elif r["status"] == "OK":
            ok.append(r)
        else:
            problemas.append(r)

    print("─" * 72)
    print(f"RESUMO: {len(afetados)} afetado(s) pintado(s) · {len(ok)} ok "
          f"(só T495 corrigida) · {len(problemas)} com problema")
    print("─" * 72)

    if afetados:
        print("\n🔴 AFETADOS (células pintadas de vinho na aba Mata-Mata):\n")
        for r in afetados:
            print(f"  ● {r['nome']:32s} chave {r['chave']} — jogos: "
                  f"{', '.join(r['jogos'])}")

    if problemas:
        print("\n⚠️  PROBLEMAS:\n")
        for r in problemas:
            print(f"  ● {r['nome']:32s} [{r['status']}] {r.get('detalhe','')}")

    print(f"\n✅ Planilhas adaptadas salvas em: {os.path.abspath(PASTA_SAIDA)}")


if __name__ == "__main__":
    main()