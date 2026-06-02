#!/usr/bin/env python3
"""
Consolidador de Apostas — Bolão Copa 2026 · Turim MFO
======================================================
Execute este script uma vez (ou sempre que alguém enviar/atualizar uma
aposta) para gerar a planilha consolidada que o app lê:

    python consolidar_apostas.py

O arquivo gerado substitui a leitura individual das ~100 planilhas,
tornando o app substancialmente mais rápido.

Saída: apostas/Bolao_Copa2026_TurimMFO - Consolidada.xlsx
"""

import re, sys
from pathlib import Path
from openpyxl import load_workbook, Workbook

APOSTAS_DIR = Path(__file__).parent / "apostas"
OUTPUT_FILE = APOSTAS_DIR / "Bolao_Copa2026_TurimMFO - Consolidada.xlsx"

N_GRUPOS = 72   # jogos da fase de grupos
N_MM     = 32   # jogos do mata-mata


def consolidar():
    part_files = sorted(APOSTAS_DIR.glob("Bolao_Copa2026_TurimMFO_*.xlsx"))

    if not part_files:
        print(f"❌  Nenhum arquivo encontrado em: {APOSTAS_DIR}")
        sys.exit(1)

    total = len(part_files)
    print(f"📊  Consolidando {total} arquivo(s)...\n")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    ws_grp = wb_out.create_sheet("Apostas - Grupos")
    ws_mm  = wb_out.create_sheet("Apostas - Mata-Mata")
    ws_bon = wb_out.create_sheet("Apostas - Bonus")

    # Cabeçalhos
    ws_grp.append(
        ["Participante"] + [f"J{m+1}_{t}" for m in range(N_GRUPOS) for t in ("T1", "T2")]
    )
    ws_mm.append(
        ["Participante"] + [f"M{m+1}_{t}" for m in range(N_MM) for t in ("T1", "T2")]
    )
    ws_bon.append(["Participante", "Campeao", "Artilheiro"])

    erros = []

    for i, f in enumerate(part_files, 1):
        nm = re.sub(r"(?i)^Bolao_Copa2026_TurimMFO[_]?", "", f.name)
        nm = nm.replace(".xlsx", "")
        nm = re.sub(r"[{}\[\]()]", "", nm)   # remove { } [ ] ( )
        nm = nm.replace("_", " ").strip()
        nm = re.sub(r"\s+", " ", nm)         # colapsa espaços duplos

        try:
            wb = load_workbook(f, data_only=True, read_only=True)
            sh = wb.sheetnames
        except Exception as e:
            erros.append(f"{nm}: {e}")
            print(f"  ⚠️   [{i:>3}/{total}] {nm} — ERRO: {e}")
            continue

        # ── Fase de Grupos
        row_grp = [nm]
        if "Apostas - Grupos" in sh:
            ws = wb["Apostas - Grupos"]
            for m in range(N_GRUPOS):
                row_grp.append(ws.cell(row=5, column=2 + m * 2).value)
                row_grp.append(ws.cell(row=5, column=3 + m * 2).value)
        else:
            row_grp += [None] * (N_GRUPOS * 2)
        ws_grp.append(row_grp)

        # ── Mata-Mata
        row_mm = [nm]
        if "Apostas - Mata-Mata" in sh:
            ws = wb["Apostas - Mata-Mata"]
            for m in range(N_MM):
                row_mm.append(ws.cell(row=6, column=2 + m * 2).value)
                row_mm.append(ws.cell(row=6, column=3 + m * 2).value)
        else:
            row_mm += [None] * (N_MM * 2)
        ws_mm.append(row_mm)

        # ── Bônus
        row_bon = [nm]
        if "Apostas - Bonus" in sh:
            ws = wb["Apostas - Bonus"]
            row_bon.append(ws.cell(row=5, column=2).value)
            row_bon.append(ws.cell(row=5, column=3).value)
        else:
            row_bon += [None, None]
        ws_bon.append(row_bon)

        wb.close()
        print(f"  ✓   [{i:>3}/{total}] {nm}")

    wb_out.save(OUTPUT_FILE)

    print(f"\n{'=' * 55}")
    print(f"✅  Arquivo consolidado salvo:")
    print(f"    {OUTPUT_FILE}")
    print(f"    {total - len(erros)}/{total} participantes incluídos")
    if erros:
        print(f"\n⚠️   {len(erros)} erro(s):")
        for e in erros:
            print(f"    • {e}")


if __name__ == "__main__":
    consolidar()
